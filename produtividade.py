import time
import os
import glob
import win32com.client
import openpyxl
import urllib.parse
import pandas as pd
import re
import shutil
import unicodedata
import pyautogui
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, NoSuchWindowException, WebDriverException
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime, timedelta
from selenium.webdriver.support.ui import Select
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from selenium.webdriver.support.ui import WebDriverWait
from datetime import date
from openpyxl import load_workbook
from urllib.parse import quote as url_quote
from openpyxl.cell.cell import MergedCell
from calendar import monthrange
import pyperclip
import subprocess
from PIL import Image

import win32gui
import win32con
import win32api
import win32process
import ctypes


# ==============================================================================
# 🚀 O PULO DO GATO: FORÇAR O WINDOWS A RECONHECER TODOS OS MONITORES
# ==============================================================================
try:
    # Diz ao Windows para usar a resolução real de múltiplos monitores (Per-Monitor DPI Aware)
    ctypes.windll.shcore.SetProcessDpiAwareness(2)
except Exception:
    try:
        # Fallback para versões mais antigas do Windows
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass
# ==============================================================================



from config import EMAIL_MRV, SENHA_MRV, PASTA_DOWNLOADS, PASTA_PRODUTIVIDADE


# --- CONFIGURAÇÃO ---
WAIT_TIME = 10

# --- FUNÇÃO DE APOIO: LOGIN MICROSOFT ---
def fazer_login_microsoft(driver, wait, email, senha):
    print("--- Iniciando rotina de Login Microsoft ---")
    try:
        try:
            email_field = wait.until(EC.presence_of_element_located((By.ID, "i0116")))
            print("Preenchendo e-mail...")
            email_field.send_keys(email)
            wait.until(EC.element_to_be_clickable((By.ID, "idSIButton9"))).click()
            
            password_field = wait.until(EC.presence_of_element_located((By.ID, "i0118")))
            print("Preenchendo senha...")
            password_field.send_keys(senha)
            
            clicked = False
            for _ in range(3):
                try:
                    wait.until(EC.element_to_be_clickable((By.ID, "idSIButton9"))).click()
                    clicked = True
                    break
                except StaleElementReferenceException:
                    time.sleep(1)
            if not clicked: raise Exception("Não clicou em Entrar")
            time.sleep(5)
            print("!!! AGUARDANDO APROVAÇÃO MFA (Se necessário) !!!")
            wait.until(EC.element_to_be_clickable((By.ID, "idSIButton9"))).click() 
            print("Login Microsoft efetuado.")
            print("Aguardando janela pop-up fechar...")
            WebDriverWait(driver, 20).until(EC.number_of_windows_to_be(1))
            
            nova_janela_principal = driver.window_handles[0]
            driver.switch_to.window(nova_janela_principal)
            print("Foco retornado para a janela principal do Podio.")
        except TimeoutException:
            print("Campo de login não apareceu. Assumindo que já estamos logados (SSO).")
        return True
    except Exception as e:
        print(f"Erro no Login Microsoft: {e}")
        return False

# ==============================================================================
# FUNÇÃO 1: EXTRAÇÃO WEB E SAP
# ==============================================================================
def extrair_dados_sistemas():
    try:
        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, WAIT_TIME)

        # --- PARTE 1: PODIO ---
        print("\n=== INICIANDO PARTE 1: PODIO ===")
        driver.get("https://podio.com/login")
        try: wait.until(EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))).click()
        except: pass 
        wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@data-provider='live']"))).click()

        janela_principal = driver.current_window_handle
        wait.until(EC.number_of_windows_to_be(2))
        for handle in driver.window_handles:
            if handle != janela_principal:
                driver.switch_to.window(handle)
                break

        fazer_login_microsoft(driver, wait, EMAIL_MRV, SENHA_MRV)
        driver.switch_to.window(janela_principal)
        
        print("Navegando no Podio...")
        menu_area = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'space-switcher-wrapper')]")))
        ActionChains(driver).move_to_element(menu_area).perform()
        wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'ADM - Núcleo Contratos')]"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@data-app-id='22830484']"))).click()

        print("Aplicando filtros (Método Robusto)...")
        time.sleep(3)
        ul_filtros = wait.until(EC.presence_of_element_located((By.XPATH, "//ul[@class='app-filter-tools']")))
        itens_lista = ul_filtros.find_elements(By.TAG_NAME, "li")
        actions = ActionChains(driver)
        for item in itens_lista: actions.move_to_element(item)
        actions.perform()
        
        wait.until(EC.element_to_be_clickable((By.XPATH, ".//li[@data-original-title='Filtros']"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@data-id='created_on']"))).click() 
        wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@data-id='-1mr:-1mr']"))).click() 

        # --- ETAPA 8 ---

        try:
            # 1. Define o seletor.
            #  Usar CSS_SELECTOR é mais fácil para classes.
            #  O ponto (.) significa "classe".
            seletor_css = ".app-header__app-menu"
        
            print(f"Procurando todos os elementos com a classe: {seletor_css}")
            # 2. Espera até que PELO MENOS 2 elementos estejam presentes
            #  (Você pode mudar o '2' para quantos você espera)
            WebDriverWait(driver, 10).until(lambda d: len(d.find_elements(By.CSS_SELECTOR, seletor_css)) >= 2)

            # 3. Pega a LISTA de todos os elementos
            elementos = driver.find_elements(By.CSS_SELECTOR, seletor_css)
        
            print(f"Encontrados {len(elementos)} elementos.")

            # 4. Clica no primeiro elemento (índice 0)
            if len(elementos) > 0:
                print("Clicando no primeiro elemento (índice 0)...")
                elementos[0].click()
        
            # 5. Espera a página reagir
            #  (MUITO IMPORTANTE: Clicar em algo pode mudar a página)
            print("Aguardando 2 segundos para a página/menu reagir...")
            time.sleep(2)

            # 6. RE-ENCONTRA a lista de elementos
            #  (É a forma mais segura, caso o primeiro clique tenha
            #  recarregado os elementos - evita o erro 'stale element')
        
            print("Re-encontrando os elementos (para segurança)...")
            elementos = driver.find_elements(By.CSS_SELECTOR, seletor_css)

            # 7. Clica no segundo elemento (índice 1)
            if len(elementos) > 1:
                print("Clicando no segundo elemento (índice 1)...")
                elementos[1].click()
            else:
                print("Erro: Não foi possível encontrar o segundo elemento após o primeiro clique.")
        
            print("Ações nos dois elementos concluídas!")
            time.sleep(3)

        except Exception as e:
            print(f"Ocorreu um erro: {e}")
            # driver.save_screenshot("erro_multiplos.png")

        # --- ETAPA 9 ---
        print("Etapa 9: Aguardando o menu dropdown abrir...")
        
        # Usando o seletor CSS (mais limpo) que discutimos
        exportar_excel_selector = "a.app-box-supermenu-v2__link.app-export-excel"
        
        exportar_link = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, exportar_excel_selector))
        )
        
        print("Link 'Exportar Excel' encontrado. Clicando...")
        exportar_link.click()
        # --- FIM DA ETAPA 9 ---

        time.sleep(3)

        # --- ETAPA 10 ---
        try:
            print("Procurando o ícone de 'Notificação' (Inbox)...")
        
            # Usando o Seletor CSS (recomendado por ser mais limpo)
            notificacao_selector = "li.navigation-link.inbox"
        
            # Espera o ícone estar presente e ser clicável
            notificacao_icon = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, notificacao_selector))
            )
        
            print("Ícone de 'Notificação' encontrado. Clicando...")
            notificacao_icon.click()
        
            time.sleep(1) # Espera o menu de notificação abrir

        except Exception as e:
            print(f"Erro ao tentar clicar no ícone de Notificação: {e}")
            driver.save_screenshot("erro_notificacao.png")
        # --- FIM DA ETAPA 10 ---


        # --- ETAPA 11 ---
        css_corrigido = "a.PodioUI__Notifications__NotificationGroup"
        item_notificacao = WebDriverWait(driver, 3).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, css_corrigido))
        )
        item_notificacao.click()
        # --- FIM DA ETAPA 11 ---

        print("Aguardando processamento do Excel (até 3 minutos)...")

        tempo_espera = 0
        sucesso_exportacao = False
        nome_do_arquivo = "Mensageria - Última vista usada.xlsx"

        # LOOP DE VERIFICAÇÃO COM REFRESH: Tenta até dar 180s (3 minutos)
        while tempo_espera < 180:
            try:            
                if EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Mensageria - Última vista usada.xlsx")):

                    print("Exportação 'Completado'!")
 
                    link_download = WebDriverWait(driver, 3).until(
                        EC.element_to_be_clickable((By.LINK_TEXT, nome_do_arquivo))
                    )
                    
                    print("Link encontrado! Clicando para baixar...")
                    link_download.click()

                    sucesso_exportacao = True
                    break # Sai do loop pois deu certo!

                    
            except Exception:
                # Se não achou, espera um pouco e ATUALIZA A PÁGINA
                print(f"Aguardando Podio... ({tempo_espera}s / 180s) - Atualizando a página (F5)...")
                time.sleep(10) # Espera 10 segundos antes de atualizar
                tempo_espera += 10
                
                driver.refresh() # Dá o F5 na página
                time.sleep(5) # Espera a página recarregar
                tempo_espera += 5
                

            
        if not sucesso_exportacao:
            raise Exception("Tempo limite de 3 minutos excedido aguardando a exportação do Podio.")
            
        print("Download Podio iniciado!")
        time.sleep(1) 
        
        # --- PARTE 2: AGILIS ---
        print("\n=== INICIANDO PARTE 2: AGILIS ===")
        driver.get("https://agilis.mrv.com.br/HomePage.do?view_type=my_view")
        try:
            btn_login = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[text()='Login Integrado Microsoft']")))
            btn_login.click()
            fazer_login_microsoft(driver, wait, EMAIL_MRV, SENHA_MRV)
        except TimeoutException:
            print("Botão de login não apareceu, seguindo...")

        print("Navegando menus Agilis...")
        wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Relatórios"))).click()
        wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Contratos - ADM"))).click()
        wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Produtividade Contratos - ADM"))).click()
        wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "linkborder"))).click() 
        wait.until(EC.element_to_be_clickable((By.XPATH, "//option[text()='Coletor de custo ADM']"))).click()
        driver.find_element(By.CLASS_NAME, "moverightButton").click()

        try:
            expand_btn = wait.until(EC.presence_of_element_located((By.ID, "rcstep2src")))
            driver.execute_script("arguments[0].click();", expand_btn)
            time.sleep(1)
        except: pass

        print("Selecionando Data 'Mês Passado' no dropdown...")
        select_elem = wait.until(EC.presence_of_element_located((By.ID, "dateFilterType")))
        Select(select_elem).select_by_visible_text("Mês passado")
        
        print("Selecionando o rádio 'Durante' (Ajuste obrigatório)...")
        selector_radio_durante = (By.CSS_SELECTOR, "input[value='predefined']")
        wait.until(EC.element_to_be_clickable(selector_radio_durante)).click()

        wait.until(EC.element_to_be_clickable((By.ID, "addnew223222"))).click() 
        print("Relatório gerando. Aguardando 10 segundos...")
        time.sleep(10) 

        print("Iniciando o download direto do relatório XLS...")
        DOWNLOAD_XLS_LINK = (By.ID, "exportxls")
        wait.until(EC.element_to_be_clickable(DOWNLOAD_XLS_LINK)).click()
        time.sleep(5)
        print("Relatório Agilis baixado com sucesso!")

        # --- PARTE 3: BÚSSOLA MRV ---
        print("\n=== INICIANDO PARTE 3: BÚSSOLA MRV ===")
        driver.get("http://bussola.mrv.com.br/Main/Big.aspx")
        time.sleep(4)
        pyautogui.write(EMAIL_MRV.strip())
        pyautogui.press('tab')
        pyautogui.write(SENHA_MRV.strip())
        pyautogui.press('enter')
        time.sleep(3)

        driver.get("http://report2.mrv.com.br/ReportServer/Pages/ReportViewer.aspx?/BIG/Administrativo/ADM013%20-%20Relat%C3%B3rio%20Protocolo%20de%20Pagamento%20MRV%20PAG/REL_PRLPGTMRV&rs:Command=Render")
        time.sleep(4)
        pyautogui.write(EMAIL_MRV.strip())
        pyautogui.press('tab')
        pyautogui.write(SENHA_MRV.strip())
        pyautogui.press('enter')
        time.sleep(3)
        
        driver.get("http://bussola.mrv.com.br/Main/Big.aspx")
        pasta_adm = wait.until(EC.element_to_be_clickable((By.ID, "pasta2")))
        pasta_adm.click()
        time.sleep(2)

        xpath_relatorio = "//div[@id='divLinha' and contains(., 'Relatório Protocolo de Pagamento MRV PAG')]"
        relatorio_link = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_relatorio)))
        relatorio_link.click()
        
        WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)
        nova_janela = driver.window_handles[-1]
        driver.switch_to.window(nova_janela)
        wait = WebDriverWait(driver, 10)

        hoje = date.today()
        primeiro_dia_mes_atual = hoje.replace(day=1)
        ultimo_dia_mes_passado = primeiro_dia_mes_atual - timedelta(days=1)
        primeiro_dia_mes_passado = ultimo_dia_mes_passado.replace(day=1)
        data_inicio_str = primeiro_dia_mes_passado.strftime("%d/%m/%Y")
        data_final_str = ultimo_dia_mes_passado.strftime("%d/%m/%Y")

        time.sleep(3) 
        xpath_input_inicio = "(//button[@aria-label='Data criação inicio']/preceding::input)[last()]"
        input_inicio = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_input_inicio)))
        input_inicio.clear() 
        input_inicio.send_keys(data_inicio_str)

        xpath_input_final = "(//button[@aria-label='Data criação final']/preceding::input)[last()]"
        input_final = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_input_final)))
        input_final.clear()
        input_final.send_keys(data_final_str)

        dropdown_status = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl07_txtValue")))
        dropdown_status.click()
        time.sleep(1)
        checkbox_todos = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl07_divDropDown_ctl00")))
        checkbox_todos.click()
        time.sleep(1)

        btn_exibir = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl00")))
        btn_exibir.click()

        wait_longo = WebDriverWait(driver, 120)
        imagem_relatorio = wait_longo.until(EC.presence_of_element_located((By.XPATH, "//img[@alt='Imagem do relatório']")))
        time.sleep(2)

        btn_exportar = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl05_ctl04_ctl00_ButtonImgDown")))
        btn_exportar.click()
        time.sleep(1)

        btn_excel = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@title='Excel' or @alt='Excel']")))
        btn_excel.click()
        time.sleep(15) 
        print("Download do Bússola finalizado!")



        # --- MOVER ARQUIVOS ---
        time.sleep(5)
        if not os.path.exists(PASTA_PRODUTIVIDADE):
            os.makedirs(PASTA_PRODUTIVIDADE)

        files = [os.path.join(PASTA_DOWNLOADS, f) for f in os.listdir(PASTA_DOWNLOADS) if os.path.isfile(os.path.join(PASTA_DOWNLOADS, f))]
        files.sort(key=os.path.getmtime, reverse=True)
        top_4_files = files[:4]

        for file_path in top_4_files:
            file_name = os.path.basename(file_path)
            try:
                shutil.move(file_path, os.path.join(PASTA_PRODUTIVIDADE, file_name))
                print(f"Sucesso: {file_name} movido para {PASTA_PRODUTIVIDADE}")
            except Exception as e:
                print(f"Erro ao mover {file_name}: {e}")

        driver.quit()

    except Exception as e:
        print(f"Erro geral na extração: {e}")
        try: driver.quit() 
        except: pass
        raise e

# ==============================================================================
# FUNÇÕES AUXILIARES E DE PROCESSAMENTO (AGORA FORA DA EXTRAÇÃO)
# ==============================================================================

def find_column_ignore_case(df, column_name):
    for col in df.columns:
        if col.lower() == column_name.lower(): return col
    return None

def processar_mensageria(filepath, new_filename):
    try:
        df = pd.read_excel(filepath, header=0) 
        df.columns = [str(col).strip() for col in df.columns]
        coluna_usuario = find_column_ignore_case(df, 'Criado por')
        coluna_data = find_column_ignore_case(df, 'Criado em')
        coluna_valores = find_column_ignore_case(df, 'Numero do chamado Agilis/Rastreio')
        df[coluna_data] = pd.to_datetime(df[coluna_data], dayfirst=True).dt.date
        pivot_table = pd.pivot_table(df, index=coluna_usuario, columns=coluna_data, values=coluna_valores, aggfunc='count', fill_value=0, margins=True, margins_name='Total Geral')
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            pivot_table.to_excel(writer, sheet_name='TabelaDinamica')
        os.rename(filepath, new_filename)
    except Exception as e: print(f"ERRO Mensageria: {e}")

def processar_produtividade(filepath, new_filename):
    try:
        df = pd.read_excel(filepath, header=0)
        df.columns = [str(col).strip() for col in df.columns]
        coluna_usuario = find_column_ignore_case(df, 'Nome do usuário')
        coluna_data = find_column_ignore_case(df, 'Data de lançamento')
        coluna_valores = find_column_ignore_case(df, 'Nº doc.faturamento')
        df[coluna_data] = pd.to_datetime(df[coluna_data], dayfirst=True).dt.date
        pivot_table = pd.pivot_table(df, index=coluna_usuario, columns=coluna_data, values=coluna_valores, aggfunc='count', fill_value=0, margins=True, margins_name='Total Geral')
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            pivot_table.to_excel(writer, sheet_name='TabelaDinamica')
        os.rename(filepath, new_filename)
    except Exception as e: print(f"ERRO Produtividade: {e}")

def processar_numerico(filepath, new_filename):
    try:
        df = pd.read_excel(filepath, header=8)
        df.columns = [str(col).strip() for col in df.columns]
        coluna_tecnico = find_column_ignore_case(df, 'Técnico')
        coluna_data = find_column_ignore_case(df, 'Hora de conclusão')
        coluna_valores = find_column_ignore_case(df, 'Identificação da solicitação')
        df[coluna_data] = pd.to_datetime(df[coluna_data], dayfirst=True).dt.date
        pivot_table = pd.pivot_table(df, index=coluna_tecnico, columns=coluna_data, values=coluna_valores, aggfunc='count', fill_value=0, margins=True, margins_name='Total Geral')
        with pd.ExcelWriter(new_filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='DadosOriginais', index=False)
            pivot_table.to_excel(writer, sheet_name='TabelaDinamica')
        os.remove(filepath)
    except Exception as e: print(f"ERRO Numérico: {e}")

def processar_relatorio_pedidos(filepath, new_filename):
    try:
        df = pd.read_excel(filepath, header=1)
        df.columns = [str(col).strip() for col in df.columns]
        coluna_linhas = find_column_ignore_case(df, 'Respons. Entrega')
        coluna_colunas = find_column_ignore_case(df, 'Data Entrada NF')
        coluna_valores = find_column_ignore_case(df, 'Nro. Pedido Compra')
        df[coluna_colunas] = pd.to_datetime(df[coluna_colunas], dayfirst=True).dt.date
        pivot_table = pd.pivot_table(df, index=coluna_linhas, columns=coluna_colunas, values=coluna_valores, aggfunc='count', fill_value=0, margins=True, margins_name='Total Geral')
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            pivot_table.to_excel(writer, sheet_name='TabelaDinamica')
        os.rename(filepath, new_filename)
    except Exception as e: print(f"ERRO Pedidos: {e}")

def step_1_prepare_and_rename_reports(diretorio):
    arquivos = glob.glob(os.path.join(diretorio, '*.*'))
    for arquivo in arquivos:
        nome_arquivo = os.path.basename(arquivo)
        if nome_arquivo.startswith('Mensageria - Última vista'):
            processar_mensageria(arquivo, os.path.join(diretorio, 'Relatório - Sedex.Malote.xlsx'))
        elif nome_arquivo.startswith('export') or nome_arquivo.startswith('EXPORT') and nome_arquivo.endswith('.xlsx'):
            processar_produtividade(arquivo, os.path.join(diretorio, 'Relatório - SAP.xlsx'))
        elif nome_arquivo.startswith('REL_PRLPGT'):
            processar_relatorio_pedidos(arquivo, os.path.join(diretorio, 'Relatório - Lançamentos.xlsx'))
        elif re.match(r'^\d+\.(xlsx|xls)$', nome_arquivo):
            processar_numerico(arquivo, os.path.join(diretorio, 'Relatório - Agilis.xlsx'))

# --- FUNÇÕES DE PREENCHIMENTO (Resumidas para economizar espaço, mas mantendo a lógica) ---
def norm_key(s):
    if s is None: return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.lower()

def col_letter_to_index(letter: str) -> int:
    letter = letter.strip().upper()
    n = 0
    for ch in letter:
        if not ('A' <= ch <= 'Z'): continue
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n

def date_keys(v):
    keys = []
    if isinstance(v, datetime):
        keys += [v.strftime("%Y-%m-%d 00:00:00"), v.strftime("%d/%m/%Y")]
    elif isinstance(v, str):
        s = v.strip()
        if re.match(r"^\d{4}-\d{2}-\d{2} 00:00:00$", s): keys.append(s)
        if re.match(r"^\d{2}/\d{2}/\d{4}$", s): keys.append(s)
        try:
            dt = pd.to_datetime(s, errors="raise")
            keys += [dt.strftime("%Y-%m-%d 00:00:00"), dt.strftime("%d/%m/%Y")]
        except: pass
    return list(dict.fromkeys(keys))

def build_header_map(ws):
    hdr = {}
    for c in range(1, ws.max_column+1):
        v = ws.cell(row=1, column=c).value
        if v is None: continue
        if isinstance(v, datetime):
            hdr[v.strftime("%Y-%m-%d 00:00:00")] = c
            hdr[v.strftime("%d/%m/%Y")] = c
        else:
            s = str(v).strip()
            hdr[s] = c
            for k in date_keys(s): hdr[k] = c
    return hdr

def extract_user_key(s: str) -> str:
    if s is None: return ""
    s = str(s).strip().split('@')[0]
    s_norm = unicodedata.normalize("NFKD", s)
    s_norm = "".join(ch for ch in s_norm if not unicodedata.combining(ch))
    tokens = s_norm.strip().split()
    for t in tokens:
        if '.' in t: return re.sub(r'[^A-Za-z\.]', '', t).lower()
    if len(tokens) >= 2:
        return f"{re.sub(r'[^A-Za-z]', '', tokens[0])}.{re.sub(r'[^A-Za-z]', '', tokens[-1])}".lower()
    return re.sub(r'[^A-Za-z]', '', tokens[0]).lower() if tokens else ""

def read_tabledinamica_with_namecol(path, name_col_hint=None):
    df = pd.read_excel(path, sheet_name="TabelaDinamica", engine="openpyxl")
    nome_col = None
    if name_col_hint:
        hint_norm = norm_key(name_col_hint)
        for c in df.columns:
            if norm_key(c) == hint_norm: nome_col = c; break
    if not nome_col:
        heads = {norm_key(c): c for c in df.columns}
        for alvo in ["criado por", "tecnico", "técnico", "respons. entrega", "nome do usuário"]:
            if alvo in heads: nome_col = heads[alvo]; break
    if not nome_col:
        for c in df.columns:
            if df[c].dtype == "O": nome_col = c; break

    day_cols = [c for c in df.columns if c != nome_col and (isinstance(c, datetime) or (isinstance(c, str) and (re.match(r"^\d{2}/\d{2}/\d{4}$", c) or re.match(r"^\d{4}-\d{2}-\d{2}", c))))]
    registros = []
    if nome_col:
        for _, row in df.iterrows():
            nome_val = str(row.get(nome_col, "")).strip()
            if not nome_val or norm_key(nome_val).startswith(norm_key("Total Geral")): continue
            for d in day_cols:
                val = row.get(d, 0)
                try: v = int(val) if pd.notna(val) else 0
                except: v = 0
                registros.append({"nome": nome_val, "data_obj": d, "valor": v})
    return pd.DataFrame(registros, columns=["nome","data_obj","valor"])

def read_lanctos_tabledinamica(path):
    df = pd.read_excel(path, sheet_name="TabelaDinamica", engine="openpyxl")
    possible_names = ["Respons. Entrega", "Técnico", "Tecnico", "Criado por", "Nome do usuário", "User"]
    nome_col = None
    cols_map = {norm_key(c): c for c in df.columns}
    for alvo in possible_names:
        if norm_key(alvo) in cols_map: nome_col = cols_map[norm_key(alvo)]; break
    if nome_col is None: nome_col = df.columns[0]

    day_cols = [c for c in df.columns if c != nome_col and (isinstance(c, datetime) or (isinstance(c, str) and (re.match(r"^\d{2}/\d{2}/\d{4}$", c) or re.match(r"^\d{4}-\d{2}-\d{2}", c))))]
    registros = []
    for _, row in df.iterrows():
        raw_name = str(row.get(nome_col, "")).strip()
        if not raw_name or norm_key(raw_name).startswith(norm_key("Total Geral")): continue
        ukey = extract_user_key(raw_name)
        for d in day_cols:
            val = row.get(d, 0)
            try: v = int(val) if pd.notna(val) else 0
            except: v = 0
            registros.append({"user_key": ukey, "data_obj": d, "valor": v})
    return pd.DataFrame(registros, columns=["user_key","data_obj","valor"])

def update_headers_to_previous_month(ws, header_row=1, start_col_letter="D", end_col_letter="AH", ref_date=None):
    start_col = col_letter_to_index(start_col_letter)
    end_col   = col_letter_to_index(end_col_letter)
    if ref_date is None: ref_date = date.today()
    first_of_month = date(ref_date.year, ref_date.month, 1)
    last_day_prev  = first_of_month - timedelta(days=1)
    ano, mes = last_day_prev.year, last_day_prev.month
    qtd_dias = monthrange(ano, mes)[1]

    for i in range(qtd_dias):
        ws.cell(row=header_row, column=start_col + i, value=datetime(ano, mes, i + 1))
    for c in range(start_col + qtd_dias, end_col + 1):
        ws.cell(row=header_row, column=c, value=None)
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=header_row, column=c)
        if isinstance(cell.value, datetime): cell.number_format = "dd/mm/yyyy"
    return ano, mes, qtd_dias

def clear_month_data_in_blocks(ws, row_ranges, start_col_letter="D", end_col_letter="AH"):
    start_col = col_letter_to_index(start_col_letter)
    end_col   = col_letter_to_index(end_col_letter)
    cleared = 0
    for (r0, r1) in row_ranges:
        for r in range(r0, r1 + 1):
            if r > ws.max_row: continue
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell): continue
                if cell.value not in (None, ""):
                    cell.value = None
                    cleared += 1
    return cleared

def fill_agilis_same_row(ws, header_map, df_long, AGILIS_POS):
    if df_long.empty: return 0
    grp_exact = {norm_key(n): sub for n, sub in df_long.groupby(df_long['nome'].apply(norm_key))}
    grp_ukey  = {extract_user_key(n): sub for n, sub in df_long.groupby(df_long['nome'].apply(extract_user_key))}
    total_writes = 0
    for item in AGILIS_POS:
        p2, p1, row_agilis, min_idx = item["p2"], item["p1"], item["row_nome"], col_letter_to_index(item["min_col_letter"])
        if min_idx > ws.max_column: min_idx = 1
        sub = grp_exact.get(norm_key(p2))
        if (sub is None) or sub.empty: sub = grp_ukey.get(extract_user_key(p2))
        if sub is None or sub.empty: continue
        writes = 0
        for _, reg in sub.iterrows():
            col_idx = next((header_map.get(k) for k in date_keys(reg["data_obj"]) if header_map.get(k)), None)
            if not col_idx or col_idx < min_idx: continue
            val = int(reg["valor"])
            if val != 0:
                ws.cell(row=row_agilis, column=col_idx, value=val)
                writes += 1
        total_writes += writes
    return total_writes

def fill_sedex(ws, header_map, df_long, MAP_SEDEX):
    if df_long.empty: return 0
    grp = {norm_key(n): sub for n, sub in df_long.groupby(df_long['nome'].apply(norm_key))}
    total_writes = 0
    for p2, p1 in MAP_SEDEX.items():
        r_nome = next((r for r in range(2, ws.max_row+1) if isinstance(ws.cell(row=r, column=2).value, str) and norm_key(ws.cell(row=r, column=2).value) == norm_key(p1)), None)
        if r_nome is None: continue
        r_sedex = next((rr for rr in range(r_nome, min(ws.max_row, r_nome+12)+1) if isinstance(ws.cell(row=rr, column=3).value, str) and 'sedex/pac/malote' in ws.cell(row=rr, column=3).value.strip().lower()), None)
        if r_sedex is None: continue
        sub = grp.get(norm_key(p2))
        if sub is None or sub.empty: continue
        writes = 0
        for _, reg in sub.iterrows():
            col_idx = next((header_map.get(k) for k in date_keys(reg["data_obj"]) if header_map.get(k)), None)
            if not col_idx: continue
            val = int(reg["valor"])
            if val != 0:
                ws.cell(row=r_sedex, column=col_idx, value=val)
                writes += 1
        total_writes += writes
    return total_writes

def fill_lanctos_fixed(ws, header_map, df_long, LANCTOS_USER_MAP):
    if df_long.empty or "user_key" not in df_long.columns: return 0
    grp = {uk: sub for uk, sub in df_long.groupby(df_long['user_key'])}
    total_writes = 0
    for ukey, meta in LANCTOS_USER_MAP.items():
        row_ativ = meta["row_ativ"]
        sub = grp.get(ukey)
        if sub is None or sub.empty: continue
        writes = 0
        for _, reg in sub.iterrows():
            col_idx = next((header_map.get(k) for k in date_keys(reg["data_obj"]) if header_map.get(k)), None)
            if not col_idx: continue
            val = int(reg["valor"])
            if val != 0:
                ws.cell(row=row_ativ, column=col_idx, value=val)
                writes += 1
        total_writes += writes
    return total_writes

def fill_sap_fixed(ws, header_map, df_long, SAP_COD_MAP):
    if df_long.empty or "nome" not in df_long.columns: return 0
    cod_grp = {cod: df_long[df_long['nome'].str.contains(cod, case=False, regex=True, na=False)] for cod in SAP_COD_MAP.keys()}
    total_writes = 0
    for cod, meta in SAP_COD_MAP.items():
        row_ativ = meta["row_ativ"]
        sub = cod_grp.get(cod)
        if sub is None or sub.empty: continue
        writes = 0
        for _, reg in sub.iterrows():
            col_idx = next((header_map.get(k) for k in date_keys(reg["data_obj"]) if header_map.get(k)), None)
            if not col_idx: continue
            val = int(reg["valor"])
            if val != 0:
                ws.cell(row=row_ativ, column=col_idx, value=val)
                writes += 1
        total_writes += writes
    return total_writes

def fill_fsf_flags(ws, header_map):
    row_ranges = [(2, 5), (7, 10), (12, 15), (17, 20), (22, 25), (27, 30), (32, 35), (37, 40), (42, 45), (47, 50), (52,52)]
    cols_to_process = {}
    for date_str, col_idx in header_map.items():
        try:
            dt = pd.to_datetime(date_str, errors='coerce') if isinstance(date_str, str) and "-" in date_str and date_str.index("-") == 4 else pd.to_datetime(date_str, dayfirst=True, errors='coerce')
            if pd.notna(dt): cols_to_process[col_idx] = dt.date()
        except: continue

    total_writes = 0
    for col_idx, data_atual in cols_to_process.items():
        is_weekend = (data_atual.weekday() >= 5)
        dia_teve_producao = any(ws.cell(row=r, column=col_idx).value not in (None, "", 0, "0") for start_row, end_row in row_ranges for r in range(start_row, end_row + 1) if r <= ws.max_row)
        if is_weekend or not dia_teve_producao:
            for start_row, end_row in row_ranges:
                if not any(ws.cell(row=r, column=col_idx).value not in (None, "", 0, "0") for r in range(start_row, end_row + 1) if r <= ws.max_row):
                    for r in range(start_row, end_row + 1):
                        if r <= ws.max_row and not isinstance(ws.cell(row=r, column=col_idx), MergedCell) and ws.cell(row=r, column=col_idx).value in (None, "", 0):
                            ws.cell(row=r, column=col_idx).value = "0"
                            total_writes += 1
    return total_writes

# ==============================================================================
# FUNÇÃO PRINCIPAL DE PROCESSAMENTO DO EXCEL
# ==============================================================================
def main():
    # ATENÇÃO: Agora os caminhos apontam para a PASTA_PRODUTIVIDADE
    PROD_PATH    = os.path.join(PASTA_PRODUTIVIDADE, "Produtividade 06 - 2026 (preenchido).xlsx")
    AGILIS_PATH  = os.path.join(PASTA_PRODUTIVIDADE, "Relatório - Agilis.xlsx")
    SEDEX_PATH   = os.path.join(PASTA_PRODUTIVIDADE, "Relatório - Sedex.Malote.xlsx")
    LANCTOS_PATH = os.path.join(PASTA_PRODUTIVIDADE, "Relatório - Lançamentos.xlsx")
    SAP_PATH     = os.path.join(PASTA_PRODUTIVIDADE, "Relatório - SAP.xlsx")
    OUT_PATH     = os.path.join(PASTA_PRODUTIVIDADE, "Produtividade 07 - 2026 (preenchido).xlsx")

    MAP_SEDEX = {
        "Alfredo Henrique Goncalves Pereira": "Alfredo.pereira MS0069532",
        "Gabriel Figueiredo Emiliano":        "gabriel.emiliano MS0073186",
        "Pedro Henrique Soares Silva":        "pedro.henrsilva MS0073814",
        "Joao Vitor Barbosa Fernandes":       "joao.vifernandes",
        "Matheus Silva De Lemos":             "matheus.lemos.silva MS0075116",
    }

    AGILIS_POS = [
        {"p2": "Alfredo Henrique Goncalves Pereira", "p1": "Alfredo.pereira MS0069532",  "row_nome":  2, "min_col_letter": "CO"},
        {"p2": "Gabriel Figueiredo Emiliano",        "p1": "gabriel.emiliano MS0073186", "row_nome":  7, "min_col_letter": "CO"},
        {"p2": "Ellen Gabrielle De Morais Gomes Da Silva", "p1": "ellen.morais",          "row_nome": 12, "min_col_letter": "CO"},
        {"p2": "maria.delgado",                       "p1": "maria.delgado",               "row_nome": 17, "min_col_letter": "CO"},
        {"p2": "Pedro Henrique Soares Silva",        "p1": "pedro.henrsilva MS0073814",  "row_nome": 22, "min_col_letter": "CO"},
        {"p2": "Pedro Henrique Marques",             "p1": "pedro.hmarques",             "row_nome": 27, "min_col_letter": "CO"},
        {"p2": "Carolina Pagnozzi Silva",            "p1": "pagnozzi.carolina",          "row_nome": 32, "min_col_letter": "CO"},
        {"p2": "maria.eduarocha",                    "p1": "maria.eduarocha",             "row_nome": 37, "min_col_letter": "CO"},
        {"p2": "Matheus Silva De Lemos",             "p1": "matheus.lemos.silva",        "row_nome": 42, "min_col_letter": "CO"},
        {"p2": "Joao Vitor Barbosa Fernandes",       "p1": "joao.vifernandes",           "row_nome": 47, "min_col_letter": "CO"},
        {"p2": "Vanessa De Brito Rodrigues",         "p1": "Vanessa",                    "row_nome": 52, "min_col_letter": "C"},
    ]

    LANCTOS_USER_MAP = {
        "alfredo.pereira":      {"p1": "Alfredo.pereira MS0069532",  "row_ativ":  4},
        "gabriel.emiliano":     {"p1": "gabriel.emiliano MS0073186", "row_ativ":  9},
        "ellen.morais":         {"p1": "ellen.morais",               "row_ativ": 14},
        "maria.delgado":        {"p1": "maria.delgado",              "row_ativ": 19},
        "pedro.henrsilva":      {"p1": "pedro.henrsilva MS0073814",  "row_ativ": 24},
        "pedro.hmarques":       {"p1": "pedro.hmarques",             "row_ativ": 29},
        "pagnozzi.carolina":    {"p1": "pagnozzi.carolina",          "row_ativ": 34},
        "maria.eduarocha":      {"p1": "maria.eduarocha",            "row_ativ": 39},
        "matheus.lemos.silva":  {"p1": "matheus.lemos.silva MS0075116", "row_ativ": 44},
    }

    SAP_COD_MAP = {
        "MS0069532": {"p1": "Alfredo.pereira MS0069532",  "row_ativ":  5},
        "MS0073186": {"p1": "gabriel.emiliano MS0073186", "row_ativ": 10},
        "MS0073814": {"p1": "pedro.henrsilva MS0073814",  "row_ativ": 25},
        "MS0075116": {"p1": "matheus.lemos.silva MS0075116", "row_ativ": 45},
    }

    wb = load_workbook(PROD_PATH)
    ws = wb["Plan1"]

    ano, mes, qtd_dias = update_headers_to_previous_month(ws, header_row=1, start_col_letter="D", end_col_letter="AH")
    
    ROW_RANGES_ATIV = [(2, 5), (7, 10), (12, 15), (17, 20), (22, 25), (27, 30), (32, 35), (37, 40), (42, 45), (47, 50), (52,52)]
    clear_month_data_in_blocks(ws, ROW_RANGES_ATIV, start_col_letter="D", end_col_letter="AH")
    
    header_map = build_header_map(ws)

    df_ag  = read_tabledinamica_with_namecol(AGILIS_PATH)
    df_sd  = read_tabledinamica_with_namecol(SEDEX_PATH)
    df_lan = read_lanctos_tabledinamica(LANCTOS_PATH)
    df_sap = read_tabledinamica_with_namecol(SAP_PATH, name_col_hint="Nome do usuário")

    fill_agilis_same_row(ws, header_map, df_ag, AGILIS_POS)
    fill_sedex(ws, header_map, df_sd, MAP_SEDEX)
    fill_lanctos_fixed(ws, header_map, df_lan, LANCTOS_USER_MAP)
    fill_sap_fixed(ws, header_map, df_sap, SAP_COD_MAP)
    fill_fsf_flags(ws, header_map)

    wb.save(OUT_PATH)
    print(f"✅ Planilha salva com sucesso em: {OUT_PATH}")

# ==============================================================================
# FUNÇÃO MESTRE (CHAMADA PELO HUB CENTRAL)
# ============================================================================
def executar_robo_produtividade_setor():
    print("Iniciando Robô de Produtividade...")
    
    # 1. Extrai tudo da Web e do SAP e move para a pasta
    extrair_dados_sistemas()
    
    # 2. Renomeia e prepara os relatórios (apontando para a pasta certa)
    print("--- Executando Etapa 1: Renomear Arquivos ---")
    step_1_prepare_and_rename_reports(PASTA_PRODUTIVIDADE)
    
    # 3. Processa o Excel
    print("--- Executando Etapa 2: Processar Produtividade ---")
    main()
    
    print("✅ Robô de Produtividade finalizado com sucesso!")

if __name__ == "__main__":
    executar_robo_produtividade_setor()
