from pathlib import Path
import os
import re
import sys
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side,
)

import config


# ==============================================================================
# CONFIGURACAO DE PASTAS
# ==============================================================================

PASTA = (
    Path(config.PASTA_ARQUIVOS)
    / "Responsaveis Viagens"
)

PASTA.mkdir(
    parents=True,
    exist_ok=True
)

PADRAO_EXPORT = "export"
PADRAO_BASE_ATIVOS = "base de ativos"
PADRAO_VIAGENS = "centro de custo viagens"

NOME_ARQUIVO_SAIDA = "Centro_Custo_Viagens_Atualizado.xlsx"

COLUNA_CC_VIAGENS = 2
COLUNA_APROVADOR_1 = 5
COLUNA_APROVADOR_2 = 6


# ==============================================================================
# FUNCOES DE NORMALIZACAO
# ==============================================================================

def norm(valor):
    """
    Normaliza textos para comparacao.

    Exemplos:
        Joao da Silva -> JOAO DA SILVA
        JOAO  DA SILVA -> JOAO DA SILVA
        João da Silva -> JOAO DA SILVA
    """

    if valor is None:
        return ""

    try:
        if pd.isna(valor):
            return ""
    except (TypeError, ValueError):
        pass

    texto = str(valor).strip()

    if not texto:
        return ""

    texto = unicodedata.normalize("NFKD", texto)

    texto = "".join(
        caractere
        for caractere in texto
        if not unicodedata.combining(caractere)
    )

    texto = texto.upper()
    texto = re.sub(r"\s+", " ", texto)

    return texto.strip()


def normalizar_codigo_cc(valor):
    """
    Normaliza o codigo do centro de custo.

    Remove espacos, caracteres invisiveis e converte para maiusculas.
    """

    texto = norm(valor)

    texto = texto.replace("\u00A0", "")
    texto = texto.replace(" ", "")

    # Corrige codigos que eventualmente tenham sido lidos como numero.
    if texto.endswith(".0"):
        texto = texto[:-2]

    return texto


# ==============================================================================
# COMPARACAO DE NOMES
# ==============================================================================

def is_word_match(palavra_1, palavra_2):
    """
    Compara duas partes de nomes.

    Permite identificar nomes abreviados, por exemplo:
        SANDRA MOURA
        SANDRA RIBEIRO DE MOURA
    """

    if palavra_1 == palavra_2:
        return True

    if (
        len(palavra_1) >= 3
        and len(palavra_2) >= 3
        and (
            palavra_1.startswith(palavra_2)
            or palavra_2.startswith(palavra_1)
        )
    ):
        return True

    if len(palavra_1) <= 2 and palavra_2.startswith(palavra_1):
        return True

    if len(palavra_2) <= 2 and palavra_1.startswith(palavra_2):
        return True

    return False


def names_match(nome_1, nome_2):
    """
    Compara dois nomes, permitindo abreviacoes.

    O primeiro nome obrigatoriamente deve ser compativel e todas as partes
    do nome menor devem existir no nome maior.
    """

    partes_1 = norm(nome_1).split()
    partes_2 = norm(nome_2).split()

    if not partes_1 or not partes_2:
        return False

    if not is_word_match(partes_1[0], partes_2[0]):
        return False

    if len(partes_1) <= len(partes_2):
        nome_menor = partes_1
        nome_maior = partes_2
    else:
        nome_menor = partes_2
        nome_maior = partes_1

    correspondencias = 0

    for parte_menor in nome_menor:
        for parte_maior in nome_maior:
            if is_word_match(parte_menor, parte_maior):
                correspondencias += 1
                break

    return correspondencias == len(nome_menor)


def get_cargo_weight(cargo):
    """
    Retorna o peso hierárquico do cargo.

    Essa pontuação é utilizada somente quando mais de uma pessoa
    da Base de Ativos corresponde ao mesmo nome informado pelo SAP.

    Quanto maior o peso, maior a prioridade.
    """

    cargo_normalizado = norm(cargo)

    if not cargo_normalizado:
        return 0

    # Alta liderança
    if "PRESIDENTE" in cargo_normalizado:
        return 1000

    if "CONSELHEIRO" in cargo_normalizado:
        return 950

    if "VICE PRESIDENTE" in cargo_normalizado:
        return 900

    if "VICE-PRESIDENTE" in cargo_normalizado:
        return 900

    if "DIRETOR EXECUTIVO" in cargo_normalizado:
        return 850

    if "DIRETOR" in cargo_normalizado:
        return 800

    if "SUPERINTENDENTE" in cargo_normalizado:
        return 750

    # Liderança gerencial
    if "GESTOR EXECUTIVO" in cargo_normalizado:
        return 700

    if "GERENTE EXECUTIVO" in cargo_normalizado:
        return 680

    if "GESTOR" in cargo_normalizado:
        return 650

    if "GERENTE" in cargo_normalizado:
        return 600

    if "COORDENADOR" in cargo_normalizado:
        return 500

    if "SUPERVISOR" in cargo_normalizado:
        return 400

    # Cargos técnicos e especializados
    if "ESPECIALISTA" in cargo_normalizado:
        return 350

    if "ENGENHEIRO" in cargo_normalizado:
        return 320

    if "ARQUITETO" in cargo_normalizado:
        return 310

    if "COMPRADOR" in cargo_normalizado:
        return 300

    if "ANALISTA" in cargo_normalizado:
        return 250

    if "TECNICO" in cargo_normalizado:
        return 200

    if "ASSISTENTE" in cargo_normalizado:
        return 150

    if "AUXILIAR" in cargo_normalizado:
        return 100

    if "ESTAGIARIO" in cargo_normalizado:
        return 50

    if "APRENDIZ" in cargo_normalizado:
        return 25

    return 10

# ==============================================================================
# LOCALIZACAO DOS ARQUIVOS
# ==============================================================================

def localizar_arquivo(palavra_chave, ignorar_nomes=None):
    """
    Localiza o arquivo Excel mais recente que contenha a palavra-chave.

    A busca nao diferencia maiusculas, minusculas ou acentos.
    """

    if ignorar_nomes is None:
        ignorar_nomes = []

    ignorar_normalizados = {
        norm(nome)
        for nome in ignorar_nomes
    }

    arquivos_encontrados = []

    for caminho in PASTA.iterdir():
        if not caminho.is_file():
            continue

        if caminho.suffix.lower() not in {".xls", ".xlsx", ".xlsm"}:
            continue

        # Ignora arquivos temporarios criados pelo Excel.
        if caminho.name.startswith("~$"):
            continue

        if norm(caminho.name) in ignorar_normalizados:
            continue

        if norm(palavra_chave) in norm(caminho.name):
            arquivos_encontrados.append(caminho)

    if not arquivos_encontrados:
        return None

    return max(
        arquivos_encontrados,
        key=lambda arquivo: arquivo.stat().st_mtime
    )


def validar_arquivos():
    """
    Localiza e valida os tres arquivos necessarios.
    """

    arquivo_saida = PASTA / NOME_ARQUIVO_SAIDA

    export_file = localizar_arquivo(PADRAO_EXPORT)

    ativos_file = localizar_arquivo(PADRAO_BASE_ATIVOS)

    viagens_file = localizar_arquivo(
        PADRAO_VIAGENS,
        ignorar_nomes=[NOME_ARQUIVO_SAIDA]
    )

    erros = []

    if export_file is None:
        erros.append(
            "- EXPORT: arquivo contendo 'export' no nome."
        )

    if ativos_file is None:
        erros.append(
            "- Base de Ativos: arquivo contendo 'Base de Ativos' no nome."
        )

    if viagens_file is None:
        erros.append(
            "- Centro de Custo Viagens: arquivo contendo "
            "'centro de custo viagens' no nome."
        )

    if erros:
        mensagem = (
            "Nao foi possivel iniciar o processo.\n\n"
            "Arquivos nao encontrados:\n"
            + "\n".join(erros)
            + f"\n\nTodos os arquivos devem estar na pasta:\n{PASTA}"
        )

        raise FileNotFoundError(mensagem)

    # Impede que o proprio arquivo de saida seja utilizado como entrada.
    if viagens_file.resolve() == arquivo_saida.resolve():
        raise RuntimeError(
            "O arquivo de entrada nao pode ser o mesmo arquivo de saida."
        )

    return export_file, ativos_file, viagens_file, arquivo_saida


# ==============================================================================
# FUNCOES PARA LOCALIZAR CABECALHOS
# ==============================================================================

def encontrar_cabecalho_dataframe(
    arquivo,
    termos_obrigatorios,
    limite_linhas=100
):
    """
    Localiza a linha de cabecalho de uma planilha usando termos obrigatorios.
    """

    dataframe_bruto = pd.read_excel(
        arquivo,
        header=None
    )

    termos_normalizados = [
        norm(termo)
        for termo in termos_obrigatorios
    ]

    quantidade_linhas = min(
        limite_linhas,
        len(dataframe_bruto)
    )

    for indice in range(quantidade_linhas):
        valores_linha = [
            norm(valor)
            for valor in dataframe_bruto.iloc[indice].tolist()
        ]

        texto_linha = " | ".join(valores_linha)

        if all(
            termo in texto_linha
            for termo in termos_normalizados
        ):
            return indice

    raise RuntimeError(
        f"Nao foi possivel localizar o cabecalho do arquivo "
        f"'{arquivo.name}'."
    )


def encontrar_cabecalho_worksheet(
    worksheet,
    termos_obrigatorios,
    limite_linhas=100,
    limite_colunas=50
):
    """
    Localiza a linha de cabecalho em uma planilha aberta pelo openpyxl.
    """

    termos_normalizados = [
        norm(termo)
        for termo in termos_obrigatorios
    ]

    ultima_linha = min(
        limite_linhas,
        worksheet.max_row
    )

    ultima_coluna = min(
        limite_colunas,
        worksheet.max_column
    )

    for numero_linha in range(1, ultima_linha + 1):
        valores = [
            norm(
                worksheet.cell(
                    row=numero_linha,
                    column=numero_coluna
                ).value
            )
            for numero_coluna in range(1, ultima_coluna + 1)
        ]

        if all(
            termo in valores
            for termo in termos_normalizados
        ):
            return numero_linha, valores

    raise RuntimeError(
        "Nao foi possivel localizar o cabecalho da planilha."
    )


def localizar_coluna_por_nomes(colunas, nomes_possiveis):
    """
    Localiza uma coluna aceitando variacoes de nome.
    """

    mapa_colunas = {
        norm(coluna): coluna
        for coluna in colunas
    }

    for nome in nomes_possiveis:
        nome_normalizado = norm(nome)

        if nome_normalizado in mapa_colunas:
            return mapa_colunas[nome_normalizado]

    for coluna_normalizada, coluna_original in mapa_colunas.items():
        for nome in nomes_possiveis:
            nome_normalizado = norm(nome)

            if nome_normalizado in coluna_normalizada:
                return coluna_original

    return None


def localizar_indice_coluna_cabecalho(
    valores_cabecalho,
    nomes_possiveis
):
    """
    Localiza o numero da coluna em um cabecalho do openpyxl.
    """

    nomes_normalizados = {
        norm(nome)
        for nome in nomes_possiveis
    }

    for indice, valor in enumerate(valores_cabecalho, start=1):
        if valor in nomes_normalizados:
            return indice

    for indice, valor in enumerate(valores_cabecalho, start=1):
        for nome in nomes_normalizados:
            if nome and nome in valor:
                return indice

    return None


# ==============================================================================
# LEITURA DO EXPORT
# ==============================================================================

def carregar_responsaveis_export(export_file):
    """
    Le o EXPORT e cria um dicionario no formato:

        centro de custo -> responsavel SAP

    Linhas contendo indicacao de inatividade ou cancelamento
    nao sao utilizadas.
    """

    print()
    print("Lendo o arquivo EXPORT...")

    cabecalho = encontrar_cabecalho_dataframe(
        export_file,
        termos_obrigatorios=[
            "Centro custo",
            "Responsável",
        ]
    )

    export_df = pd.read_excel(
        export_file,
        header=cabecalho
    )

    export_df.columns = [
        str(coluna).strip()
        for coluna in export_df.columns
    ]

    coluna_cc = localizar_coluna_por_nomes(
        export_df.columns,
        [
            "Centro custo",
            "Centro de custo",
        ]
    )

    coluna_responsavel = localizar_coluna_por_nomes(
        export_df.columns,
        [
            "Responsável",
            "Responsavel",
        ]
    )

    if coluna_cc is None:
        raise RuntimeError(
            "A coluna 'Centro custo' nao foi encontrada no EXPORT."
        )

    if coluna_responsavel is None:
        raise RuntimeError(
            "A coluna 'Responsavel' nao foi encontrada no EXPORT."
        )

    export_df = export_df.dropna(
        how="all"
    ).copy()

    # ==========================================================
    # REMOVE LINHAS INATIVAS OU CANCELADAS
    # ==========================================================

    mascara_registro_invalido = pd.Series(
        False,
        index=export_df.index
    )

    padrao_registro_invalido = (
        r"\bINATIVO\b"
        r"|\bINATIVA\b"
        r"|\bINATIVOS\b"
        r"|\bINATIVAS\b"
        r"|\bINAT\b"
        r"|\bINAT-"
        r"|\bCANCELADO\b"
        r"|\bCANCELADA\b"
        r"|\bCANCELADOS\b"
        r"|\bCANCELADAS\b"
        r"|\bCANCEL\b"
    )

    for coluna in export_df.columns:
        valores = (
            export_df[coluna]
            .fillna("")
            .astype(str)
        )

        mascara_registro_invalido = (
            mascara_registro_invalido
            | valores.str.contains(
                padrao_registro_invalido,
                case=False,
                na=False,
                regex=True
            )
        )

    quantidade_invalidos = int(
        mascara_registro_invalido.sum()
    )

    export_df = export_df.loc[
        ~mascara_registro_invalido
    ].copy()

    print(
        "Registros inativos ou cancelados ignorados: "
        f"{quantidade_invalidos}"
    )

    # ==========================================================
    # NORMALIZA CENTRO DE CUSTO E RESPONSAVEL
    # ==========================================================

    export_df["_CC_NORMALIZADO"] = export_df[
        coluna_cc
    ].map(normalizar_codigo_cc)

    export_df["_RESPONSAVEL_NORMALIZADO"] = export_df[
        coluna_responsavel
    ].map(norm)

    export_df = export_df.loc[
        (export_df["_CC_NORMALIZADO"] != "")
        & (export_df["_RESPONSAVEL_NORMALIZADO"] != "")
    ].copy()

    # Se restarem varios registros validos para o mesmo CC,
    # mantem o ultimo registro encontrado no EXPORT.
    export_df = export_df.drop_duplicates(
        subset=["_CC_NORMALIZADO"],
        keep="last"
    )

    responsavel_por_cc = dict(
        zip(
            export_df["_CC_NORMALIZADO"],
            export_df[coluna_responsavel]
            .astype(str)
            .str.strip()
        )
    )

    print(
        "Centros de custo ativos carregados: "
        f"{len(responsavel_por_cc)}"
    )

    return responsavel_por_cc


# ==============================================================================
# LEITURA DA BASE DE ATIVOS
# ==============================================================================

def carregar_base_ativos(ativos_file):
    """
    Le a Base de Ativos e retorna uma lista contendo:

        nome
        nome normalizado
        superior
        cargo
        email
    """

    print()
    print("Lendo a Base de Ativos...")

    workbook = load_workbook(
        ativos_file,
        read_only=True,
        data_only=True
    )

    worksheet = workbook.worksheets[0]

    numero_cabecalho, valores_cabecalho = (
        encontrar_cabecalho_worksheet(
            worksheet,
            termos_obrigatorios=[
                "Nome Funcionário",
                "Nome Superior",
            ]
        )
    )

    coluna_nome = localizar_indice_coluna_cabecalho(
        valores_cabecalho,
        [
            "Nome Funcionário",
            "Nome Funcionario",
        ]
    )

    coluna_superior = localizar_indice_coluna_cabecalho(
        valores_cabecalho,
        [
            "Nome Superior",
        ]
    )

    coluna_cargo = localizar_indice_coluna_cabecalho(
        valores_cabecalho,
        [
            "Nome da Função",
            "Nome da Funcao",
            "Função",
            "Funcao",
            "Cargo",
        ]
    )

    coluna_email = localizar_indice_coluna_cabecalho(
        valores_cabecalho,
        [
            "E-mail",
            "Email",
            "E Mail",
        ]
    )

    if coluna_nome is None:
        workbook.close()

        raise RuntimeError(
            "A coluna 'Nome Funcionario' nao foi encontrada "
            "na Base de Ativos."
        )

    if coluna_superior is None:
        workbook.close()

        raise RuntimeError(
            "A coluna 'Nome Superior' nao foi encontrada "
            "na Base de Ativos."
        )

    pessoas = []

    for linha in worksheet.iter_rows(
        min_row=numero_cabecalho + 1,
        values_only=True
    ):
        nome = (
            linha[coluna_nome - 1]
            if coluna_nome - 1 < len(linha)
            else None
        )

        if not nome or not norm(nome):
            continue

        superior = (
            linha[coluna_superior - 1]
            if coluna_superior - 1 < len(linha)
            else None
        )

        cargo = (
            linha[coluna_cargo - 1]
            if coluna_cargo
            and coluna_cargo - 1 < len(linha)
            else None
        )

        email = (
            linha[coluna_email - 1]
            if coluna_email
            and coluna_email - 1 < len(linha)
            else None
        )

        pessoas.append(
            {
                "nome": str(nome).strip(),
                "nome_norm": norm(nome),
                "superior": (
                    ""
                    if superior is None
                    else str(superior).strip()
                ),
                "cargo": (
                    ""
                    if cargo is None
                    else str(cargo).strip()
                ),
                "email": (
                    ""
                    if email is None
                    else str(email).strip()
                ),
            }
        )

    workbook.close()

    print(
        f"Registros carregados da Base de Ativos: "
        f"{len(pessoas)}"
    )

    return pessoas

def pontuacao_pessoa(nome_sap, pessoa):
    """
    Calcula a pontuacao de uma pessoa candidata.

    Regra principal:

    Se mais de uma pessoa for compativel com o nome informado
    pelo SAP, o maior cargo deve ter prioridade.

    O superior sera obtido diretamente da mesma linha da pessoa
    selecionada na Base de Ativos.
    """

    nome_sap_normalizado = norm(nome_sap)
    nome_base_normalizado = pessoa["nome_norm"]

    partes_sap = nome_sap_normalizado.split()
    partes_base = nome_base_normalizado.split()

    if not partes_sap or not partes_base:
        return (
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
        )

    nome_exato = int(
        nome_sap_normalizado == nome_base_normalizado
    )

    primeiro_nome_exato = int(
        partes_sap[0] == partes_base[0]
    )

    ultimo_nome_exato = int(
        partes_sap[-1] == partes_base[-1]
    )

    quantidade_partes_exatas = len(
        set(partes_sap).intersection(
            set(partes_base)
        )
    )

    diferenca_quantidade_partes = abs(
        len(partes_sap) - len(partes_base)
    )

    peso_cargo = get_cargo_weight(
        pessoa.get("cargo", "")
    )

    superior_preenchido = int(
        bool(
            norm(
                pessoa.get("superior", "")
            )
        )
    )

    email_preenchido = int(
        bool(
            pessoa.get(
                "email",
                ""
            ).strip()
        )
    )

    return (
        nome_exato,
        primeiro_nome_exato,
        peso_cargo,
        ultimo_nome_exato,
        quantidade_partes_exatas,
        -diferenca_quantidade_partes,
        superior_preenchido,
        email_preenchido,
    )

def deduplicar_pessoas(pessoas):
    """
    Remove registros completamente repetidos.

    Registros com o mesmo nome, mas cargos ou superiores diferentes,
    permanecem na lista para que o desempate seja realizado corretamente.
    """

    pessoas_unicas = []
    chaves_encontradas = set()

    for pessoa in pessoas:
        chave = (
            pessoa.get("nome_norm", ""),
            norm(pessoa.get("cargo", "")),
            norm(pessoa.get("superior", "")),
            pessoa.get("email", "").strip().lower(),
        )

        if chave in chaves_encontradas:
            continue

        chaves_encontradas.add(chave)
        pessoas_unicas.append(pessoa)

    return pessoas_unicas


def localizar_pessoa_na_base(nome_sap, pessoas):
    """
    Localiza o responsavel informado pelo SAP na Base de Ativos.

    Regras:

    1. Procura primeiro o nome exato.
    2. Se nao encontrar, utiliza a comparacao aproximada.
    3. Se houver mais de um candidato compativel, utiliza a pontuacao.
    4. O cargo atua como desempate.
    5. Retorna o registro completo da linha da Base de Ativos.

    O superior nao passa por uma nova pesquisa. Ele sera obtido
    diretamente do campo 'superior' desse registro.
    """

    nome_sap_normalizado = norm(nome_sap)

    if not nome_sap_normalizado:
        return None

    # ==========================================================
    # PRIMEIRA TENTATIVA: NOME EXATO
    # ==========================================================

    correspondencias_exatas = [
        pessoa
        for pessoa in pessoas
        if pessoa["nome_norm"] == nome_sap_normalizado
    ]

    correspondencias_exatas = deduplicar_pessoas(
        correspondencias_exatas
    )

    if correspondencias_exatas:
        correspondencias_exatas.sort(
            key=lambda pessoa: pontuacao_pessoa(
                nome_sap,
                pessoa
            ),
            reverse=True
        )

        return correspondencias_exatas[0]

    # ==========================================================
    # SEGUNDA TENTATIVA: NOME APROXIMADO
    # ==========================================================

    correspondencias_flexiveis = [
        pessoa
        for pessoa in pessoas
        if names_match(
            nome_sap,
            pessoa["nome"]
        )
    ]

    correspondencias_flexiveis = deduplicar_pessoas(
        correspondencias_flexiveis
    )

    if not correspondencias_flexiveis:
        return None

    correspondencias_flexiveis.sort(
        key=lambda pessoa: pontuacao_pessoa(
            nome_sap,
            pessoa
        ),
        reverse=True
    )

    pessoa_selecionada = correspondencias_flexiveis[0]

    # ==========================================================
    # LOG DE DESEMPATE
    # ==========================================================

    if len(correspondencias_flexiveis) > 1:
        print()
        print(
            "Mais de uma pessoa compativel encontrada para: "
            f"{nome_sap}"
        )

        limite_log = 5

        for pessoa in correspondencias_flexiveis[:limite_log]:
            print(
                f"   - {pessoa['nome']} | "
                f"Cargo: {pessoa['cargo']} | "
                f"Peso: {get_cargo_weight(pessoa['cargo'])}"
            )

        quantidade_restante = (
            len(correspondencias_flexiveis)
            - limite_log
        )

        if quantidade_restante > 0:
            print(
                f"   ... mais {quantidade_restante} "
                "candidato(s) compativel(is)"
            )

        print(
            f"   Selecionado: "
            f"{pessoa_selecionada['nome']} | "
            f"{pessoa_selecionada['cargo']}"
        )

        print(
            f"   Superior da mesma linha: "
            f"{pessoa_selecionada['superior'] or 'VAZIO'}"
        )

    return pessoa_selecionada

# ==============================================================================
# FORMATACAO DA PLANILHA DE SAIDA
# ==============================================================================

def configurar_coluna_status(worksheet, coluna_status):
    """
    Aplica uma formatacao simples ao cabecalho da coluna de status.
    """

    preenchimento = PatternFill(
        fill_type="solid",
        fgColor="FFC000"
    )

    fonte = Font(
        bold=True,
        color="000000"
    )

    alinhamento = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    lado = Side(
        style="thin",
        color="A6A6A6"
    )

    borda = Border(
        left=lado,
        right=lado,
        top=lado,
        bottom=lado
    )

    celula = worksheet.cell(
        row=1,
        column=coluna_status
    )

    celula.value = "Status Atualização SAP"
    celula.fill = preenchimento
    celula.font = fonte
    celula.alignment = alinhamento
    celula.border = borda

    worksheet.column_dimensions[
        celula.column_letter
    ].width = 55


def definir_status(
    worksheet,
    linha,
    coluna_status,
    texto,
    cor_fonte
):
    """
    Grava e formata o status de uma linha.
    """

    celula = worksheet.cell(
        row=linha,
        column=coluna_status
    )

    celula.value = texto

    celula.font = Font(
        color=cor_fonte
    )

    celula.alignment = Alignment(
        vertical="center",
        wrap_text=True
    )


# ==============================================================================
# PROCESSO PRINCIPAL
# ==============================================================================

def atualizar_centro_custo_viagens():
    print("[PROGRESSO: 5]")

    print("=" * 72)
    print(
        " ATUALIZACAO DE APROVADORES - CENTRO DE CUSTO VIAGENS ".center(72)
    )
    print("=" * 72)

    (
        export_file,
        ativos_file,
        viagens_file,
        arquivo_saida
    ) = validar_arquivos()

    print("[PROGRESSO: 15]")

    print()
    print("Arquivos encontrados:")
    print(f"  EXPORT: {export_file.name}")
    print(f"  Base de Ativos: {ativos_file.name}")
    print(f"  Centro de Custo Viagens: {viagens_file.name}")

    responsavel_por_cc = carregar_responsaveis_export(
        export_file
    )

    print("[PROGRESSO: 35]")

    pessoas_base = carregar_base_ativos(
        ativos_file
    )

    print("[PROGRESSO: 55]")

    print()
    print("Atualizando a planilha de viagens...")

    workbook = load_workbook(
        viagens_file
    )

    worksheet = workbook.active

    # ==========================================================
    # LOCALIZA OU CRIA A COLUNA DE STATUS
    # ==========================================================

    coluna_status = None

    for coluna in range(
        1,
        worksheet.max_column + 1
    ):
        valor_cabecalho = norm(
            worksheet.cell(
                row=1,
                column=coluna
            ).value
        )

        if valor_cabecalho == norm(
            "Status Atualização SAP"
        ):
            coluna_status = coluna
            break

    if coluna_status is None:
        coluna_status = worksheet.max_column + 1

    configurar_coluna_status(
        worksheet,
        coluna_status
    )

    print("[PROGRESSO: 60]")

    # ==========================================================
    # CONTADORES
    # ==========================================================

    quantidade_ok = 0
    quantidade_alterados = 0
    quantidade_cc_nao_encontrado = 0
    quantidade_nome_nao_encontrado = 0
    quantidade_superior_vazio = 0
    quantidade_linhas_vazias = 0

    # ==========================================================
    # CACHE DE RESPONSAVEIS
    # ==========================================================

    # Cada nome do SAP sera pesquisado apenas uma vez.
    # O valor armazenado sera o registro completo da Base de Ativos.
    cache_pessoas_localizadas = {}

    total_linhas = max(
        worksheet.max_row - 1,
        1
    )

    # ==========================================================
    # PROCESSAMENTO DAS LINHAS
    # ==========================================================

    for indice_processamento, numero_linha in enumerate(
        range(2, worksheet.max_row + 1),
        start=1
    ):
        valor_cc = worksheet.cell(
            row=numero_linha,
            column=COLUNA_CC_VIAGENS
        ).value

        cc_normalizado = normalizar_codigo_cc(
            valor_cc
        )

        if not cc_normalizado:
            quantidade_linhas_vazias += 1
            continue

        # ======================================================
        # LOCALIZA O RESPONSAVEL DO CC NO EXPORT
        # ======================================================

        responsavel_sap = responsavel_por_cc.get(
            cc_normalizado
        )

        if not responsavel_sap:
            definir_status(
                worksheet,
                numero_linha,
                coluna_status,
                "Centro de custo não encontrado no SAP",
                "C00000"
            )

            quantidade_cc_nao_encontrado += 1
            continue

        # ======================================================
        # LOCALIZA O RESPONSAVEL NA BASE DE ATIVOS
        # ======================================================

        chave_cache = norm(
            responsavel_sap
        )

        if chave_cache not in cache_pessoas_localizadas:
            cache_pessoas_localizadas[chave_cache] = (
                localizar_pessoa_na_base(
                    responsavel_sap,
                    pessoas_base
                )
            )

        pessoa_encontrada = cache_pessoas_localizadas[
            chave_cache
        ]

        if pessoa_encontrada is None:
            definir_status(
                worksheet,
                numero_linha,
                coluna_status,
                (
                    "Nome não encontrado na Base de Ativos: "
                    f"{responsavel_sap}"
                ),
                "C00000"
            )

            quantidade_nome_nao_encontrado += 1
            continue

        # ======================================================
        # DADOS DA MESMA LINHA DA BASE DE ATIVOS
        # ======================================================

        # O nome, cargo e superior pertencem ao mesmo registro.
        nome_correto = pessoa_encontrada["nome"]
        nome_superior = pessoa_encontrada["superior"]
        cargo_encontrado = pessoa_encontrada["cargo"]

        aprovador_atual = worksheet.cell(
            row=numero_linha,
            column=COLUNA_APROVADOR_1
        ).value

        superior_atual = worksheet.cell(
            row=numero_linha,
            column=COLUNA_APROVADOR_2
        ).value

        # Para o status, compara o texto normalizado completo.
        # A planilha recebera o nome completo da Base de Ativos.
        aprovador_igual = (
            norm(nome_correto)
            == norm(aprovador_atual)
        )

        if nome_superior:
            superior_igual = (
                norm(nome_superior)
                == norm(superior_atual)
            )
        else:
            superior_igual = not norm(
                superior_atual
            )

        # ======================================================
        # ATUALIZA O PRIMEIRO APROVADOR
        # ======================================================

        worksheet.cell(
            row=numero_linha,
            column=COLUNA_APROVADOR_1
        ).value = nome_correto

        # ======================================================
        # ATUALIZA O SUPERIOR DA MESMA LINHA
        # ======================================================

        # Nao realiza nova busca aproximada.
        # Apenas copia Nome Superior do registro selecionado.
        if nome_superior:
            worksheet.cell(
                row=numero_linha,
                column=COLUNA_APROVADOR_2
            ).value = nome_superior
        else:
            quantidade_superior_vazio += 1

        # ======================================================
        # DEFINE O STATUS
        # ======================================================

        if aprovador_igual and superior_igual:
            definir_status(
                worksheet,
                numero_linha,
                coluna_status,
                "OK",
                "008000"
            )

            quantidade_ok += 1

        else:
            detalhes = []

            if not aprovador_igual:
                detalhes.append(
                    "1º aprovador atualizado"
                )

            if nome_superior and not superior_igual:
                detalhes.append(
                    "superior atualizado"
                )

            if not nome_superior:
                detalhes.append(
                    "Nome Superior vazio na Base de Ativos"
                )

            texto_status = "Alterado"

            if detalhes:
                texto_status += (
                    ": "
                    + " | ".join(detalhes)
                )

            definir_status(
                worksheet,
                numero_linha,
                coluna_status,
                texto_status,
                "C65911"
            )

            quantidade_alterados += 1

        # Progresso dinamico entre 60% e 90%.
        if indice_processamento % 100 == 0:
            percentual = 60 + int(
                (
                    indice_processamento
                    / total_linhas
                )
                * 30
            )

            percentual = min(
                percentual,
                90
            )

            print(
                f"[PROGRESSO: {percentual}]"
            )

    print("[PROGRESSO: 90]")

    # ==========================================================
    # SALVA O RESULTADO
    # ==========================================================

    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

    workbook.save(
        arquivo_saida
    )

    print("[PROGRESSO: 100]")

    # ==========================================================
    # RESUMO
    # ==========================================================

    print()
    print("=" * 72)
    print(
        " PROCESSO CONCLUIDO ".center(72)
    )
    print("=" * 72)

    print(f"OK: {quantidade_ok}")
    print(f"Alterados: {quantidade_alterados}")

    print(
        "Centros de custo não encontrados no SAP: "
        f"{quantidade_cc_nao_encontrado}"
    )

    print(
        "Nomes não encontrados na Base de Ativos: "
        f"{quantidade_nome_nao_encontrado}"
    )

    print(
        "Responsáveis sem Nome Superior na Base de Ativos: "
        f"{quantidade_superior_vazio}"
    )

    print(
        "Linhas sem centro de custo: "
        f"{quantidade_linhas_vazias}"
    )

    print(
        "Responsáveis distintos pesquisados na Base de Ativos: "
        f"{len(cache_pessoas_localizadas)}"
    )

    print()
    print(
        f"Arquivo gerado: {arquivo_saida}"
    )
    print("=" * 72)

# ==============================================================================
# EXECUCAO
# ==============================================================================

if __name__ == "__main__":
    try:
        atualizar_centro_custo_viagens()

    except Exception as erro:
        print()
        print("=" * 72)
        print(" ERRO DURANTE O PROCESSAMENTO ".center(72))
        print("=" * 72)
        print(str(erro))
        print("=" * 72)

        sys.exit(1)