#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
malote_web_scraper.py - Módulo Selenium para consulta ao Malote Web dos Correios
Integrado ao Rateio Malote v14
"""

import os
import re
import json
import time
import glob
from difflib import SequenceMatcher
import openpyxl
import sys
from pathlib import Path

import config
# ==============================================================================
# CONFIGURAÇÃO DE PASTAS DINÂMICAS E CREDENCIAIS
# ==============================================================================
sys.path.append(str(Path(__file__).parent.parent))
from config import EMAIL_MRV, SENHA_MRV

# Aponta dinamicamente para a nova pasta usando o Radar do config
PASTA_MALOTE = Path(config.PASTA_ARQUIVOS) / "rateio_malote"
CACHE_FILE = str(PASTA_MALOTE / "cache_percursos_cc.json")

CORREIOS_EMAIL = EMAIL_MRV
CORREIOS_SENHA = SENHA_MRV

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.service import Service as ChromeService
    from selenium.webdriver.chrome.options import Options as ChromeOptions
    from selenium.webdriver.common.action_chains import ActionChains
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.common.exceptions import (
        TimeoutException, NoSuchElementException,
        ElementClickInterceptedException, StaleElementReferenceException
    )
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False
    print("⚠️ Selenium não instalado. Execute: pip install selenium webdriver-manager")

# ============================================================
# CACHE LOCAL
# ============================================================
def load_cache():
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_cache(cache):
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

# ============================================================
# BASE DE CENTROS DE CUSTO
# ============================================================
class BaseCentroCusto:
    def __init__(self):
        self.centros = []
        self._load()

    def _load(self):
        arquivo = self._find_file()
        if not arquivo:
            raise RuntimeError("A planilha 'BASE CENTRO DE CUSTO' não foi encontrada nas pastas de arquivos.")

        print(f"  📄 Base CC: {Path(arquivo).name}")
        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = self._find_sheet(wb)
            if not ws:
                print("  ⚠️ Nenhuma aba encontrada na base de CC")
                wb.close()
                return

            print(f"  📋 Aba: '{ws.title}' ({ws.max_row} linhas)")
            cc_col, desc_col = self._find_columns(ws)

            if cc_col is None:
                print("  ⚠️ Não conseguiu identificar coluna de CC")
                wb.close()
                return

            print(f"  Colunas → CC:{cc_col} Descrição:{desc_col}")

            for row in ws.iter_rows(min_row=2, values_only=True):
                rl = list(row)
                if len(rl) <= cc_col:
                    continue
                cc = str(rl[cc_col]).strip() if rl[cc_col] else ""
                if not cc or cc.lower() in ('none', '', 'nan'):
                    continue

                desc = ""
                if desc_col is not None and len(rl) > desc_col:
                    desc = str(rl[desc_col]).strip() if rl[desc_col] else ""

                self.centros.append({
                    'cc': cc,
                    'descricao': desc,
                    'desc_norm': self._normalize(desc),
                })

            wb.close()
            print(f"  ✅ Base CC: {len(self.centros)} centros de custo carregados")
        except Exception as e:
            print(f"  ⚠️ Erro ao ler base CC: {e}")

    def _find_file(self):
        pasta_arquivos = PASTA_MALOTE.parent 
        if not pasta_arquivos.exists(): return None
        
        for f in pasta_arquivos.rglob("*.xlsx"):
            fl = f.name.lower()
            if ("centro" in fl and "custo" in fl) or "diagrama" in fl:
                return str(f)
            if "base" in fl and ("cc" in fl or "custo" in fl):
                return str(f)
        return None

    def _find_sheet(self, wb):
        for name in wb.sheetnames:
            if "outubro" in name.lower() or "out" in name.lower():
                return wb[name]
        best, mx = None, 0
        for name in wb.sheetnames:
            sheet = wb[name]
            if (sheet.max_row or 0) > mx:
                mx = sheet.max_row or 0
                best = sheet
        return best

    def _find_columns(self, ws):
        cc_col, desc_col = None, None
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    vl = cell.value.strip().lower()
                    if 'centro' in vl and ('cst' in vl or 'custo' in vl or 'ordem' in vl):
                        cc_col = cell.column - 1
                    elif vl in ('descrição', 'descricao', 'descriçao'):
                        desc_col = cell.column - 1

        if cc_col is None:
            for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
                for i, v in enumerate(row):
                    if v and isinstance(v, (str, int, float)):
                        sv = str(v).strip()
                        if re.match(r'^[A-Z0-9]{8,15}$', sv) or re.match(r'^\d{10,15}$', sv):
                            cc_col = i
                            break
                if cc_col is not None:
                    break

        if desc_col is None and cc_col is not None:
            desc_col = cc_col + 1
        return cc_col, desc_col

    def _normalize(self, text):
        if not text:
            return ""
        result = text.upper().strip()
        for k, v in {'Á': 'A', 'À': 'A', 'Ã': 'A', 'Â': 'A', 'É': 'E', 'Ê': 'E',
                      'Í': 'I', 'Ó': 'O', 'Ô': 'O', 'Õ': 'O', 'Ú': 'U', 'Ç': 'C'}.items():
            result = result.replace(k, v)
        return result

    def find_cc_by_name(self, nome_obra):
        if not nome_obra or not self.centros:
            return "", "", 0

        nome_norm = self._normalize(nome_obra)
        if not nome_norm:
            return "", "", 0

        best_cc, best_desc, best_score = "", "", 0

        generic_words = r'\b(DE|DO|DA|DOS|DAS|E|O|A|OS|AS|RESIDENCIAL|RES|CONDOMINIO|COND|PARQUE|JARDIM|VILA|RUA|AV|AVENIDA|TRAVESSA|EDIFICIO|ED|OBRA|ESCRITORIO|SEDE)\b'
        nome_core = re.sub(generic_words, '', nome_norm).strip()
        nome_core = re.sub(r'\s+', ' ', nome_core)

        for centro in self.centros:
            desc_norm = centro['desc_norm']
            if not desc_norm:
                continue

            if nome_norm == desc_norm:
                return centro['cc'], centro['descricao'], 1.0

            if nome_core and len(nome_core) >= 4:
                if f" {nome_core} " in f" {desc_norm} ":
                    score = 0.95
                    if score > best_score:
                        best_score = score
                        best_cc = centro['cc']
                        best_desc = centro['descricao']
                    continue

            if nome_core:
                words_nome = set(nome_core.split())
                desc_core = re.sub(generic_words, '', desc_norm).strip()
                words_desc = set(desc_core.split())
                
                if words_nome and words_nome.issubset(words_desc):
                    score = 0.90
                    if score > best_score:
                        best_score = score
                        best_cc = centro['cc']
                        best_desc = centro['descricao']
                    continue

            score = SequenceMatcher(None, nome_norm, desc_norm).ratio()
            if score > best_score:
                best_score = score
                best_cc = centro['cc']
                best_desc = centro['descricao']

        if best_score >= 0.80:
            return best_cc, best_desc, best_score
        return "", "", 0

# ============================================================
# ACOMPANHAMENTO VSC
# ============================================================
class AcompanhamentoVSC:
    def __init__(self):
        self.percursos = {}
        self._load()

    def _load(self):
        arquivo = self._find_file()
        if not arquivo:
            raise RuntimeError("A planilha 'Acompanhamento VSC' não foi encontrada nas pastas de arquivos.")

        print(f"  📄 Acompanhamento VSC: {Path(arquivo).name}")
        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            ws = self._find_sheet(wb)
            if not ws:
                print("  ⚠️ Aba 'Percursos ativos' não encontrada")
                wb.close()
                return

            print(f"  📋 Aba: '{ws.title}' ({ws.max_row} linhas)")
            col_map, header_row = self._find_columns(ws)
            print(f"  Colunas mapeadas: {col_map}")

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                rl = list(row)
                perc_idx = col_map.get('percurso', 1)
                if len(rl) <= perc_idx or not rl[perc_idx]:
                    continue

                perc_raw = rl[perc_idx]
                perc_str = str(perc_raw).strip().split()[0] if perc_raw else ""
                perc_num = re.sub(r'\D', '', perc_str)
                if not perc_num or len(perc_num) < 7:
                    continue

                def get_val(key):
                    idx = col_map.get(key)
                    if idx is not None and len(rl) > idx and rl[idx]:
                        return str(rl[idx]).strip()
                    return ""

                self.percursos[perc_num] = {
                    'cep_origem': get_val('cep_origem'),
                    'endereco_origem': get_val('endereco_origem'),
                    'cidade_origem': get_val('cidade_origem'),
                    'cep_destino': get_val('cep_destino'),
                    'endereco_destino': get_val('endereco_destino'),
                    'cidade_destino': get_val('cidade_destino'),
                    'situacao': get_val('situacao'),
                }

            wb.close()
            print(f"  ✅ Percursos carregados: {len(self.percursos)}")
        except Exception as e:
            print(f"  ⚠️ Erro ao ler Acompanhamento VSC: {e}")

    def _find_file(self):
        pasta_arquivos = PASTA_MALOTE.parent 
        if not pasta_arquivos.exists(): return None
        
        for f in pasta_arquivos.rglob("*.xlsx"):
            fl = f.name.lower()
            if "acompanhamento" in fl and "vsc" in fl:
                return str(f)
        return None

    def _find_sheet(self, wb):
        for name in wb.sheetnames:
            nl = name.lower()
            if 'percurso' in nl and ('ativo' in nl or 'ativos' in nl):
                return wb[name]
        for name in wb.sheetnames:
            if 'percurso' in name.lower():
                return wb[name]
        return None

    def _find_columns(self, ws):
        col_map = {}
        header_row = 1

        for row in ws.iter_rows(min_row=1, max_row=3, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    vl = cell.value.strip().lower()
                    col = cell.column - 1

                    if 'percurso' in vl:
                        col_map['percurso'] = col
                        header_row = cell.row
                    elif 'cep' in vl and 'origem' in vl:
                        col_map['cep_origem'] = col
                    elif ('endereco' in vl or 'endereço' in vl) and 'origem' in vl:
                        col_map['endereco_origem'] = col
                    elif ('cidade' in vl or 'city' in vl) and 'uf' in vl:
                        if 'cidade_origem' not in col_map:
                            col_map['cidade_origem'] = col
                        elif 'cidade_destino' not in col_map:
                            col_map['cidade_destino'] = col
                    elif 'cep' in vl and 'destino' in vl:
                        col_map['cep_destino'] = col
                    elif ('endereco' in vl or 'endereço' in vl) and 'destino' in vl:
                        col_map['endereco_destino'] = col
                    elif 'situação' in vl or 'situacao' in vl:
                        col_map['situacao'] = col

        if 'percurso' not in col_map:
            col_map = {
                'percurso': 1, 'cep_origem': 2, 'endereco_origem': 3,
                'cidade_origem': 4, 'cep_destino': 5, 'endereco_destino': 6,
                'cidade_destino': 7, 'situacao': 8,
            }
        return col_map, header_row

    def get_percurso(self, nr_percurso):
        return self.percursos.get(str(nr_percurso), None)

# ============================================================
# SELENIUM - MALOTE WEB SCRAPER
# ============================================================
class MaloteWebScraper:
    URL_LOGIN = "https://apps.correios.com.br/malote/?metodo=actionLogin"

    def __init__(self):
        self.driver = None
        self.logged_in = False

    def start(self):
        if not SELENIUM_AVAILABLE:
            print("  ❌ Selenium não disponível.")
            return False

        try:
            options = ChromeOptions()
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-gpu")
            options.add_argument("--log-level=3")
            options.add_argument("--disable-extensions")
            options.add_argument("--start-maximized")

            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option("useAutomationExtension", False)

            try:
                service = ChromeService(ChromeDriverManager().install())
                self.driver = webdriver.Chrome(service=service, options=options)
            except Exception as e:
                print(f"  ⚠️ webdriver-manager falhou: {e}")
                self.driver = webdriver.Chrome(options=options)

            self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
                "source": """
                    Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                """
            })

            self.driver.implicitly_wait(0)
            self.driver.set_page_load_timeout(60)
            print("  ✅ Chrome iniciado (modo anti-detecção)")
            return True

        except Exception as e:
            print(f"  ❌ Falha ao iniciar Chrome: {e}")
            return False

    def login(self):
        try:
            import pyautogui
            import pyperclip
            pyautogui.FAILSAFE = True
            pyautogui.PAUSE = 0.3
        except ImportError:
            print("  ❌ Execute: pip install pyautogui pyperclip pillow")
            return False

        try:
            print("  🔐 Abrindo Malote Web no Chrome...")
            self.driver.get(self.URL_LOGIN)
            time.sleep(6)

            screen_w, screen_h = pyautogui.size()
            pyautogui.click(screen_w // 2, screen_h // 2)
            time.sleep(0.5)

            print("  ⌨️ Navegando até campo Email...")
            pyautogui.press('tab')
            time.sleep(0.5)
            pyperclip.copy(CORREIOS_EMAIL)
            time.sleep(0.3)
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(0.5)

            print("  ⌨️ Navegando até campo Senha...")
            pyautogui.press('tab')
            time.sleep(0.5)
            pyperclip.copy(CORREIOS_SENHA)
            time.sleep(0.3)
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(0.5)

            print("  ⌨️ Pressionando Enter...")
            pyautogui.press('enter')
            time.sleep(10)

            page_source = self.driver.page_source.lower()
            logado = any(["boa tarde" in page_source, "boa manhã" in page_source, "boa noite" in page_source, "sair" in page_source])
            falhou = any(["senha inválida" in page_source, "usuário inválido" in page_source, "acesso negado" in page_source])

            if logado and not falhou:
                self.logged_in = True
                print("  ✅ Login realizado com sucesso!")
                return True

            sem_formulario = "e-mail do usuário" not in page_source
            if sem_formulario:
                self.logged_in = True
                print("  ✅ Login realizado (formulário não encontrado na página)!")
                return True

            print("  ❌ Login não confirmado.")
            return False

        except Exception as e:
            print(f"  ❌ Erro no login: {e}")
            return False

    def _navigate_to_percursos(self):
        try:
            print("    🔗 Navegando para Percursos...")
            base_url = self.driver.current_url.split("/malote/")[0]
            self.driver.get(f"{base_url}/malote/percurso.do")
            time.sleep(3)
            return True
        except Exception as e:
            print(f"    ❌ Erro ao navegar para percursos: {e}")
            return False

    def _preencher_numero_percurso(self, nr_percurso):
        try:
            campo = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[name='percurso_pk_nrPercurso']"))
            )
            self.driver.execute_script("arguments[0].value = arguments[1];", campo, str(nr_percurso))
            print(f"    ✅ Número preenchido: {nr_percurso}")
            return True
        except Exception as e:
            print(f"    ❌ Erro ao preencher número: {e}")
            return False

    def _set_status_combo(self, status):
        try:
            arrow = WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//input[@id='comboSituacao']/following-sibling::img"))
            )
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", arrow)
            time.sleep(0.5)

            actions = ActionChains(self.driver)
            actions.move_to_element(arrow).click().perform()
            print("    🖱️ Clique físico simulado na seta do combo")
            time.sleep(1.5) 

            opcoes = self.driver.find_elements(By.CSS_SELECTOR, "div.x-combo-list-item")
            for opcao in opcoes:
                if opcao.is_displayed() and opcao.text.strip().lower() == status.lower():
                    actions_item = ActionChains(self.driver)
                    actions_item.move_to_element(opcao).click().perform()
                    print(f"    ✅ Status selecionado com mouse real: {status}")
                    time.sleep(1)
                    return True
                    
            print(f"    ⚠️ Status '{status}' não encontrado nas opções visíveis")
            return False
        except Exception as e:
            print(f"    ❌ Erro ao selecionar status '{status}': {e}")
            return False

    def _click_pesquisar(self):
        try:
            btn = WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "button.pesquisar"))
            )
            self.driver.execute_script("arguments[0].click();", btn)
            print("    ✅ Pesquisar clicado")
            time.sleep(3)
            return True
        except Exception as e:
            print(f"    ❌ Erro ao clicar Pesquisar: {e}")
            return False

    def _check_not_found_dialog(self):
        try:
            textos = self.driver.find_elements(By.CSS_SELECTOR, "span.ext-mb-text")
            for t in textos:
                if "nenhum percurso" in t.text.lower():
                    print("    ℹ️ Nenhum percurso encontrado (dialog detectado)")
                    botoes_ok = self.driver.find_elements(By.XPATH, "//button[text()='OK']")
                    for btn in botoes_ok:
                        if btn.is_displayed():
                            self.driver.execute_script("arguments[0].click();", btn)
                            time.sleep(1)
                            print("    ✅ Dialog fechado (OK)")
                            return True
                    return True
            return False
        except Exception:
            return False

    def _has_result(self):
        try:
            icons = self.driver.find_elements(By.CSS_SELECTOR, "img[title='Detalhar Percurso']")
            return len(icons) > 0
        except:
            return False

    def _click_detalhar(self):
        try:
            script_selecionar_linha = """
                var radio = document.querySelector("input[type='radio'][name='radio']");
                if(radio) {
                    var divInner = radio.parentElement;       
                    var divCell = divInner.parentElement;     
                    var td = divCell.parentElement;           
                    var tr = td.parentElement;                

                    function forceClick(el) {
                        if(!el) return;
                        var evDown = new MouseEvent('mousedown', {bubbles: true, cancelable: true, view: window});
                        el.dispatchEvent(evDown);
                        el.click();
                    }

                    forceClick(tr);
                    forceClick(td);
                    forceClick(divCell);
                    forceClick(divInner);
                    forceClick(radio);
                    
                    radio.checked = true;
                    return true;
                }
                return false;
            """
            self.driver.execute_script(script_selecionar_linha)
            print("    ✅ Linha selecionada (Cliques forçados em toda a estrutura)")
            time.sleep(1.5)

            script_clicar_detalhar = """
                var icon = document.querySelector("img[title='Detalhar Percurso']");
                if(icon) {
                    var evDown = new MouseEvent('mousedown', {bubbles: true, cancelable: true, view: window});
                    icon.dispatchEvent(evDown);
                    icon.click();
                    return true;
                }
                return false;
            """
            self.driver.execute_script(script_clicar_detalhar)
            print("    ✅ Detalhar clicado")
            time.sleep(3)
            return True
        except Exception as e:
            print(f"    ❌ Erro ao selecionar linha ou clicar em detalhar: {e}")
            return False

    def _get_dados_origem(self):
        try:
            campo_comp = self.driver.find_element(By.CSS_SELECTOR, "input[name='dsComplementoO']")
            valor = self.driver.execute_script("return arguments[0].value;", campo_comp).strip()
            if valor:
                print(f"    📍 Origem (dsComplementoO): '{valor}'")
                return valor, "dsComplementoO"
        except: pass

        try:
            campo_fant = self.driver.find_element(By.CSS_SELECTOR, "input[name='nmFantasiaO']")
            valor = self.driver.execute_script("return arguments[0].value;", campo_fant).strip()
            if valor:
                print(f"    📍 Origem (nmFantasiaO): '{valor}'")
                return valor, "nmFantasiaO"
        except: pass

        print("    ⚠️ Nenhum dado de origem encontrado")
        return "", ""

    def _click_tab_destino(self):
        try:
            aba = WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//span[contains(@class, 'x-tab-strip-text') and text()='Destino do Percurso']"))
            )
            self.driver.execute_script("arguments[0].click();", aba)
            print("    ✅ Aba 'Destino do Percurso' clicada")
            time.sleep(2)
            return True
        except Exception as e:
            print(f"    ❌ Erro ao clicar aba destino: {e}")
            return False

    def _get_dados_destino(self):
        try:
            campo_comp = self.driver.find_element(By.CSS_SELECTOR, "input[name='dsComplementoD']")
            valor = self.driver.execute_script("return arguments[0].value;", campo_comp).strip()
            if valor:
                print(f"    📍 Destino (dsComplementoD): '{valor}'")
                return valor, "dsComplementoD"
        except: pass

        try:
            campo_fant = self.driver.find_element(By.CSS_SELECTOR, "input[name='nmFantasiaD']")
            valor = self.driver.execute_script("return arguments[0].value;", campo_fant).strip()
            if valor:
                print(f"    📍 Destino (nmFantasiaD): '{valor}'")
                return valor, "nmFantasiaD"
        except: pass

        print("    ⚠️ Nenhum dado de destino encontrado")
        return "", ""

    def consultar_percurso(self, nr_percurso):
        if not self.logged_in:
            print("    ❌ Não está logado")
            return None

        print(f"    🔍 Consultando percurso {nr_percurso}...")
        found = False
        status_usado = ""

        for status in ["Ativo", "Cancelado", "Suspenso", "Todas"]:
            print(f"      🔄 Tentando status: {status}...")

            if not self._navigate_to_percursos():
                continue

            if not self._preencher_numero_percurso(nr_percurso):
                continue

            if status != "Ativo":
                self._set_status_combo(status)

            self._click_pesquisar()

            if self._check_not_found_dialog():
                print(f"      ❌ Não encontrado como '{status}'")
                continue

            if self._has_result():
                found = True
                status_usado = status
                print(f"      ✅ Encontrado como '{status}'!")
                break
            else:
                print(f"      ❌ Sem resultado para '{status}'")

        if not found:
            print(f"    ❌ Percurso {nr_percurso} não encontrado em nenhum status")
            return None

        if not self._click_detalhar():
            return None

        valor_origem, campo_origem = self._get_dados_origem()
        self._click_tab_destino()
        valor_destino, campo_destino = self._get_dados_destino()

        resultado = {
            'percurso': nr_percurso,
            'status': status_usado,
            'valor_origem': valor_origem,
            'campo_origem': campo_origem,
            'valor_destino': valor_destino,
            'campo_destino': campo_destino,
        }

        print(f"    📦 Resultado: {resultado}")
        return resultado
    
    def close(self):
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass

# ============================================================
# RESOLVEDOR DE CC
# ============================================================
class ResolvedorCC:
    def __init__(self, use_selenium=True):
        self.cache = load_cache()
        self.base_cc = BaseCentroCusto()
        self.acomp_vsc = AcompanhamentoVSC()
        self.scraper = None
        self.use_selenium = use_selenium and SELENIUM_AVAILABLE

    def resolver_cc(self, nr_percurso, nome_obra_agilis=""):
        perc_str = str(nr_percurso)

        if perc_str in self.cache:
            cached = self.cache[perc_str]
            if cached.get('cc'):
                print(f"      💾 Cache: {cached['cc']} ({cached.get('descricao', '')})")
                return cached['cc'], cached.get('descricao', ''), "cache"

        if nome_obra_agilis:
            cc, desc, score = self.base_cc.find_cc_by_name(nome_obra_agilis)
            if cc and score >= 0.80:
                print(f"      📋 Agilis: {cc} = {desc} (score:{score:.2f})")
                self._save_cache(perc_str, cc, desc, "agilis_nome")
                return cc, desc, "agilis_nome"

        perc_data = self.acomp_vsc.get_percurso(perc_str)
        if perc_data:
            endereco_dest = perc_data.get('endereco_destino', '')
            if endereco_dest:
                nome_extraido = self._extract_nome(endereco_dest)
                if nome_extraido:
                    cc, desc, score = self.base_cc.find_cc_by_name(nome_extraido)
                    if cc and score >= 0.80:
                        print(f"      📋 Acomp VSC: {cc} = {desc} (score:{score:.2f})")
                        self._save_cache(perc_str, cc, desc, "acomp_vsc")
                        return cc, desc, "acomp_vsc"

        if self.use_selenium:
            dados = self._consultar_selenium(perc_str)
            if dados:
                for campo in ['complemento_destino', 'complemento_origem',
                              'endereco_destino', 'endereco_origem',
                              'valor_destino', 'valor_origem']: 
                    valor = dados.get(campo, '')
                    if not valor:
                        continue

                    valor_limpo = re.sub(r'\s*/\s*[A-Z]{2}\b', '', valor, flags=re.IGNORECASE)
                    valor_limpo = re.sub(r'\s*-\s*\d+', '', valor_limpo)
                    valor_limpo = re.sub(r'\s*-\s*SEDE\b', '', valor_limpo, flags=re.IGNORECASE)
                    valor_limpo = valor_limpo.strip()

                    nome = self._extract_nome(valor_limpo)
                    buscar = nome if nome else valor_limpo

                    cc, desc, score = self.base_cc.find_cc_by_name(buscar)
                    if cc and score >= 0.80:
                        print(f"      🌐 Selenium [{campo}]: {cc} = {desc} (score:{score:.2f})")
                        self._save_cache(perc_str, cc, desc, f"selenium_{campo}")
                        return cc, desc, f"selenium_{campo}"

                self._save_cache(perc_str, "", "", "selenium_not_found", extra=dados)

        print(f"      ❌ CC não encontrado para percurso {perc_str}")
        return "", "", ""

    def _extract_nome(self, endereco):
        if not endereco:
            return ""

        end_upper = endereco.upper().strip()

        patterns = [
            r'(?:RUA|AV|AVENIDA|ESTRADA|TRAVESSA|ALAMEDA|ROD|RODOVIA)\s+.+?\s+\d+\s+(.+?)(?:\s*[-,]|$)',
            r'((?:RESIDENCIAL|CONDOMINIO|COND\.?|PARQUE|JARDIM|VILLAGE|VILLE|TORRES?)\s+.+?)(?:\s*[-,]\s*(?:BLOCO|BL|APTO|AP|LOTE|LT|QUADRA|QD)|$)',
            r'\d+\s+(.{5,}?)(?:\s*[-,]|$)',
        ]

        for pattern in patterns:
            match = re.search(pattern, end_upper)
            if match:
                nome = match.group(1).strip()
                nome = re.sub(r'\s*(BLOCO|BL|APTO|AP|LOTE|LT|QUADRA|QD|SALA|SL)\s*.*$', '', nome)
                nome = nome.strip(' -,.')
                if len(nome) >= 4:
                    return nome

        if not re.match(r'^(RUA|AV|AVENIDA|ESTRADA|TRAVESSA|ALAMEDA|ROD)', end_upper):
            return end_upper

        return ""

    def _consultar_selenium(self, nr_percurso):
        if not self.scraper:
            self.scraper = MaloteWebScraper()
            if not self.scraper.start():
                self.use_selenium = False
                return None
            if not self.scraper.login():
                self.use_selenium = False
                return None

        return self.scraper.consultar_percurso(nr_percurso)

    def _save_cache(self, percurso, cc, descricao, fonte, extra=None):
        entry = {'cc': cc, 'descricao': descricao, 'fonte': fonte}
        if extra:
            entry.update(extra)
        self.cache[percurso] = entry
        save_cache(self.cache)

    def close(self):
        if self.scraper:
            self.scraper.close()
        save_cache(self.cache)
