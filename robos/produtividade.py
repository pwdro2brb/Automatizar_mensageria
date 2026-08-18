import time
import os
import glob
import win32com.client
import openpyxl
import urllib.parse
import pandas as pd
import re
import shutil
import unicodedata
import pyautogui
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException, NoSuchWindowException, WebDriverException
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime, timedelta
from selenium.webdriver.support.ui import Select
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from selenium.webdriver.support.ui import WebDriverWait
from datetime import date
from openpyxl import load_workbook
from urllib.parse import quote as url_quote
from openpyxl.cell.cell import MergedCell
from calendar import monthrange
import pyperclip
import subprocess
from PIL import Image

import win32gui
import win32con
import win32api
import win32process
import ctypes
import sys
from pathlib import Path

import config
# ==============================================================================
# CONFIGURAÇÃO DE PASTAS DINÂMICAS
# ==============================================================================
from config import EMAIL_MRV, SENHA_MRV, PASTA_DOWNLOADS

# Aponta dinamicamente para a nova pasta usando o Radar do config
PASTA_PRODUTIVIDADE = str(Path(config.PASTA_ARQUIVOS) / "produtividade")

# ==============================================================================
# 🚀 O PULO DO GATO: FORÇAR O WINDOWS A RECONHECER TODOS OS MONITORES
# ==============================================================================
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(2)
except Exception:
    try:
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass
# ==============================================================================

# --- CONFIGURAÇÃO ---
WAIT_TIME = 10

# --- FUNÇÃO DE APOIO: LOGIN MICROSOFT ---
def fazer_login_microsoft(driver, wait, email, senha):
    print("--- Iniciando rotina de Login Microsoft ---")
    try:
        try:
            email_field = wait.until(EC.presence_of_element_located((By.ID, "i0116")))
            print("Preenchendo e-mail...")
            email_field.send_keys(email)
            wait.until(EC.element_to_be_clickable((By.ID, "idSIButton9"))).click()
            
            password_field = wait.until(EC.presence_of_element_located((By.ID, "i0118")))
            print("Preenchendo senha...")
            password_field.send_keys(senha)
            
            clicked = False
            for _ in range(3):
                try:
                    wait.until(EC.element_to_be_clickable((By.ID, "idSIButton9"))).click()
                    clicked = True
                    break
                except StaleElementReferenceException:
                    time.sleep(1)
            if not clicked: raise Exception("Não clicou em Entrar")
            time.sleep(5)
            print("!!! AGUARDANDO APROVAÇÃO MFA (Se necessário) !!!")
            wait.until(EC.element_to_be_clickable((By.ID, "idSIButton9"))).click() 
            print("Login Microsoft efetuado.")
            print("Aguardando janela pop-up fechar...")
            WebDriverWait(driver, 20).until(EC.number_of_windows_to_be(1))
            
            nova_janela_principal = driver.window_handles[0]
            driver.switch_to.window(nova_janela_principal)
            print("Foco retornado para a janela principal do Podio.")
        except TimeoutException:
            print("Campo de login não apareceu. Assumindo que já estamos logados (SSO).")
        return True
    except Exception as e:
        print(f"Erro no Login Microsoft: {e}")
        return False

# ==============================================================================
# FUNÇÃO 1: EXTRAÇÃO WEB E SAP
# ==============================================================================
def extrair_dados_sistemas():
    try:
        driver = webdriver.Chrome()
        driver.maximize_window()
        wait = WebDriverWait(driver, WAIT_TIME)

        # --- PARTE 1: PODIO ---
        print("[PROGRESSO: 10]")
        print("\n=== INICIANDO PARTE 1: PODIO ===")
        driver.get("https://podio.com/login")
        try: wait.until(EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))).click()
        except: pass 
        wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@data-provider='live']"))).click()

        janela_principal = driver.current_window_handle
        wait.until(EC.number_of_windows_to_be(2))
        for handle in driver.window_handles:
            if handle != janela_principal:
                driver.switch_to.window(handle)
                break

        fazer_login_microsoft(driver, wait, EMAIL_MRV, SENHA_MRV)
        driver.switch_to.window(janela_principal)
        
        print("Navegando no Podio...")
        menu_area = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'space-switcher-wrapper')]")))
        ActionChains(driver).move_to_element(menu_area).perform()
        wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'ADM - Núcleo Contratos')]"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@data-app-id='22830484']"))).click()

        print("Aplicando filtros (Método Robusto)...")
        time.sleep(3)
        ul_filtros = wait.until(EC.presence_of_element_located((By.XPATH, "//ul[@class='app-filter-tools']")))
        itens_lista = ul_filtros.find_elements(By.TAG_NAME, "li")
        actions = ActionChains(driver)
        for item in itens_lista: actions.move_to_element(item)
        actions.perform()
        
        wait.until(EC.element_to_be_clickable((By.XPATH, ".//li[@data-original-title='Filtros']"))).click()
        wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@data-id='created_on']"))).click() 
        wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@data-id='-1mr:-1mr']"))).click() 

        # --- ETAPA 8 ---

        try:
            seletor_css = ".app-header__app-menu"
        
            print(f"Procurando todos os elementos com a classe: {seletor_css}")
            WebDriverWait(driver, 10).until(lambda d: len(d.find_elements(By.CSS_SELECTOR, seletor_css)) >= 2)

            elementos = driver.find_elements(By.CSS_SELECTOR, seletor_css)
        
            print(f"Encontrados {len(elementos)} elementos.")

            if len(elementos) > 0:
                print("Clicando no primeiro elemento (índice 0)...")
                elementos[0].click()
        
            print("Aaguardando 2 segundos para a página/menu reagir...")
            time.sleep(2)
        
            print("Re-encontrando os elementos (para segurança)...")
            elementos = driver.find_elements(By.CSS_SELECTOR, seletor_css)

            if len(elementos) > 1:
                print("Clicando no segundo elemento (índice 1)...")
                elementos[1].click()
            else:
                print("Erro: Não foi possível encontrar o segundo elemento após o primeiro clique.")
        
            print("Ações nos dois elementos concluídas!")
            time.sleep(3)

        except Exception as e:
            print(f"Ocorreu um erro: {e}")

        # --- ETAPA 9 ---
        print("Etapa 9: Aguardando o menu dropdown abrir...")
        
        exportar_excel_selector = "a.app-box-supermenu-v2__link.app-export-excel"
        
        exportar_link = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, exportar_excel_selector))
        )
        
        print("Link 'Exportar Excel' encontrado. Clicando...")
        exportar_link.click()

        time.sleep(3)

        # --- ETAPA 10 ---
        try:
            print("Procurando o ícone de 'Notificação' (Inbox)...")
        
            notificacao_selector = "li.navigation-link.inbox"
        
            notificacao_icon = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, notificacao_selector))
            )
        
            print("Ícone de 'Notificação' encontrado. Clicando...")
            notificacao_icon.click()
        
            time.sleep(1) 

        except Exception as e:
            print(f"Erro ao tentar clicar no ícone de Notificação: {e}")
            driver.save_screenshot("erro_notificacao.png")

        # --- ETAPA 11 ---
        css_corrigido = "a.PodioUI__Notifications__NotificationGroup"
        item_notificacao = WebDriverWait(driver, 3).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, css_corrigido))
        )
        item_notificacao.click()

        print("Aguardando processamento do Excel (até 3 minutos)...")

        tempo_espera = 0
        sucesso_exportacao = False
        nome_do_arquivo = "Mensageria - Última vista usada.xlsx"

        while tempo_espera < 180:
            # Progresso dinâmico enquanto espera o Podio (vai de 10% a 24%)
            prog = 10 + int((tempo_espera / 180) * 14)
            print(f"[PROGRESSO: {prog}]")
            
            try:            
                if EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Mensageria - Última vista usada.xlsx")):

                    print("Exportação 'Completado'!")
 
                    link_download = WebDriverWait(driver, 3).until(
                        EC.element_to_be_clickable((By.LINK_TEXT, nome_do_arquivo))
                    )
                    
                    print("Link encontrado! Clicando para baixar...")
                    link_download.click()

                    sucesso_exportacao = True
                    break 

                    
            except Exception:
                print(f"Aguardando Podio... ({tempo_espera}s / 180s) - Atualizando a página (F5)...")
                time.sleep(10) 
                tempo_espera += 10
                
                driver.refresh() 
                time.sleep(5) 
                tempo_espera += 5
                

            
        if not sucesso_exportacao:
            raise Exception("Tempo limite de 3 minutos excedido aguardando a exportação do Podio.")
            
        print("Download Podio iniciado!")
        time.sleep(1) 
        
        # --- PARTE 2: AGILIS ---
        print("[PROGRESSO: 25]")
        print("\n=== INICIANDO PARTE 2: AGILIS ===")
        driver.get("https://agilis.mrv.com.br/HomePage.do?view_type=my_view")
        try:
            btn_login = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[text()='Login Integrado Microsoft']")))
            btn_login.click()
            fazer_login_microsoft(driver, wait, EMAIL_MRV, SENHA_MRV)
        except TimeoutException:
            print("Botão de login não apareceu, seguindo...")

        print("Navegando menus Agilis...")
        wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Relatórios"))).click()
        wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Contratos - ADM"))).click()
        wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Produtividade Contratos - ADM"))).click()
        wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "linkborder"))).click() 
        wait.until(EC.element_to_be_clickable((By.XPATH, "//option[text()='Coletor de custo ADM']"))).click()
        driver.find_element(By.CLASS_NAME, "moverightButton").click()

        try:
            expand_btn = wait.until(EC.presence_of_element_located((By.ID, "rcstep2src")))
            driver.execute_script("arguments[0].click();", expand_btn)
            time.sleep(1)
        except: pass

        print("Selecionando Data 'Mês Passado' no dropdown...")
        select_elem = wait.until(EC.presence_of_element_located((By.ID, "dateFilterType")))
        Select(select_elem).select_by_visible_text("Mês passado")
        
        print("Selecionando o rádio 'Durante' (Ajuste obrigatório)...")
        selector_radio_durante = (By.CSS_SELECTOR, "input[value='predefined']")
        wait.until(EC.element_to_be_clickable(selector_radio_durante)).click()

        wait.until(EC.element_to_be_clickable((By.ID, "addnew223222"))).click() 
        print("Relatório gerando. Aguardando 10 segundos...")
        time.sleep(10) 

        print("Iniciando o download direto do relatório XLS...")
        DOWNLOAD_XLS_LINK = (By.ID, "exportxls")
        wait.until(EC.element_to_be_clickable(DOWNLOAD_XLS_LINK)).click()
        time.sleep(5)
        print("Relatório Agilis baixado com sucesso!")

        # --- PARTE 3: BÚSSOLA MRV ---
        print("[PROGRESSO: 40]")
        print("\n=== INICIANDO PARTE 3: BÚSSOLA MRV ===")
        driver.get("http://bussola.mrv.com.br/Main/Big.aspx")
        time.sleep(4)
        pyautogui.write(EMAIL_MRV.strip())
        pyautogui.press('tab')
        pyautogui.write(SENHA_MRV.strip())
        pyautogui.press('enter')
        time.sleep(3)

        driver.get("http://report2.mrv.com.br/ReportServer/Pages/ReportViewer.aspx?/BIG/Administrativo/ADM013%20-%20Relat%C3%B3rio%20Protocolo%20de%20Pagamento%20MRV%20PAG/REL_PRLPGTMRV&rs:Command=Render")
        time.sleep(4)
        pyautogui.write(EMAIL_MRV.strip())
        pyautogui.press('tab')
        pyautogui.write(SENHA_MRV.strip())
        pyautogui.press('enter')
        time.sleep(3)
        
        driver.get("http://bussola.mrv.com.br/Main/Big.aspx")
        pasta_adm = wait.until(EC.element_to_be_clickable((By.ID, "pasta2")))
        pasta_adm.click()
        time.sleep(2)

        xpath_relatorio = "//div[@id='divLinha' and contains(., 'Relatório Protocolo de Pagamento MRV PAG')]"
        relatorio_link = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_relatorio)))
        relatorio_link.click()
        
        WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)
        nova_janela = driver.window_handles[-1]
        driver.switch_to.window(nova_janela)
        wait = WebDriverWait(driver, 10)

        hoje = date.today()
        primeiro_dia_mes_atual = hoje.replace(day=1)
        ultimo_dia_mes_passado = primeiro_dia_mes_atual - timedelta(days=1)
        primeiro_dia_mes_passado = ultimo_dia_mes_passado.replace(day=1)
        data_inicio_str = primeiro_dia_mes_passado.strftime("%d/%m/%Y")
        data_final_str = ultimo_dia_mes_passado.strftime("%d/%m/%Y")

        time.sleep(3) 
        xpath_input_inicio = "(//button[@aria-label='Data criação inicio']/preceding::input)[last()]"
        input_inicio = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_input_inicio)))
        input_inicio.clear() 
        input_inicio.send_keys(data_inicio_str)

        xpath_input_final = "(//button[@aria-label='Data criação final']/preceding::input)[last()]"
        input_final = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_input_final)))
        input_final.clear()
        input_final.send_keys(data_final_str)

        dropdown_status = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl07_txtValue")))
        dropdown_status.click()
        time.sleep(1)
        checkbox_todos = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl07_divDropDown_ctl00")))
        checkbox_todos.click()
        time.sleep(1)

        btn_exibir = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl00")))
        btn_exibir.click()

        wait_longo = WebDriverWait(driver, 120)
        imagem_relatorio = wait_longo.until(EC.presence_of_element_located((By.XPATH, "//img[@alt='Imagem do relatório']")))
        time.sleep(2)

        btn_exportar = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl05_ctl04_ctl00_ButtonImgDown")))
        btn_exportar.click()
        time.sleep(1)

        btn_excel = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@title='Excel' or @alt='Excel']")))
        btn_excel.click()
        
        print("[PROGRESSO: 55]")
        time.sleep(15) 
        print("Download do Bússola finalizado!")



        # --- MOVER ARQUIVOS ---
        print("[PROGRESSO: 60]")
        time.sleep(5)
        if not os.path.exists(PASTA_PRODUTIVIDADE):
            os.makedirs(PASTA_PRODUTIVIDADE)

        files = [os.path.join(PASTA_DOWNLOADS, f) for f in os.listdir(PASTA_DOWNLOADS) if os.path.isfile(os.path.join(PASTA_DOWNLOADS, f))]
        files.sort(key=os.path.getmtime, reverse=True)
        top_4_files = files[:4]

        for file_path in top_4_files:
            file_name = os.path.basename(file_path)
            try:
                shutil.move(file_path, os.path.join(PASTA_PRODUTIVIDADE, file_name))
                print(f"Sucesso: {file_name} movido para {PASTA_PRODUTIVIDADE}")
            except Exception as e:
                print(f"Erro ao mover {file_name}: {e}")

        driver.quit()

    except Exception as e:
        print(f"Erro geral na extração: {e}")
        try: driver.quit() 
        except: pass
        raise e

# ==============================================================================
# FUNÇÕES AUXILIARES E DE PROCESSAMENTO (AGORA FORA DA EXTRAÇÃO)
# ==============================================================================
def gerar_sap_map(colaboradores):

    retorno = {}

    for c in colaboradores:

        if not c["matricula"]:
            continue

        retorno[c["matricula"]] = {
            "p1": (
                f"{c['login']} {c['matricula']}"
            )
        }

    return retorno

def gerar_cadastro_agilis(colaboradores):

    retorno = []

    for c in colaboradores:

        identificador = c["login"]

        if c["matricula"]:
            identificador += f" {c['matricula']}"

        retorno.append({
            "p2": c["nome"],
            "p1": identificador,
            "min_col_letter": "D"
        })

    return retorno

def gerar_lanctos_map(colaboradores):

    retorno = {}

    for c in colaboradores:

        retorno[c["login"]] = {
            "p1": (
                f"{c['login']} {c['matricula']}"
                if c["matricula"]
                else c["login"]
            )
        }

    return retorno

def gerar_sedex_map(colaboradores):

    retorno = {}

    for c in colaboradores:

        identificador = c["login"]

        if c["matricula"]:
            identificador += f" {c['matricula']}"

        retorno[c["nome"]] = identificador

    return retorno

def carregar_cadastro_colaboradores(prod_path):

    df = pd.read_excel(
        prod_path,
        sheet_name="Nomes"
    )

    colaboradores = []

    for _, row in df.iterrows():

        nome = str(
            row["Nome Completo"]
        ).strip()

        login = str(
            row["Login"]
        ).strip()

        matricula = ""

        if pd.notna(row["Matrícula"]):
            matricula = str(
                row["Matrícula"]
            ).strip()

        colaboradores.append({
            "nome": nome,
            "login": login,
            "matricula": matricula
        })

    return colaboradores

def descobrir_faixas_atividades(ws):
    """
    Descobre dinamicamente as linhas de atividades de cada colaborador.

    Considera como atividade:
    - Agilis;
    - Sedex/Pac/Malote;
    - Lançamentos;
    - SAP.

    Não inclui a linha Total.
    """

    atividades_validas = (
        "agilis",
        "sedex",
        "lancamentos",
        "sap"
    )

    faixas = []
    linha_colaborador = None
    linhas_atividades = []

    for row in range(2, ws.max_row + 1):
        valor_colaborador = ws.cell(row=row, column=2).value
        valor_atividade = ws.cell(row=row, column=3).value

        if valor_colaborador not in (None, ""):
            # Salva o bloco anterior.
            if linha_colaborador is not None and linhas_atividades:
                faixas.append(
                    (min(linhas_atividades), max(linhas_atividades))
                )

            linha_colaborador = row
            linhas_atividades = []

        if linha_colaborador is None or valor_atividade is None:
            continue

        atividade_norm = norm_key(valor_atividade)

        if atividade_norm == "total":
            if linhas_atividades:
                faixas.append(
                    (min(linhas_atividades), max(linhas_atividades))
                )

            linha_colaborador = None
            linhas_atividades = []
            continue

        if any(
            chave in atividade_norm
            for chave in atividades_validas
        ):
            linhas_atividades.append(row)

    if linha_colaborador is not None and linhas_atividades:
        faixas.append(
            (min(linhas_atividades), max(linhas_atividades))
        )

    # Remove duplicações, preservando a ordem.
    return list(dict.fromkeys(faixas))

def localizar_bloco_colaborador(ws, texto_busca):
    """
    Localiza um colaborador na coluna B.

    A busca é normalizada, ignorando:
    - maiúsculas e minúsculas;
    - acentos;
    - espaços nas extremidades.

    Retorna a linha do colaborador ou None.
    """

    busca = norm_key(texto_busca)

    if not busca:
        return None

    for row in range(2, ws.max_row + 1):
        valor = ws.cell(row=row, column=2).value

        if valor is None:
            continue

        valor_norm = norm_key(valor)

        if busca == valor_norm or busca in valor_norm:
            return row

    return None

def localizar_linha_atividade(ws, linha_colaborador, atividade):
    """
    Localiza uma atividade na coluna C dentro do bloco do colaborador.

    A busca termina quando:
    - encontra a atividade;
    - encontra outro colaborador na coluna B;
    - encontra a linha Total;
    - percorre no máximo 8 linhas.
    """

    if not linha_colaborador:
        return None

    atividade_norm = norm_key(atividade)

    limite = min(linha_colaborador + 8, ws.max_row + 1)

    for row in range(linha_colaborador, limite):
        # Se encontrou outro colaborador, saiu do bloco atual.
        if row > linha_colaborador:
            outro_colaborador = ws.cell(row=row, column=2).value

            if outro_colaborador not in (None, ""):
                break

        valor_atividade = ws.cell(row=row, column=3).value

        if valor_atividade is None:
            continue

        texto = norm_key(valor_atividade)

        if texto == "total":
            break

        if atividade_norm in texto:
            return row

    return None

def find_column_ignore_case(df, column_name):
    for col in df.columns:
        if col.lower() == column_name.lower(): return col
    return None

def processar_mensageria(filepath, new_filename):
    try:
        df = pd.read_excel(filepath, header=0) 
        df.columns = [str(col).strip() for col in df.columns]
        coluna_usuario = find_column_ignore_case(df, 'Criado por')
        coluna_data = find_column_ignore_case(df, 'Criado em')
        coluna_valores = find_column_ignore_case(df, 'Numero do chamado Agilis/Rastreio')
        df[coluna_data] = pd.to_datetime(df[coluna_data], dayfirst=True).dt.date
        pivot_table = pd.pivot_table(df, index=coluna_usuario, columns=coluna_data, values=coluna_valores, aggfunc='count', fill_value=0, margins=True, margins_name='Total Geral')
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            pivot_table.to_excel(writer, sheet_name='TabelaDinamica')
        os.rename(filepath, new_filename)
    except Exception as e: print(f"ERRO Mensageria: {e}")

def processar_produtividade(filepath, new_filename):
    try:
        df = pd.read_excel(filepath, header=0)
        df.columns = [str(col).strip() for col in df.columns]
        coluna_usuario = find_column_ignore_case(df, 'Nome do usuário')
        coluna_data = find_column_ignore_case(df, 'Data de lançamento')
        coluna_valores = find_column_ignore_case(df, 'Nº doc.faturamento')
        df[coluna_data] = pd.to_datetime(df[coluna_data], dayfirst=True).dt.date
        pivot_table = pd.pivot_table(df, index=coluna_usuario, columns=coluna_data, values=coluna_valores, aggfunc='count', fill_value=0, margins=True, margins_name='Total Geral')
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            pivot_table.to_excel(writer, sheet_name='TabelaDinamica')
        os.rename(filepath, new_filename)
    except Exception as e: print(f"ERRO Produtividade: {e}")

def processar_numerico(filepath, new_filename):
    try:
        df = pd.read_excel(filepath, header=8)
        df.columns = [str(col).strip() for col in df.columns]
        coluna_tecnico = find_column_ignore_case(df, 'Técnico')
        coluna_data = find_column_ignore_case(df, 'Hora de conclusão')
        coluna_valores = find_column_ignore_case(df, 'Identificação da solicitação')
        df[coluna_data] = pd.to_datetime(df[coluna_data], dayfirst=True).dt.date
        pivot_table = pd.pivot_table(df, index=coluna_tecnico, columns=coluna_data, values=coluna_valores, aggfunc='count', fill_value=0, margins=True, margins_name='Total Geral')
        with pd.ExcelWriter(new_filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='DadosOriginais', index=False)
            pivot_table.to_excel(writer, sheet_name='TabelaDinamica')
        os.remove(filepath)
    except Exception as e: print(f"ERRO Numérico: {e}")

def processar_relatorio_pedidos(filepath, new_filename):
    try:
        df = pd.read_excel(filepath, header=1)
        df.columns = [str(col).strip() for col in df.columns]
        coluna_linhas = find_column_ignore_case(df, 'Respons. Entrega')
        coluna_colunas = find_column_ignore_case(df, 'Data Entrada NF')
        coluna_valores = find_column_ignore_case(df, 'Nro. Pedido Compra')
        df[coluna_colunas] = pd.to_datetime(df[coluna_colunas], dayfirst=True).dt.date
        pivot_table = pd.pivot_table(df, index=coluna_linhas, columns=coluna_colunas, values=coluna_valores, aggfunc='count', fill_value=0, margins=True, margins_name='Total Geral')
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            pivot_table.to_excel(writer, sheet_name='TabelaDinamica')
        os.rename(filepath, new_filename)
    except Exception as e: print(f"ERRO Pedidos: {e}")

def step_1_prepare_and_rename_reports(diretorio):
    arquivos = glob.glob(os.path.join(diretorio, '*.*'))
    for arquivo in arquivos:
        nome_arquivo = os.path.basename(arquivo)
        if nome_arquivo.startswith('Mensageria - Última vista'):
            processar_mensageria(arquivo, os.path.join(diretorio, 'Relatório - Sedex.Malote.xlsx'))
        elif (nome_arquivo.lower().startswith("export")and nome_arquivo.lower().endswith(".xlsx")):
            processar_produtividade(arquivo, os.path.join(diretorio, 'Relatório - SAP.xlsx'))
        elif nome_arquivo.startswith('REL_PRLPGT'):
            processar_relatorio_pedidos(arquivo, os.path.join(diretorio, 'Relatório - Lançamentos.xlsx'))
        elif re.match(r'^\d+\.(xlsx|xls)$', nome_arquivo):
            processar_numerico(arquivo, os.path.join(diretorio, 'Relatório - Agilis.xlsx'))

# --- FUNÇÕES DE PREENCHIMENTO (Resumidas para economizar espaço, mas mantendo a lógica) ---
def norm_key(s):
    if s is None: return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.lower()

def col_letter_to_index(letter: str) -> int:
    letter = letter.strip().upper()
    n = 0
    for ch in letter:
        if not ('A' <= ch <= 'Z'): continue
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n

def date_keys(v):
    keys = []
    if isinstance(v, datetime):
        keys += [v.strftime("%Y-%m-%d 00:00:00"), v.strftime("%d/%m/%Y")]
    elif isinstance(v, str):
        s = v.strip()
        if re.match(r"^\d{4}-\d{2}-\d{2} 00:00:00$", s): keys.append(s)
        if re.match(r"^\d{2}/\d{2}/\d{4}$", s): keys.append(s)
        try:
            dt = pd.to_datetime(s, errors="raise")
            keys += [dt.strftime("%Y-%m-%d 00:00:00"), dt.strftime("%d/%m/%Y")]
        except: pass
    return list(dict.fromkeys(keys))

def build_header_map(ws):
    hdr = {}
    for c in range(1, ws.max_column+1):
        v = ws.cell(row=1, column=c).value
        if v is None: continue
        if isinstance(v, datetime):
            hdr[v.strftime("%Y-%m-%d 00:00:00")] = c
            hdr[v.strftime("%d/%m/%Y")] = c
        else:
            s = str(v).strip()
            hdr[s] = c
            for k in date_keys(s): hdr[k] = c
    return hdr

def extract_user_key(s: str) -> str:
    if s is None: return ""
    s = str(s).strip().split('@')[0]
    s_norm = unicodedata.normalize("NFKD", s)
    s_norm = "".join(ch for ch in s_norm if not unicodedata.combining(ch))
    tokens = s_norm.strip().split()
    for t in tokens:
        if '.' in t: return re.sub(r'[^A-Za-z\.]', '', t).lower()
    if len(tokens) >= 2:
        return f"{re.sub(r'[^A-Za-z]', '', tokens[0])}.{re.sub(r'[^A-Za-z]', '', tokens[-1])}".lower()
    return re.sub(r'[^A-Za-z]', '', tokens[0]).lower() if tokens else ""

def read_tabledinamica_with_namecol(path, name_col_hint=None):
    df = pd.read_excel(path, sheet_name="TabelaDinamica", engine="openpyxl")
    nome_col = None
    if name_col_hint:
        hint_norm = norm_key(name_col_hint)
        for c in df.columns:
            if norm_key(c) == hint_norm: nome_col = c; break
    if not nome_col:
        heads = {norm_key(c): c for c in df.columns}
        for alvo in ["criado por", "tecnico", "técnico", "respons. entrega", "nome do usuário"]:
            if alvo in heads: nome_col = heads[alvo]; break
    if not nome_col:
        for c in df.columns:
            if df[c].dtype == "O": nome_col = c; break

    day_cols = [c for c in df.columns if c != nome_col and (isinstance(c, datetime) or (isinstance(c, str) and (re.match(r"^\d{2}/\d{2}/\d{4}$", c) or re.match(r"^\d{4}-\d{2}-\d{2}", c))))]
    registros = []
    if nome_col:
        for _, row in df.iterrows():
            nome_val = str(row.get(nome_col, "")).strip()
            if not nome_val or norm_key(nome_val).startswith(norm_key("Total Geral")): continue
            for d in day_cols:
                val = row.get(d, 0)
                try: v = int(val) if pd.notna(val) else 0
                except: v = 0
                registros.append({"nome": nome_val, "data_obj": d, "valor": v})
    return pd.DataFrame(registros, columns=["nome","data_obj","valor"])

def read_lanctos_tabledinamica(path):
    df = pd.read_excel(path, sheet_name="TabelaDinamica", engine="openpyxl")
    possible_names = ["Respons. Entrega", "Técnico", "Tecnico", "Criado por", "Nome do usuário", "User"]
    nome_col = None
    cols_map = {norm_key(c): c for c in df.columns}
    for alvo in possible_names:
        if norm_key(alvo) in cols_map: nome_col = cols_map[norm_key(alvo)]; break
    if nome_col is None: nome_col = df.columns[0]

    day_cols = [c for c in df.columns if c != nome_col and (isinstance(c, datetime) or (isinstance(c, str) and (re.match(r"^\d{2}/\d{2}/\d{4}$", c) or re.match(r"^\d{4}-\d{2}-\d{2}", c))))]
    registros = []
    for _, row in df.iterrows():
        raw_name = str(row.get(nome_col, "")).strip()
        if not raw_name or norm_key(raw_name).startswith(norm_key("Total Geral")): continue
        ukey = extract_user_key(raw_name)
        for d in day_cols:
            val = row.get(d, 0)
            try: v = int(val) if pd.notna(val) else 0
            except: v = 0
            registros.append({"user_key": ukey, "data_obj": d, "valor": v})
    return pd.DataFrame(registros, columns=["user_key","data_obj","valor"])

def update_headers_to_previous_month(ws, header_row=1, start_col_letter="D", end_col_letter="AH", ref_date=None):
    start_col = col_letter_to_index(start_col_letter)
    end_col   = col_letter_to_index(end_col_letter)
    if ref_date is None: ref_date = date.today()
    first_of_month = date(ref_date.year, ref_date.month, 1)
    last_day_prev  = first_of_month - timedelta(days=1)
    ano, mes = last_day_prev.year, last_day_prev.month
    qtd_dias = monthrange(ano, mes)[1]

    for i in range(qtd_dias):
        ws.cell(row=header_row, column=start_col + i, value=datetime(ano, mes, i + 1))
    for c in range(start_col + qtd_dias, end_col + 1):
        ws.cell(row=header_row, column=c, value=None)
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=header_row, column=c)
        if isinstance(cell.value, datetime): cell.number_format = "dd/mm/yyyy"
    return ano, mes, qtd_dias

def clear_month_data_in_blocks(ws, row_ranges, start_col_letter="D", end_col_letter="AH"):
    start_col = col_letter_to_index(start_col_letter)
    end_col   = col_letter_to_index(end_col_letter)
    cleared = 0
    for (r0, r1) in row_ranges:
        for r in range(r0, r1 + 1):
            if r > ws.max_row: continue
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell): continue
                if cell.value not in (None, ""):
                    cell.value = None
                    cleared += 1
    return cleared

def fill_agilis_same_row(ws, header_map, df_long, AGILIS_POS):
    if df_long.empty:
        print("⚠️ Relatório do Agilis sem dados.")
        return 0

    grp_exact = {
        norm_key(nome): grupo
        for nome, grupo in df_long.groupby(
            df_long["nome"].apply(norm_key)
        )
    }

    grp_ukey = {
        chave: grupo
        for chave, grupo in df_long.groupby(
            df_long["nome"].apply(extract_user_key)
        )
        if chave
    }

    total_writes = 0

    for item in AGILIS_POS:
        p2 = item["p2"]
        p1 = item["p1"]

        # Coluna mínima configurada para o colaborador.
        min_idx = col_letter_to_index(
            item.get("min_col_letter", "D")
        )

        # A área mensal atual termina em AH. Se a configuração antiga
        # apontar para uma coluna inexistente, começa pela coluna D.
        if min_idx > ws.max_column:
            min_idx = col_letter_to_index("D")

        linha_colaborador = localizar_bloco_colaborador(
            ws,
            p1
        )

        if linha_colaborador is None:
            print(
                f"⚠️ Agilis: colaborador não encontrado na planilha: {p1}"
            )
            continue

        row_agilis = localizar_linha_atividade(
            ws,
            linha_colaborador,
            "Agilis"
        )

        if row_agilis is None:
            print(
                f"⚠️ Agilis: atividade não encontrada para {p1}"
            )
            continue

        # Primeiro tenta encontrar pelo nome exato do relatório.
        sub = grp_exact.get(norm_key(p2))

        # Se não encontrar, tenta pelo login extraído.
        if sub is None or sub.empty:
            chave_usuario = extract_user_key(p2)
            sub = grp_ukey.get(chave_usuario)

        if sub is None or sub.empty:
            print(
                f"ℹ️ Agilis: nenhum registro encontrado para "
                f"{p2}."
            )
            continue

        writes = 0

        for _, reg in sub.iterrows():
            col_idx = next(
                (
                    header_map.get(chave)
                    for chave in date_keys(reg["data_obj"])
                    if header_map.get(chave) is not None
                ),
                None
            )

            if col_idx is None:
                continue

            if col_idx < min_idx:
                continue

            try:
                valor = int(reg["valor"])
            except (TypeError, ValueError):
                valor = 0

            if valor != 0:
                ws.cell(
                    row=row_agilis,
                    column=col_idx,
                    value=valor
                )
                writes += 1

        print(
            f"✅ Agilis: {p1} | linha {row_agilis} | "
            f"{writes} valores preenchidos."
        )

        total_writes += writes

    print(
        f"✅ Total Agilis: {total_writes} valores preenchidos."
    )

    return total_writes


def fill_sedex(ws, header_map, df_long, MAP_SEDEX):
    if df_long.empty: return 0
    grp = {norm_key(n): sub for n, sub in df_long.groupby(df_long['nome'].apply(norm_key))}
    total_writes = 0
    for p2, p1 in MAP_SEDEX.items():
        r_nome = next((r for r in range(2, ws.max_row+1) if isinstance(ws.cell(row=r, column=2).value, str) and norm_key(ws.cell(row=r, column=2).value) == norm_key(p1)), None)
        if r_nome is None: continue
        r_sedex = next((rr for rr in range(r_nome, min(ws.max_row, r_nome+12)+1) if isinstance(ws.cell(row=rr, column=3).value, str) and 'sedex/pac/malote' in ws.cell(row=rr, column=3).value.strip().lower()), None)
        if r_sedex is None: continue
        sub = grp.get(norm_key(p2))
        if sub is None or sub.empty: continue
        writes = 0
        for _, reg in sub.iterrows():
            col_idx = next((header_map.get(k) for k in date_keys(reg["data_obj"]) if header_map.get(k)), None)
            if not col_idx: continue
            val = int(reg["valor"])
            if val != 0:
                ws.cell(row=r_sedex, column=col_idx, value=val)
                writes += 1
        total_writes += writes
    return total_writes

def fill_lanctos_fixed(ws, header_map, df_long, LANCTOS_USER_MAP):
    if df_long.empty or "user_key" not in df_long.columns:
        print("⚠️ Relatório de Lançamentos sem dados.")
        return 0

    grp = {
        norm_key(ukey): sub
        for ukey, sub in df_long.groupby(df_long["user_key"].apply(norm_key))
    }

    total_writes = 0

    for ukey, meta in LANCTOS_USER_MAP.items():
        linha_colaborador = localizar_bloco_colaborador(
            ws,
            meta["p1"]
        )

        if linha_colaborador is None:
            print(
                f"⚠️ Lançamentos: colaborador não encontrado na planilha: "
                f"{meta['p1']}"
            )
            continue

        row_ativ = localizar_linha_atividade(
            ws,
            linha_colaborador,
            "Lançamentos"
        )

        if row_ativ is None:
            print(
                f"⚠️ Lançamentos: atividade não encontrada para "
                f"{meta['p1']}"
            )
            continue

        sub = grp.get(norm_key(ukey))

        if sub is None or sub.empty:
            print(
                f"ℹ️ Lançamentos: nenhum registro encontrado para {ukey}."
            )
            continue

        writes = 0

        for _, reg in sub.iterrows():
            col_idx = next(
                (
                    header_map.get(k)
                    for k in date_keys(reg["data_obj"])
                    if header_map.get(k) is not None
                ),
                None
            )

            if col_idx is None:
                continue

            try:
                valor = int(reg["valor"])
            except (TypeError, ValueError):
                valor = 0

            if valor != 0:
                ws.cell(
                    row=row_ativ,
                    column=col_idx,
                    value=valor
                )
                writes += 1

        print(
            f"✅ Lançamentos: {meta['p1']} | "
            f"linha {row_ativ} | {writes} valores preenchidos."
        )

        total_writes += writes

    return total_writes


def fill_sap_fixed(ws, header_map, df_long, SAP_COD_MAP):
    if df_long.empty or "nome" not in df_long.columns:
        print("⚠️ Relatório SAP sem dados.")
        return 0

    total_writes = 0

    for codigo, meta in SAP_COD_MAP.items():
        linha_colaborador = localizar_bloco_colaborador(
            ws,
            meta["p1"]
        )

        if linha_colaborador is None:
            print(
                f"⚠️ SAP: colaborador não encontrado na planilha: "
                f"{meta['p1']}"
            )
            continue

        row_ativ = localizar_linha_atividade(
            ws,
            linha_colaborador,
            "SAP"
        )

        if row_ativ is None:
            print(
                f"⚠️ SAP: atividade não encontrada para "
                f"{meta['p1']}"
            )
            continue

        # regex=False evita interpretar caracteres do código como regex.
        sub = df_long[
            df_long["nome"].astype(str).str.contains(
                codigo,
                case=False,
                regex=False,
                na=False
            )
        ]

        if sub.empty:
            print(
                f"ℹ️ SAP: nenhum registro encontrado para "
                f"{codigo} - {meta['p1']}."
            )
            continue

        writes = 0

        for _, reg in sub.iterrows():
            col_idx = next(
                (
                    header_map.get(k)
                    for k in date_keys(reg["data_obj"])
                    if header_map.get(k) is not None
                ),
                None
            )

            if col_idx is None:
                continue

            try:
                valor = int(reg["valor"])
            except (TypeError, ValueError):
                valor = 0

            if valor != 0:
                ws.cell(
                    row=row_ativ,
                    column=col_idx,
                    value=valor
                )
                writes += 1

        print(
            f"✅ SAP: {meta['p1']} | "
            f"linha {row_ativ} | {writes} valores preenchidos."
        )

        total_writes += writes

    return total_writes

def fill_fsf_flags(ws, header_map, row_ranges):
    cols_to_process = {}

    for date_str, col_idx in header_map.items():
        try:
            if (
                isinstance(date_str, str)
                and "-" in date_str
                and date_str.index("-") == 4
            ):
                dt = pd.to_datetime(
                    date_str,
                    errors="coerce"
                )
            else:
                dt = pd.to_datetime(
                    date_str,
                    dayfirst=True,
                    errors="coerce"
                )

            if pd.notna(dt):
                cols_to_process[col_idx] = dt.date()

        except Exception:
            continue

    total_writes = 0

    for col_idx, data_atual in cols_to_process.items():
        is_weekend = data_atual.weekday() >= 5

        dia_teve_producao = any(
            ws.cell(row=r, column=col_idx).value
            not in (None, "", 0, "0")
            for start_row, end_row in row_ranges
            for r in range(start_row, end_row + 1)
            if r <= ws.max_row
        )

        if is_weekend or not dia_teve_producao:
            for start_row, end_row in row_ranges:
                bloco_teve_producao = any(
                    ws.cell(row=r, column=col_idx).value
                    not in (None, "", 0, "0")
                    for r in range(start_row, end_row + 1)
                    if r <= ws.max_row
                )

                if bloco_teve_producao:
                    continue

                for r in range(start_row, end_row + 1):
                    if r > ws.max_row:
                        continue

                    cell = ws.cell(row=r, column=col_idx)

                    if isinstance(cell, MergedCell):
                        continue

                    if cell.value in (None, "", 0):
                        cell.value = "0"
                        total_writes += 1

    return total_writes

# ==============================================================================
# FUNÇÃO PRINCIPAL DE PROCESSAMENTO DO EXCEL
# ==============================================================================
def main(nome_arquivo_base, nome_arquivo_saida):
    print("[PROGRESSO: 82]")
    PROD_PATH    = os.path.join(PASTA_PRODUTIVIDADE, nome_arquivo_base)
    
    if not os.path.exists(PROD_PATH):
        raise Exception(f"O arquivo base não foi encontrado!\n\nCaminho procurado:\n{PROD_PATH}\n\nVerifique se a pasta existe e se o arquivo está lá com este nome exato.")

    OUT_PATH     = os.path.join(PASTA_PRODUTIVIDADE, nome_arquivo_saida)
    
    AGILIS_PATH  = os.path.join(PASTA_PRODUTIVIDADE, "Relatório - Agilis.xlsx")
    SEDEX_PATH   = os.path.join(PASTA_PRODUTIVIDADE, "Relatório - Sedex.Malote.xlsx")
    LANCTOS_PATH = os.path.join(PASTA_PRODUTIVIDADE, "Relatório - Lançamentos.xlsx")
    SAP_PATH     = os.path.join(PASTA_PRODUTIVIDADE, "Relatório - SAP.xlsx")

    wb = load_workbook(PROD_PATH)
    ws = wb["Plan1"]

    COLABORADORES = carregar_cadastro_colaboradores(
        PROD_PATH
    )

    print(
        f"✅ {len(COLABORADORES)} colaboradores carregados."
    )

    print("\n=== CADASTRO CARREGADO ===")

    for c in COLABORADORES:
        print(c)

    print("=========================\n")

    MAP_SEDEX = gerar_sedex_map(
        COLABORADORES
    )

    AGILIS_POS = gerar_cadastro_agilis(
        COLABORADORES
    )

    LANCTOS_USER_MAP = gerar_lanctos_map(
        COLABORADORES
    )

    SAP_COD_MAP = gerar_sap_map(
        COLABORADORES
    )

    print(f"✅ {len(COLABORADORES)} colaboradores carregados.")

    ano, mes, qtd_dias = update_headers_to_previous_month(ws, header_row=1, start_col_letter="D", end_col_letter="AH")
    
    ROW_RANGES_ATIV = descobrir_faixas_atividades(ws)

    print(
        f"✅ Foram encontrados {len(ROW_RANGES_ATIV)} "
        f"blocos de colaboradores."
    )
    print(f"Faixas identificadas: {ROW_RANGES_ATIV}")

    clear_month_data_in_blocks(ws, ROW_RANGES_ATIV, start_col_letter="D", end_col_letter="AH")
    
    print("[PROGRESSO: 85]")
    header_map = build_header_map(ws)

    df_ag  = read_tabledinamica_with_namecol(AGILIS_PATH)
    df_sd  = read_tabledinamica_with_namecol(SEDEX_PATH)
    df_lan = read_lanctos_tabledinamica(LANCTOS_PATH)
    df_sap = read_tabledinamica_with_namecol(SAP_PATH, name_col_hint="Nome do usuário")

    print("[PROGRESSO: 90]")
    fill_agilis_same_row(ws, header_map, df_ag, AGILIS_POS)
    fill_sedex(ws, header_map, df_sd, MAP_SEDEX)
    fill_lanctos_fixed(ws, header_map, df_lan, LANCTOS_USER_MAP)
    fill_sap_fixed(ws, header_map, df_sap, SAP_COD_MAP)
    
    print("[PROGRESSO: 95]")
    fill_fsf_flags(ws, header_map, ROW_RANGES_ATIV)

    wb.save(OUT_PATH)
    print(f"✅ Planilha salva com sucesso em: {OUT_PATH}")

# ==============================================================================
# FUNÇÃO MESTRE (CHAMADA PELO HUB CENTRAL)
# ============================================================================
def executar_robo_produtividade_setor(pular_extracao=False):
    import sys 
    
    print("[PROGRESSO: 2]")
    print("Iniciando Robô de Produtividade...")
    
    hoje = date.today()
    
    primeiro_dia_mes_atual = hoje.replace(day=1)
    ultimo_dia_mes_passado = primeiro_dia_mes_atual - timedelta(days=1)
    mes_relatorio = ultimo_dia_mes_passado.month
    ano_relatorio = ultimo_dia_mes_passado.year
    
    primeiro_dia_mes_passado = ultimo_dia_mes_passado.replace(day=1)
    ultimo_dia_mes_retrasado = primeiro_dia_mes_passado - timedelta(days=1)
    mes_base = ultimo_dia_mes_retrasado.month
    ano_base = ultimo_dia_mes_retrasado.year
    
    nome_arquivo_base = f"Produtividade {mes_base:02d} - {ano_base} (preenchido).xlsx"
    nome_arquivo_saida = f"Produtividade {mes_relatorio:02d} - {ano_relatorio} (preenchido).xlsx"
    
    caminho_base = os.path.join(PASTA_PRODUTIVIDADE, nome_arquivo_base)
    
    print("-" * 50)
    print("🔍 VERIFICAÇÃO DE PRÉ-REQUISITOS")
    print(f"Arquivo base esperado: {nome_arquivo_base}")
    print(f"Arquivo que será gerado: {nome_arquivo_saida}")
    
    if not os.path.exists(PASTA_PRODUTIVIDADE):
        print(f"\n❌ ERRO FATAL: A pasta '{PASTA_PRODUTIVIDADE}' não existe!")
        print("Crie a pasta e tente novamente.")
        sys.exit(1) 
        
    if not os.path.exists(caminho_base):
        print(f"\n❌ ERRO FATAL: O arquivo base '{nome_arquivo_base}' não foi encontrado!")
        print("Coloque o arquivo do mês retrasado na pasta e tente novamente.")
        sys.exit(1) 
        
    if not pular_extracao:
        print("[PROGRESSO: 5]")
        print("✅ Pré-requisitos validados! Iniciando extração...\n")
        print("-" * 50)
        extrair_dados_sistemas()
    else:
        print("[PROGRESSO: 60]")
        print("⚠️ Opção 'Pular Extração' ativada. Usando arquivos já existentes na pasta...")
        print("-" * 50)
    
    print("[PROGRESSO: 70]")
    print("--- Executando Etapa 1: Renomear Arquivos ---")
    step_1_prepare_and_rename_reports(PASTA_PRODUTIVIDADE)
    
    print("[PROGRESSO: 80]")
    print("--- Executando Etapa 2: Processar Produtividade ---")
    main(nome_arquivo_base, nome_arquivo_saida)
    
    print("[PROGRESSO: 100]")
    print("✅ Robô de Produtividade finalizado com sucesso!")

if __name__ == "__main__":
    executar_robo_produtividade_setor()
