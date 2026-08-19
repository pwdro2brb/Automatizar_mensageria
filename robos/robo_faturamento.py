import os
import re
import time
import getpass
import unicodedata
import win32timezone
import pandas as pd
import json
from pathlib import Path
from PyPDF2 import PdfReader
from typing import Optional, Dict, List, Union, Tuple

# --- Selenium Imports ---
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    StaleElementReferenceException, TimeoutException, 
    ElementClickInterceptedException, ElementNotInteractableException, WebDriverException
)
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options 

# --- Pywinauto Imports ---
from pywinauto.application import Application
from pywinauto import Desktop
from pywinauto.timings import wait_until

import win32com.client as win32
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import numbers
import traceback
import sys

def obter_caminho_base():
    """Retorna o diretório real do projeto/executável."""
    if getattr(sys, 'frozen', False):
        # Rodando como executável compilado (.exe)
        return os.path.dirname(sys.executable)
    else:
        # Rodando via script Python (.py) no VS Code
        # Sobe um nível se este script estiver dentro da pasta /robos
        pasta_atual = os.path.dirname(os.path.abspath(__file__))
        return os.path.abspath(os.path.join(pasta_atual, ".."))

# Define o caminho absoluto correto
PASTA_BASE = obter_caminho_base()
CAMINHO_CONFIG_EMAILS = os.path.join(PASTA_BASE, "config_emails.json")
 
try:
    with open(CAMINHO_CONFIG_EMAILS, "r", encoding="utf-8") as f:
        CONFIG = json.load(f)
except FileNotFoundError:
    raise FileNotFoundError(
        f"Arquivo config_emails.json não foi encontrado em: {CAMINHO_CONFIG_EMAILS}"
    )

CONFIG_EMAILS = CONFIG["robo_faturamento"]

EMAILS_IGNORADOS = CONFIG_EMAILS["emails_ignorados"]

# ==============================================================================
# CONFIGURAÇÃO DE PASTAS DINÂMICAS
# ==============================================================================
sys.path.append(str(Path(__file__).parent.parent))
import config
from config import EMAIL_MRV, SENHA_MRV

PASTA_ARQUIVOS_RATEIO = Path(config.PASTA_ARQUIVOS) / "faturamento"
CNPJ_CORREIOS_FIXO = "34028316001509"   
DATE_RE = r"([0-3]?\d/[01]?\d/\d{4})"   


# ==============================================================================
# 1. FUNÇÃO: GERAR RASCUNHOS (REFATORADA COM JSON)
# ==============================================================================
def criar_rascunhos_correios():
    print("[PROGRESSO: 5]")
    print("Iniciando a criação de rascunhos no Outlook...")
    caminho_base = r"\\Bhz-fls-app1\mrvbh\Gerência Administrativa\Pública\NUCLEO DE CONTRATOS E APOIO A GESTÃO\CONTRATOS\Contratos Serviços\1. CORREIOS\2. Faturamento\2026"
    
    pastas_meses = [f for f in os.listdir(caminho_base) if os.path.isdir(os.path.join(caminho_base, f))]
    if not pastas_meses:
        print("Nenhuma pasta de mês encontrada no diretório.")
        return

    pastas_meses.sort()
    pasta_mes_recente = pastas_meses[-1]
    caminho_mes_recente = os.path.join(caminho_base, pasta_mes_recente)
    nome_mes = pasta_mes_recente.split("-")[-1].strip()

    print(f"Pasta mais recente encontrada: {pasta_mes_recente}")
    print("[PROGRESSO: 15]")

    # Extrai os dados parametrizados do JSON
    contatos_para = CONFIG_EMAILS["destinatarios_por_regional"]
    cc_padrao = CONFIG_EMAILS["cc_padrao"]
    cc_extra = CONFIG_EMAILS["cc_extra_regional"]
    regionais_extra = CONFIG_EMAILS["regionais_com_cc_extra"]

    agora = datetime.now()
    saudacao = "Bom dia" if agora.hour < 12 else "Boa tarde"
    prazo_rateio = agora + timedelta(hours=32)
    prazo_formatado = prazo_rateio.strftime("%d/%m/%Y às %H:%M")

    corpo_email = f"""
    <p style="font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #000000;">
        {saudacao}, Prezado(s)!<br><br>
        Segue em anexo o extrato dos Correios. O rateio deverá ser enviado até <b>{prazo_formatado}</b>.<br><br>
        Atenciosamente,
    </p>
    """

    outlook = win32.Dispatch('outlook.application')
    pastas_regionais = os.listdir(caminho_mes_recente)
    
    regionais_validas = [r for r in pastas_regionais if os.path.isdir(os.path.join(caminho_mes_recente, r)) and r.upper() != "BH"]
    total_regionais = len(regionais_validas)
    
    print("[PROGRESSO: 25]")
    
    for i, regional in enumerate(regionais_validas):
        caminho_regional = os.path.join(caminho_mes_recente, regional)

        print(f"Gerando rascunho para: {regional}...")
        mail = outlook.CreateItem(0)
        mail.To = contatos_para.get(regional, "")
        
        # Aplicação dinâmica das regras de cópia
        if regional in regionais_extra:
            mail.CC = f"{cc_padrao}; {cc_extra}"
        else:
            mail.CC = cc_padrao

        mail.Subject = f"RES: Extrato Correios - {regional} ({nome_mes})"

        arquivos_na_pasta = os.listdir(caminho_regional)
        for arquivo in arquivos_na_pasta:
            caminho_arquivo = os.path.join(caminho_regional, arquivo)
            if os.path.isfile(caminho_arquivo):
                mail.Attachments.Add(caminho_arquivo)

        mail.Display() 
        assinatura_outlook = mail.HTMLBody
        mail.HTMLBody = f"<html><body>{corpo_email}{assinatura_outlook}</body></html>"
        mail.Save()
        mail.Close(0)
        
        progresso_atual = 25 + int(((i + 1) / total_regionais) * 70)
        print(f"[PROGRESSO: {progresso_atual}]")

    print("[PROGRESSO: 100]")
    print("\nProcesso concluído! Verifique a pasta 'Rascunhos' no seu Outlook.")
    
# ==============================================================================
# 2. NOVA FUNÇÃO: FATURAMENTO END-TO-END (E-MAIL -> MRV PAG)
# ==============================================================================
def extrair_dados_email_inteligente(texto_email, caminho_correios):
    """Lê o corpo do e-mail e extrai os Centros de Custo e Valores (ou busca o valor pelo rastreio)"""
    dados = []
    rastreios_encontrados = []
    
    # 1. Isolar apenas a resposta mais recente (Corta o histórico do e-mail)
    marcadores_historico = ["De:", "From:", "________________________________", "-----Mensagem original-----", "-----Original Message-----"]
    texto_recente = texto_email
    for marcador in marcadores_historico:
        if marcador in texto_recente:
            texto_recente = texto_recente.split(marcador)[0]
            
    # 2. Analisar linha por linha
    linhas = texto_recente.splitlines()
    
    # Padrões Regex super flexíveis (Aceita com hífen, sem hífen, em tabela, etc.)
    # CC: 6 a 12 caracteres alfanuméricos (com letras e números) OU 10 a 12 números puros
    padrao_cc = re.compile(r'\b(?:(?=[A-Z0-9]*[A-Z])(?=[A-Z0-9]*[0-9])[A-Z0-9]{6,12}|\d{10,12})\b', re.IGNORECASE)
    # Valor: R$ 123,45 ou apenas 123,45
    padrao_valor = re.compile(r'(?:R\$?\s*)?(\d{1,3}(?:\.\d{3})*,\d{2})', re.IGNORECASE)
    # Rastreio: 2 letras + 9 números + 2 letras
    padrao_rastreio = re.compile(r'\b([A-Z]{2}\d{9}[A-Z]{2})\b', re.IGNORECASE)
    
    for linha in linhas:
        linha = linha.strip()
        if not linha: continue
        
        cc_match = padrao_cc.search(linha)
        val_match = padrao_valor.search(linha)
        rastreio_match = padrao_rastreio.search(linha)
        
        # Caso A: Tem CC e Valor na mesma linha (Tabela ou Texto)
        if cc_match and val_match:
            # CORREÇÃO AQUI: cc_match.group(0) em vez de group(1)
            cc = cc_match.group(0).upper()
            valor = float(val_match.group(1).replace('.', '').replace(',', '.'))
            dados.append({"COLETOR": cc, "VALOR": valor})
            continue # Já achou valor, não precisa procurar rastreio nessa linha
            
        # Caso B: Tem CC e Rastreio na mesma linha
        if cc_match and rastreio_match:
            # CORREÇÃO AQUI: cc_match.group(0) em vez de group(1)
            cc = cc_match.group(0).upper()
            rastreio = rastreio_match.group(1).upper()
            rastreios_encontrados.append((rastreio, cc))
           
    # 3. Se achou rastreios, busca os valores na planilha dos Correios
    if rastreios_encontrados and os.path.exists(caminho_correios):
        print(f"   -> {len(rastreios_encontrados)} rastreios encontrados. Buscando valores na planilha...")
        try:
            df_corr = pd.read_excel(caminho_correios, engine='openpyxl')
            df_str = df_corr.astype(str).apply(lambda x: x.str.upper())
            
            for rastreio, cc in rastreios_encontrados:
                mask = df_str.apply(lambda row: row.str.contains(rastreio).any(), axis=1)
                if mask.any():
                    idx = mask.idxmax()
                    row = df_corr.iloc[idx]
                    
                    valor_encontrado = 0.0
                    for val in row.values:
                        if isinstance(val, (int, float)):
                            valor_encontrado = float(val)
                            break
                        elif isinstance(val, str) and 'R$' in val:
                            try:
                                v = val.replace('R$', '').replace('.', '').replace(',', '.').strip()
                                valor_encontrado = float(v)
                                break
                            except: pass
                            
                    if valor_encontrado > 0:
                        dados.append({"COLETOR": cc, "VALOR": valor_encontrado})
                        print(f"      Rastreio {rastreio} -> Valor R$ {valor_encontrado} -> CC {cc}")
        except Exception as e:
            print(f"   -> Erro ao buscar rastreio na planilha: {e}")
            
    if dados:
        # Agrupar por CC caso a pessoa tenha mandado o mesmo CC duas vezes
        df = pd.DataFrame(dados)
        df = df.groupby("COLETOR", as_index=False)["VALOR"].sum()
        return df
    return None

def extrair_substituicao_cc(texto_email):
    """Lê o histórico do e-mail para descobrir quais CCs falharam e quais são os novos"""
    marcadores_historico = ["De:", "From:", "________________________________", "-----Mensagem original-----", "-----Original Message-----"]
    texto_recente = texto_email
    for marcador in marcadores_historico:
        if marcador in texto_recente:
            texto_recente = texto_recente.split(marcador)[0]
            
    padrao_cc = re.compile(r'\b(?:(?=[A-Z0-9]*[A-Z])(?=[A-Z0-9]*[0-9])[A-Z0-9]{6,12}|\d{10,12})\b', re.IGNORECASE)
    
    # Pega todos os CCs novos na resposta da regional (sem duplicatas)
    ccs_novos = list(dict.fromkeys([m.upper() for m in padrao_cc.findall(texto_recente)]))
    
    # Pega o histórico (o que você escreveu)
    historico = texto_email.replace(texto_recente, "")
    
    # Procura a linha onde você escreveu "substituto" e extrai os CCs inválidos de lá
    ccs_antigos = []
    for linha in historico.splitlines():
        if "substituto" in linha.lower():
            ccs_antigos.extend(padrao_cc.findall(linha))
            
    ccs_antigos = list(dict.fromkeys([m.upper() for m in ccs_antigos]))
    
    return ccs_antigos, ccs_novos

def atualizar_cc_rateio_pag(caminho_rateio_pag, ccs_antigos, ccs_novos):
    """Abre o RATEIO PAG existente, troca os CCs antigos pelos novos e salva"""
    df = pd.read_excel(caminho_rateio_pag, engine='openpyxl')
    sucesso = False
    
    # Pega os CCs que atualmente estão na planilha
    ccs_na_planilha = df['COLETOR'].astype(str).values
    
    # Cenário 1: Você pediu 2 substitutos e mandaram 2 novos (troca um por um)
    if len(ccs_antigos) == len(ccs_novos) and len(ccs_antigos) > 0:
        for antigo, novo in zip(ccs_antigos, ccs_novos):
            if antigo in ccs_na_planilha:
                df['COLETOR'] = df['COLETOR'].astype(str).replace(antigo, novo)
                sucesso = True
                
    # Cenário 2: Achou vários antigos no histórico, mas só mandaram 1 novo agora
    elif len(ccs_antigos) >= 1 and len(ccs_novos) >= 1:
        novo = ccs_novos[0]
        # Procura qual dos antigos do histórico realmente está na planilha precisando ser trocado
        for antigo in ccs_antigos:
            if antigo in ccs_na_planilha:
                df['COLETOR'] = df['COLETOR'].astype(str).replace(antigo, novo)
                sucesso = True
                break # Achou o culpado, faz a troca e para de procurar
                
    if sucesso:
        # Atualiza as colunas de tipo e operação para o novo CC
        df['TIPOCOLETOR'] = df['COLETOR'].apply(_tipo_de_coletor)
        df['OPERACAO'] = df['TIPOCOLETOR'].apply(lambda t: 10 if t == 'N' else '')
        
        # Salva por cima
        with pd.ExcelWriter(caminho_rateio_pag, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Planilha1', index=False)
            
        _formatar_planilha_final(caminho_rateio_pag, 'Planilha1')
        
    return sucesso


def executar_faturamento_completo():
    print("[PROGRESSO: 5]")
    print("Iniciando Faturamento Ponta a Ponta (E-mail -> MRV Pag)...")
    
    caminho_base = r"\\Bhz-fls-app1\mrvbh\Gerência Administrativa\Pública\NUCLEO DE CONTRATOS E APOIO A GESTÃO\CONTRATOS\Contratos Serviços\1. CORREIOS\2. Faturamento\2026"
    
    pastas_meses = [f for f in os.listdir(caminho_base) if os.path.isdir(os.path.join(caminho_base, f))]
    if not pastas_meses:
        print("Nenhuma pasta de mês encontrada no diretório da rede.")
        return
    pastas_meses.sort()
    pasta_mes_recente = pastas_meses[-1]
    caminho_mes_recente = os.path.join(caminho_base, pasta_mes_recente)
    
    print(f"Lendo e-mails do Outlook (Caixa de Entrada)...")
    outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(6) 
    mensagens = inbox.Items
    mensagens.Sort("[ReceivedTime]", True) 
    
    emails_processar = []
    
    print("[PROGRESSO: 15]")
    print("Analisando os 200 e-mails mais recentes...")
    
    contador = 0
    for msg in mensagens:
        # 43 é o código interno do Outlook para "MailItem" (E-mail normal)
        # Isso evita que ele tente ler relatórios de "Não foi possível entregar"
        if getattr(msg, "Class", 0) != 43:
            continue
            
        contador += 1
        if contador > 200:
            break 
            
        try:
            assunto = getattr(msg, "Subject", "")
            if not assunto:
                continue
                
            if "EXTRATO CORREIOS" in assunto.upper():
                print(f"\n🔍 Achou: {assunto}")
                
                try:
                    # Agora com o win32timezone importado, isso vai funcionar perfeitamente
                    data_msg = msg.ReceivedTime
                    data_python = datetime(data_msg.year, data_msg.month, data_msg.day)
                    dias_passados = (datetime.now() - data_python).days
                except Exception as e:
                    print(f"   ❌ Erro ao ler a data do e-mail: {e}")
                    continue
                    
                if dias_passados > 2:
                    print(f"   ❌ Ignorado (Muito antigo: {dias_passados} dias atrás)")
                    continue
                    
                try:
                    remetente = msg.SenderEmailAddress.lower()
                except:
                    remetente = "desconhecido"
                    
                if remetente in EMAILS_IGNORADOS:
                    print(f"   ❌ Ignorado (Remetente na lista de ignorados: {remetente})")
                    continue
                    
                try:
                    regional = assunto.split("-")[1].split("(")[0].strip()
                    emails_processar.append({
                        "mensagem": msg,
                        "regional": regional,
                        "assunto": assunto
                    })
                    print(f"   ✅ APROVADO! Regional identificada: {regional}")
                except Exception as e:
                    print(f"   ❌ Ignorado (Não consegui descobrir a regional pelo assunto)")
                    
        except Exception as e:
            continue

    if not emails_processar:
        print("[PROGRESSO: 100]")
        print("\nNenhum e-mail válido para processar.")
        return

    print(f"Encontrados {len(emails_processar)} e-mails para processar.")
    
    ARQUIVO_REGRAS_XLSX = PASTA_ARQUIVOS_RATEIO / "dados_puxados_preenchimento.xlsx"
    if not ARQUIVO_REGRAS_XLSX.exists():
        print(f"ERRO: Planilha de regras não encontrada em {ARQUIVO_REGRAS_XLSX}")
        return
    df_regras = pd.read_excel(ARQUIVO_REGRAS_XLSX, engine="openpyxl")

    driver = None 
    regionais_ja_tratadas = [] # <--- NOVA MEMÓRIA DO ROBÔ
    
    for i, item in enumerate(emails_processar):
        msg = item["mensagem"]
        regional = item["regional"]
        
        # Se já tratou essa regional neste ciclo, pula para o próximo e-mail
        if regional in regionais_ja_tratadas:
            print(f"⚠️ Regional {regional} já foi processada. Pulando e-mail duplicado.")
            continue

        print(f"\n" + "="*50)
        print(f"Processando Regional: {regional}") 
        
        caminho_regional_rede = os.path.join(caminho_mes_recente, regional)
        if not os.path.exists(caminho_regional_rede):
            print(f"⚠️ Pasta da regional não encontrada na rede: {caminho_regional_rede}")
            continue
            
        arquivos_rede = os.listdir(caminho_regional_rede)
        caminho_correios = None
        caminho_boleto_pdf = None
        
        for arq in arquivos_rede:
            if re.match(r'^\d{7}\.xlsx$', arq, re.IGNORECASE):
                caminho_correios = os.path.join(caminho_regional_rede, arq)
            elif arq.lower().endswith('.pdf'):
                caminho_boleto_pdf = os.path.join(caminho_regional_rede, arq)
                
        if not caminho_correios or not caminho_boleto_pdf:
            print(f"⚠️ Faltam arquivos (PDF ou Planilha dos Correios) na pasta da rede: {caminho_regional_rede}")
            continue

        caminho_rr = os.path.join(caminho_regional_rede, "Rateio Recebido.xlsx")
        caminho_rateio_pag = os.path.join(caminho_regional_rede, "RATEIO PAG.xlsx")
        
        pular_geracao_rateio = False

        # 🧠 INTELIGÊNCIA DE SUBSTITUIÇÃO DE CC
        if os.path.exists(caminho_rateio_pag):
            print("🔍 RATEIO PAG já existe na pasta. Verificando se é um e-mail de substituição...")
            ccs_antigos, ccs_novos = extrair_substituicao_cc(msg.Body)
            
            if ccs_antigos and ccs_novos:
                print(f"🔄 Substituição detectada! Trocando {ccs_antigos} por {ccs_novos}...")
                sucesso_troca = atualizar_cc_rateio_pag(caminho_rateio_pag, ccs_antigos, ccs_novos)
                
                if sucesso_troca:
                    print("✅ RATEIO PAG atualizado com o novo Centro de Custo!")
                    pular_geracao_rateio = True
                    
                    # SALVA O E-MAIL DE SUBSTITUIÇÃO COM DATA E HORA PARA NÃO SOBRESCREVER
                    agora_str = datetime.now().strftime("%d%m%Y_%H%M%S")
                    nome_substituto = f"Substituto_CC_{regional}_{agora_str}.msg"
                    msg.SaveAs(os.path.join(caminho_regional_rede, nome_substituto), 3)
                    print(f"✅ E-mail de substituição salvo como: {nome_substituto}")
                else:
                    print(f"⚠️ Os CCs antigos não foram encontrados na planilha. Vou regerar o rateio do zero.")
            else:
                print("ℹ️ Não é um e-mail de substituição claro. Regerando rateio...")

        # Se não for substituição, faz o processo normal de ler anexo/corpo e gerar do zero
        if not pular_geracao_rateio:
            df_email = extrair_dados_email_inteligente(msg.Body, caminho_correios)
            
            if df_email is not None and not df_email.empty:
                print("✅ Dados extraídos do corpo do e-mail com sucesso!")
                df_email.to_excel(caminho_rr, index=False)
                
                nome_comprovante = f"Comprovante_Aprovacao_{regional}.msg"
                msg.SaveAs(os.path.join(caminho_regional_rede, nome_comprovante), 3)
                print("✅ E-mail salvo como comprovante (.msg).")
            else:
                print("Procurando anexo no e-mail...")
                anexo_salvo = False
                for anexo in msg.Attachments:
                    nome_anexo = anexo.FileName.lower()
                    if ".xls" in nome_anexo:
                        if nome_anexo.endswith(".xls"):
                            caminho_temp = os.path.join(caminho_regional_rede, "temp_rateio.xls")
                            anexo.SaveAsFile(caminho_temp)
                            print("🔄 Convertendo arquivo .xls para .xlsx...")
                            try:
                                excel = win32.DispatchEx("Excel.Application")
                                excel.Visible = False
                                excel.DisplayAlerts = False
                                wb = excel.Workbooks.Open(caminho_temp)
                                wb.SaveAs(caminho_rr, FileFormat=51)
                                wb.Close()
                                excel.Quit()
                                os.remove(caminho_temp)
                                anexo_salvo = True
                                print("✅ Anexo convertido e salvo como 'Rateio Recebido.xlsx'.")
                            except Exception as e:
                                print(f"⚠️ Erro ao converter .xls para .xlsx: {e}")
                                try: excel.Quit() 
                                except: pass
                        else:
                            anexo.SaveAsFile(caminho_rr)
                            anexo_salvo = True
                            print("✅ Anexo salvo como 'Rateio Recebido.xlsx'.")
                        break
                        
                if not anexo_salvo:
                    print("⚠️ Nenhum dado no corpo e nenhum anexo Excel encontrado. Pulando regional.")
                    continue

            print("Gerando RATEIO PAG.xlsx...")
            try:
                gerar_rateio_pag(caminho_correios=caminho_correios, caminho_rr=caminho_rr, saida=caminho_rateio_pag, debug=False)
                print("✅ RATEIO PAG gerado com sucesso!")
            except Exception as e:
                print(f"⚠️ Erro ao gerar RATEIO PAG: {e}")
                continue

        print("Extraindo dados do Boleto PDF...")
        try:
            campos = extrair_campos_boleto(caminho_boleto_pdf)
            num_doc = campos["numero_documento"]
            cnpj_mrv = campos["cnpj_pagador"]
            valor_boleto = campos["valor_total_str"]
            
            if not num_doc or not cnpj_mrv or not valor_boleto:
                print("⚠️ Falha ao extrair dados essenciais do PDF.")
                continue
            print(f"📌 Dados: CNPJ MRV: {cnpj_mrv} | Valor: R$ {valor_boleto} | Nº Doc: {num_doc}")
        except Exception as e:
            print(f"⚠️ Erro ao ler PDF: {e}")
            continue

        if not driver:
            print("Abrindo MRV Pag...")
            chrome_options = Options()
            chrome_options.add_experimental_option("detach", True) 
            driver = webdriver.Chrome(options=chrome_options) 
            driver.get("https://mrvpag2.mrv.com.br/home")
            driver.maximize_window()
            wait_longo = WebDriverWait(driver, 180)
            wait = WebDriverWait(driver, 15)
            wait_rapido = WebDriverWait(driver, 2)
            
            print("Aguardando login...")
            wait.until(EC.presence_of_element_located((By.ID, "i0116"))).send_keys(EMAIL_MRV)
            click_anti_stale(wait, By.ID, "idSIButton9")
            wait.until(EC.presence_of_element_located((By.ID, "i0118"))).send_keys(SENHA_MRV)
            click_anti_stale(wait_longo, By.ID, "idSIButton9")
            print("!!! APROVE O MFA NO CELULAR !!!")
            click_anti_stale(wait_longo, By.ID, "idSIButton9") 
            
            fechar_mensagem = WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "mat-icon.btnCancelTest")))
            fechar_mensagem.click()
            print("Login concluído.")

        print("Verificando se o documento já foi lançado...")
        driver.get("https://mrvpag2.mrv.com.br/home")
        wait_overlays_to_hide(wait)
        
        try:
            dropdown_paginacao = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "mat-select[aria-label='Itens por página:']")))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", dropdown_paginacao)
            dropdown_paginacao.click()
            
            opcao_1000 = wait.until(EC.element_to_be_clickable((By.XPATH, "//mat-option[.//span[contains(text(), '1000')]]")))
            opcao_1000.click()
            time.sleep(3) 
        except Exception as e:
            print("Aviso: Não foi possível alterar a paginação para 1000.")

        xpath_doc = f"//td[contains(@class, 'cdk-column-documento') and contains(text(), '{num_doc}')]"
        documento_encontrado = driver.find_elements(By.XPATH, xpath_doc)
        
        if len(documento_encontrado) > 0:
            print(f"⚠️ Documento {num_doc} JÁ ESTÁ LANÇADO no MRV Pag. Pulando lançamento.")
        else:
            print(f"✅ Documento {num_doc} não encontrado. Iniciando lançamento...")
            try:
                _realizar_lancamento_mrvpag(driver, wait, wait_rapido, caminho_boleto_pdf, caminho_rateio_pag, campos, df_regras)
                print(f"🎉 Lançamento da regional {regional} preenchido com sucesso!")
                
                regionais_ja_tratadas.append(regional) # Salva na memória
                
                # --- PARADA DE SEGURANÇA ---
                print("\n🛑 PARADA DE SEGURANÇA: O robô preencheu os dados.")
                print("Por favor, confira na tela do Chrome e clique em SALVAR/CONFIRMAR manualmente.")
                print("Para lançar a próxima regional, rode o robô novamente no Hub.")
                break # <--- ISSO FAZ O ROBÔ PARAR AQUI E DEIXAR A TELA ABERTA PARA VOCÊ
                
            except Exception as e:
                print(f"❌ Erro ao lançar no MRV Pag: {e}")

        progresso = 15 + int(((i + 1) / len(emails_processar)) * 85)
        print(f"[PROGRESSO: {progresso}]")

    print("[PROGRESSO: 100]")
    print("\nFaturamento Ponta a Ponta finalizado com sucesso!")

def _realizar_lancamento_mrvpag(driver, wait, wait_rapido, caminho_boleto_pdf, caminho_planilha_rateio, campos, df_regras):
    """Função interna que executa os cliques do MRV Pag"""
    cnpj_correios = campos["cnpj_beneficiario"]
    num_doc       = campos["numero_documento"]
    vencimento    = campos["vencimento"]
    valor_boleto  = campos["valor_total_str"]
    cnpj_mrv      = campos["cnpj_pagador"]
    emissao_proc  = campos["data_processamento"]

    texto_completo_pdf = norm_text(read_pdf_text(caminho_boleto_pdf)).upper()
    norm_limpo = texto_completo_pdf.replace(".", "").replace("/", "").replace("-", "")

    ID_REGIONAL = None
    candidatos = []

    for index, linha in df_regras.iterrows():
        palavra_chave = str(linha.get("PALAVRA_CHAVE", "")).upper()
        if not palavra_chave or palavra_chave == "NAN": continue
        palavra_chave_limpa = palavra_chave.replace(".", "").replace("/", "").replace("-", "")
        if palavra_chave_limpa in norm_limpo:
            candidatos.append(linha)

    if len(candidatos) == 0:
        raise ValueError("Nenhuma 'PALAVRA_CHAVE' da planilha Excel foi encontrada no boleto.")
    elif len(candidatos) == 1:
        linha_escolhida = candidatos[0]
        ID_REGIONAL  = linha_escolhida["ID"]
        descr        = linha_escolhida["DESCR"]
        material_cod = str(linha_escolhida["material_cod"])
    else:
        resultado = determinar_id_por_valor(valor_boleto, cnpj_mrv, df_regras)
        if resultado is None:
            raise ValueError("Não foi possível determinar o ID regional pelo valor do boleto.")
        ID_REGIONAL  = resultado["ID"]
        descr        = resultado["DESCR"]
        material_cod = resultado["material_cod"]

    novo_protocolo = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.item-menu[routerlink='/protocolo'], a.item-menu[href='/protocolo']")))
    novo_protocolo.click()
    
    tipo_nota = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "mat-card a.pointer")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", tipo_nota)
    tipo_nota.click()
    
    tipo_de_documento = wait.until(EC.presence_of_element_located((By.XPATH, "//mat-expansion-panel-header[.//mat-panel-title[contains(normalize-space(),'Tipo de Documento')]]")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", tipo_de_documento)
    if tipo_de_documento.get_attribute("aria-expanded") == "false":
        wait.until(EC.element_to_be_clickable(tipo_de_documento)).click()
    
    select_el = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "mat-select[formcontrolname='frmTipoDoc'], mat-select[aria-label='Qual o tipo do documento'], #mat-select-2")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", select_el)
    select_el.click()

    wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".cdk-overlay-pane .mat-select-panel")))
    opcao = wait.until(EC.element_to_be_clickable((By.XPATH, "//mat-option[.//span[contains(@class,'mat-option-text') and normalize-space()='NF somente de Serviços']]")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", opcao)
    opcao.click()

    input_enviar_arquivo = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#frmFile")))
    input_enviar_arquivo.send_keys(caminho_boleto_pdf)

    btn_continuar = wait.until(EC.element_to_be_clickable((By.ID, "btnTipoDoc")))
    safe_click(driver, btn_continuar) 

    aprovador_select = wait.until(EC.element_to_be_clickable((By.XPATH, "//mat-select[@aria-label='APROVADOR' or @placeholder='APROVADOR']")))
    safe_click(driver, aprovador_select) 

    opcao_vanessa = wait.until(EC.element_to_be_clickable((By.XPATH, "//mat-option//span[normalize-space(.)='VANESSA DE BRITO RODRIGUES (VANESSA.BRODRIGUES)']/ancestor::mat-option")))
    safe_click(driver, opcao_vanessa) 

    cnpj_input = driver.find_element(By.CSS_SELECTOR, "input[placeholder='CNPJ DA EMPRESA MRV'], input[aria-label='CNPJ DA EMPRESA MRV']")
    cnpj_input.clear()
    cnpj_input.send_keys(cnpj_mrv)
    cnpj_input.send_keys(Keys.TAB)

    xpath_linha_mrv = "(//tr[contains(@class,'mat-row')][.//td[contains(normalize-space(.),'MRV')]])[1]"
    linha_mrv = wait.until(EC.presence_of_element_located((By.XPATH, xpath_linha_mrv)))
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", linha_mrv)

    try:
        wait.until(EC.element_to_be_clickable((By.XPATH, xpath_linha_mrv)))
    except TimeoutException:
        pass 

    try:
        try:
            ActionChains(driver).move_to_element(linha_mrv).pause(0.1).click(linha_mrv).perform()
        except (ElementClickInterceptedException, StaleElementReferenceException):
            primeiro_td = linha_mrv.find_element(By.XPATH, ".//td[1]")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", primeiro_td)
            try:
                ActionChains(driver).move_to_element(primeiro_td).pause(0.1).click(primeiro_td).perform()
            except Exception:
                driver.execute_script("arguments[0].click();", primeiro_td)
    except Exception as e:
        raise RuntimeError("Erro ao clicar na linha MRV.")

    wait_overlay_gone(driver, wait)
    wait_no_overlay(driver, wait)
    
    for _ in range(3):
        try:
            inp_num_doc = get_input_by_formcontrol(driver, wait, "frmNumDocumento")
            type_safely(driver, wait, inp_num_doc, num_doc)
            if (inp_num_doc.get_attribute("value") or "").strip() == num_doc: break
        except StaleElementReferenceException: continue

    wait_no_overlay(driver, wait)
    for _ in range(3):
        try:
            inp_cnpj_cor = get_input_by_formcontrol(driver, wait, "frmCnpjFornecedor")
            editable = ensure_enabled_and_editable(driver, inp_cnpj_cor, allow_force=True)
            if editable: type_safely(driver, wait, inp_cnpj_cor, cnpj_correios)
            else: js_set_value_and_dispatch(driver, inp_cnpj_cor, cnpj_correios)
            if (re.sub(r"\D", "", inp_cnpj_cor.get_attribute("value") or "")) == cnpj_correios: break
        except StaleElementReferenceException: continue

    wait_no_overlay(driver, wait)
    for _ in range(3):
        try:
            inp_emissao = get_input_by_formcontrol(driver, wait, "frmDtEmissao")
            try: click_with_fallback(driver, inp_emissao)
            except Exception: pass
            inp_emissao.send_keys(Keys.ESCAPE)
            type_safely(driver, wait, inp_emissao, emissao_proc)
            if (inp_emissao.get_attribute("value") or "").strip() == emissao_proc: break
        except StaleElementReferenceException: continue

    wait_no_overlay(driver, wait)
    for _ in range(3):
        try:
            inp_venc = get_input_by_formcontrol(driver, wait, "frmVencimento")
            ensure_enabled_and_editable(driver, inp_venc, allow_force=True)
            try: click_with_fallback(driver, inp_venc)
            except Exception: pass
            inp_venc.send_keys(Keys.ESCAPE)
            type_safely(driver, wait, inp_venc, vencimento)
            if (inp_venc.get_attribute("value") or "").strip() == vencimento: break
        except StaleElementReferenceException: continue

    wait_no_overlay(driver, wait)
    for _ in range(3):
        try:
            inp_valor = get_input_by_formcontrol(driver, wait, "frmValorTotalNf")
            click_with_fallback(driver, inp_valor)
            inp_valor.send_keys(Keys.CONTROL, 'a', Keys.DELETE)
            for ch in valor_boleto:
                inp_valor.send_keys(ch)
                time.sleep(0.01)
            if (inp_valor.get_attribute("value") or "").strip(): break
        except StaleElementReferenceException: continue

    click_ok_confirm(driver, wait_rapido, timeout=1, max_tentativas=1)

    campo_desc = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[formcontrolname="frmDescNota"]')))
    driver.execute_script("""
        arguments[0].value = arguments[1];
        arguments[0].dispatchEvent(new Event('input', {bubbles: true}));
        arguments[0].dispatchEvent(new Event('change', {bubbles: true}));
    """, campo_desc, descr)
    campo_desc.send_keys(Keys.TAB) 
    time.sleep(0.2) 

    driver.execute_script("""
        let btns = Array.from(document.querySelectorAll("button"));
        let btn = btns.find(b => b.textContent.includes("CONTINUAR") && !b.disabled && !b.classList.contains("mat-button-disabled"));
        if(btn) btn.click();
    """)

    locator_link = (By.XPATH, "//a[.//span[normalize-space(.)='CONTINUAR']]")
    wait.until(EC.presence_of_element_located(locator_link)) 
    
    driver.execute_script("""
        let links = Array.from(document.querySelectorAll("a"));
        let link = links.find(a => a.textContent.includes("CONTINUAR") && !a.classList.contains("mat-button-disabled"));
        if(link) link.click();
    """)

    btn_adicionar = wait_rapido.until(EC.element_to_be_clickable((By.XPATH, "//span[normalize-space(.)='Adicionar']/ancestor::button[1]")))
    btn_adicionar.click()

    preencher_codigo_material_ultima_linha(driver, wait_rapido, material_cod, timeout=10)
    click_pesquisar(driver, wait_rapido)

    wait_longo = WebDriverWait(driver, 40)
    locator_checkbox = (By.XPATH, "(//td[contains(@class,'mat-column-select')]//mat-checkbox)[1]")
    wait_longo.until(EC.presence_of_element_located(locator_checkbox))
    
    driver.execute_script("""
        let matCheckbox = document.querySelector("td.mat-column-select mat-checkbox");
        if (matCheckbox) {
            let label = matCheckbox.querySelector("label");
            if (label) {
                label.click();
            } else {
                matCheckbox.click();
            }
        }
    """)
    
    click_incluir_produtos(driver, wait_rapido)
    preencher_quantidade_e_valor(driver, wait_rapido, quantidade="1", valor_boleto=valor_boleto)
    abrir_select_justificativa(driver, wait_rapido)
    selecionar_opcao_justificativa_com_hover(driver, wait_rapido, texto_alvo="2 - Orientações do gestor/coordendor da área")
    click_continuar_proximo_ao_select(driver, wait_rapido)
    
    try:
        inputs_file = driver.find_elements(By.XPATH, "//input[@type='file']")
        if inputs_file:
            input_planilha = inputs_file[-1]
            driver.execute_script(
                "arguments[0].style.display = 'block'; "
                "arguments[0].style.visibility = 'visible'; "
                "arguments[0].style.opacity = 1;", 
                input_planilha
            )
            input_planilha.send_keys(caminho_planilha_rateio)
            time.sleep(2) 
    except Exception as e:
        print(f"⚠️ Erro ao tentar anexar silenciosamente: {e}")
        
    click_ok_confirm_repeatedly(driver, wait, max_clicks=3)

# ==============================================================================
# 3. FUNÇÕES AUXILIARES (MANTIDAS INTACTAS)
# ==============================================================================

def click_anti_stale(wait, by, seletor, tentativas=3):
    for _ in range(tentativas):
        try:
            elemento = wait.until(EC.element_to_be_clickable((by, seletor)))
            elemento.click()
            return True 
        except StaleElementReferenceException:
            time.sleep(0.5) 
    raise RuntimeError(f"O elemento {seletor} sumiu repetidas vezes.")

def scroll_center(driver, el):
    try: driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    except: pass

def wait_overlays_to_hide(wait):
    try: wait.until_not(lambda d: len(d.find_elements(By.CSS_SELECTOR, ".cdk-overlay-backdrop, .mat-progress-spinner")) > 0)
    except: pass

def safe_click(driver, element):
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
    try:
        element.click()
    except Exception:
        driver.execute_script("arguments[0].click();", element)

def safe_click_diferenciado(driver, element):
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
    try:
        element.click()
    except Exception:
        driver.execute_script("arguments[0].click();", element)

def fazer_upload_janela_windows(caminho_do_arquivo):
    from pywinauto import Desktop
    from pywinauto.timings import wait_until
    import time

    caminho_absoluto = str(Path(caminho_do_arquivo).resolve())
    titulos_possiveis = ["Abrir", "Open", "Abrir arquivo", "Escolher arquivo para carregar", "Escolher arquivo", "Select file", "File Upload", "Carregar"]
    classes_possiveis = ["#32770"]  

    janela = None
    backend_usado = None

    for backend in ["uia", "win32"]:
        if janela: break
        try:
            desktop = Desktop(backend=backend)
            for titulo in titulos_possiveis:
                try:
                    win = desktop.window(title=titulo)
                    if win.exists(timeout=1):
                        janela = win
                        backend_usado = backend
                        break
                except Exception: continue
        except Exception: continue

    if not janela:
        import re
        for backend in ["uia", "win32"]:
            if janela: break
            try:
                desktop = Desktop(backend=backend)
                for win in desktop.windows():
                    try:
                        titulo = win.window_text()
                        if any(t.lower() in titulo.lower() for t in ["abrir", "open", "upload", "file"]):
                            janela = win
                            backend_usado = backend
                            break
                    except Exception: continue
            except Exception: continue

    if not janela:
        for backend in ["uia", "win32"]:
            if janela: break
            try:
                desktop = Desktop(backend=backend)
                for classe in classes_possiveis:
                    try:
                        win = desktop.window(class_name=classe)
                        if win.exists(timeout=1):
                            janela = win
                            backend_usado = backend
                            break
                    except Exception: continue
            except Exception: continue

    if not janela:
        raise RuntimeError("⚠️ Janela de upload não encontrada!")

    try: janela.wait("ready", timeout=10)
    except Exception: time.sleep(1)  

    campo_preenchido = False
    if backend_usado == "uia":
        tentativas_campo = [
            lambda: janela.child_window(title="Nome do arquivo:", control_type="ComboBox").child_window(control_type="Edit"),
            lambda: janela.child_window(title="Nome do arquivo:", control_type="Edit"),
            lambda: janela.child_window(title="File name:", control_type="ComboBox").child_window(control_type="Edit"),
            lambda: janela.child_window(title="File name:", control_type="Edit"),
            lambda: janela.child_window(control_type="Edit", found_index=0),
        ]
    else:
        tentativas_campo = [
            lambda: janela.child_window(class_name="Edit", found_index=0),
            lambda: janela.child_window(title="Nome do arquivo:", class_name="ComboBoxEx32").child_window(class_name="Edit"),
            lambda: janela.child_window(class_name="ComboBoxEx32").child_window(class_name="Edit"),
        ]

    for get_campo in tentativas_campo:
        try:
            campo = get_campo()
            if campo.exists(timeout=2):
                campo.set_edit_text(caminho_absoluto)
                campo_preenchido = True
                break
        except Exception: continue

    if not campo_preenchido:
        try:
            from pywinauto.keyboard import send_keys
            janela.set_focus()
            time.sleep(0.3)
            send_keys("^a{DELETE}", pause=0.05)
            time.sleep(0.1)
            caminho_escaped = caminho_absoluto.replace("{", "{{").replace("}", "}}")
            send_keys(caminho_escaped, pause=0.02, with_spaces=True)
            campo_preenchido = True
        except Exception as e:
            raise RuntimeError(f"⚠️ Não consegui digitar o caminho do arquivo: {e}")

    time.sleep(0.5)
    botao_clicado = False
    nomes_botao = ["Abrir", "Open", "&Abrir", "&Open"]

    for nome_btn in nomes_botao:
        if botao_clicado: break
        try:
            if backend_usado == "uia": btn = janela.child_window(title=nome_btn, control_type="Button")
            else: btn = janela.child_window(title=nome_btn, class_name="Button")
            if btn.exists(timeout=2):
                btn.click()
                botao_clicado = True
        except Exception: continue

    if not botao_clicado:
        try:
            from pywinauto.keyboard import send_keys
            send_keys("{ENTER}")
            botao_clicado = True
        except Exception as e:
            raise RuntimeError(f"⚠️ Não consegui clicar no botão Abrir: {e}")

    for _ in range(20):
        try:
            if not janela.exists(timeout=0): return
        except Exception: return
        time.sleep(0.5)

def click_ok_confirm_repeatedly(driver, wait, max_clicks=5):    
    clicks = 0
    locators = [
        (By.ID, "btnTipoDoc"),
        (By.XPATH, "//button[contains(@class,'confirm') and normalize-space(.)='OK']")
    ]
    for _ in range(max_clicks):
        wait_overlays_to_hide(wait)
        btn = None
        for by, sel in locators:
            try:
                elem = driver.find_element(by, sel)
                if elem.is_displayed(): btn = elem; break
            except: pass
        if not btn: break
        
        safe_click(driver, btn)
        clicks += 1
        time.sleep(1)
    return clicks

def click_anexar_planilha(driver, wait, caminho_planilha=None):
    wait_overlays_to_hide(wait)
    locator_btn = (By.XPATH, "//button[.//span[normalize-space(.)='ANEXAR PLANILHA'] and not(@disabled) and not(contains(@class,'mat-button-disabled'))]")
    try: btn = wait.until(EC.presence_of_element_located(locator_btn))
    except TimeoutException: raise RuntimeError("Botão 'ANEXAR PLANILHA' não encontrado.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
    time.sleep(0.05)

    try: wait.until(EC.element_to_be_clickable(locator_btn))
    except TimeoutException: pass

    try:
        if not is_center_clickable_js(driver, btn):
            driver.execute_script("window.scrollBy(0, -80);")
            time.sleep(0.05)
        btn.click()
    except Exception:
        try: ActionChains(driver).move_to_element(btn).pause(0.05).click().perform()
        except Exception:
            try: driver.execute_script("arguments[0].click();", btn)
            except Exception as e3: raise RuntimeError(f"Falha ao clicar em 'ANEXAR PLANILHA': {repr(e3)}")

    if caminho_planilha:
        try: input_file = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='file']")))
        except TimeoutException: return
        caminho_planilha = str(Path(caminho_planilha).resolve())
        try: input_file.send_keys(caminho_planilha)
        except Exception:
            try:
                driver.execute_script("arguments[0].style.display='block'; arguments[0].style.visibility='visible';", input_file)
                time.sleep(0.05)
                input_file.send_keys(caminho_planilha)
            except Exception as e2: raise RuntimeError(f"Falha ao enviar arquivo para input[type=file]: {repr(e2)}")

def wait_overlay_gone(driver, wait, timeout=40):
    try:
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((
                By.CSS_SELECTOR,
                ".cdk-overlay-backdrop.cdk-overlay-backdrop-showing, .mat-progress-bar, .mat-spinner, .ngx-spinner-overlay"
            ))
        )
    except TimeoutException: pass

def get_visible_input(driver, formcontrol: str):
    candidates = driver.find_elements(By.CSS_SELECTOR, f"input[formcontrolname='{formcontrol}']")
    visibles = [el for el in candidates if el.is_displayed() and el.is_enabled()]
    if not visibles: raise TimeoutException(f"Input visível '{formcontrol}' não encontrado.")
    return visibles[0]

def focus_input(driver, wait, el):
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        wait.until(lambda d: el.is_displayed() and el.is_enabled())
        try: el.click()
        except ElementClickInterceptedException:
            container = el.find_element(By.XPATH, "./ancestor::mat-form-field//div[contains(@class,'mat-form-field-flex')]")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", container)
            try: ActionChains(driver).move_to_element(container).pause(0.05).click(container).perform()
            except Exception: driver.execute_script("arguments[0].click();", container)
    except StaleElementReferenceException: pass

def clear_and_type(el, text: str):
    driver = el.parent  
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        time.sleep(0.15)
    except Exception: pass

    try: el.click()
    except Exception:
        try: driver.execute_script("arguments[0].click(); arguments[0].focus();", el)
        except Exception: pass

    try:
        el.send_keys(Keys.CONTROL, 'a', Keys.DELETE)
        time.sleep(0.1)
        if text: el.send_keys(str(text))
    except Exception:
        driver.execute_script("""
            const el = arguments[0], val = arguments[1] ?? '';
            el.value = '';
            el.dispatchEvent(new Event('input', {bubbles: true}));
            el.value = val;
            el.dispatchEvent(new Event('input', {bubbles: true}));
            el.dispatchEvent(new Event('change', {bubbles: true}));
        """, el, text or '')

def fill_input(driver, wait, formcontrol: str, value: str, numeric=False):
    wait_overlay_gone(driver, wait)
    el = get_visible_input(driver, formcontrol)
    focus_input(driver, wait, el)
    if numeric and value is not None: value = re.sub(r'\D', '', str(value))
    clear_and_type(el, value)
    try: WebDriverWait(driver, 5).until(lambda d: (el.get_attribute("value") or "") != "")
    except Exception: pass

def somente_digitos(s: str) -> str:
    return re.sub(r"\D", "", s or "")

def norm_text(s: str) -> str:
    s = unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode("ASCII")
    s = s.replace("\xa0", " ")
    return re.sub(r"[ \t]+", " ", s)

def read_pdf_text(pdf_path: str) -> str:
    reader = PdfReader(pdf_path)
    txt = [p.extract_text() or "" for p in reader.pages]
    texto = "\n".join(txt)
    if not texto.strip(): raise ValueError("PDF sem texto. Use OCR.")
    return texto

def extrair_cnpj_pagador(text_norm: str) -> Optional[str]:
    padrao_cnpj = r"\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b"
    cnpjs_encontrados = re.findall(padrao_cnpj, text_norm)
    cnpjs_correios = ["34.028.316/0015-09", "34.028.316/0001-03"]
    for cnpj in cnpjs_encontrados:
        if cnpj not in cnpjs_correios: return cnpj
    return None

def extrair_numero_documento_7d(text_norm: str) -> Optional[str]:
    mdoc = re.search(r"DOCUMENTO.{0,100}(\d{7})", text_norm, flags=re.I | re.S)
    return mdoc.group(1) if mdoc else None

def extrair_valor_total(text_norm: str) -> Optional[str]:
    anchor = r"(?i)VALOR\s*(?:DO\s*)?DOCUMENTO(?:\s*\(R\$\))?"
    money_re = r"(?:R\$\s*)?(\d{1,3}(?:\.\d{3})*,\d{2})"
    lines = text_norm.splitlines()
    for i, line in enumerate(lines):
        if re.search(anchor, line, flags=re.I):
            for j in (i, i+1, i+2): 
                if 0 <= j < len(lines):
                    m = re.search(money_re, lines[j], flags=re.I)
                    if m: return m.group(1)
    return None

def extrair_datas_correios(text_norm: str) -> dict:
    from datetime import datetime
    todas_datas_str = re.findall(r"\d{2}/\d{2}/\d{4}", text_norm)
    vistas = set()
    unicas_str = []
    for d in todas_datas_str:
        if d not in vistas:
            vistas.add(d)
            unicas_str.append(d)

    datas_parsed = []
    for d_str in unicas_str:
        try:
            dt = datetime.strptime(d_str, "%d/%m/%Y")
            datas_parsed.append((d_str, dt))
        except ValueError: continue  

    datas_parsed.sort(key=lambda x: x[1])
    if not datas_parsed: return {"emissao": "", "vencimento": ""}

    vencimento = datas_parsed[-1][0]  
    if len(datas_parsed) >= 3: emissao = datas_parsed[1][0]
    elif len(datas_parsed) == 2: emissao = datas_parsed[0][0]
    else: emissao = datas_parsed[0][0]

    dt_emissao = datetime.strptime(emissao, "%d/%m/%Y")
    dt_vencimento = datetime.strptime(vencimento, "%d/%m/%Y")
    if dt_vencimento < dt_emissao: emissao, vencimento = vencimento, emissao

    return {"emissao": emissao, "vencimento": vencimento}

def extrair_campos_boleto(pdf_path: str) -> Dict[str, Optional[str]]:
    texto = read_pdf_text(pdf_path)
    norm  = norm_text(texto)
    num_doc = extrair_numero_documento_7d(norm)
    datas = extrair_datas_correios(norm) 
    valor_total_str = extrair_valor_total(norm)
    cnpj_pagador_extraido = extrair_cnpj_pagador(norm)
    cnpj_pag = somente_digitos(cnpj_pagador_extraido) if cnpj_pagador_extraido else None
    
    return {
        "numero_documento": num_doc,
        "cnpj_pagador": cnpj_pag,
        "cnpj_beneficiario": CNPJ_CORREIOS_FIXO,
        "data_processamento": datas["emissao"],
        "vencimento": datas["vencimento"], 
        "valor_total_str": valor_total_str     
    }

def determinar_id_por_valor(valor_str: str, cnpj_pagador: str, df: pd.DataFrame) -> dict:
    valor_float = float(valor_str.replace(".", "").replace(",", "."))
    cnpj_limpo = cnpj_pagador.replace(".", "").replace("/", "").replace("-", "")
    df["PALAVRA_CHAVE_LIMPA"] = df["PALAVRA_CHAVE"].astype(str).str.replace(".", "", regex=False).str.replace("/", "", regex=False).str.replace("-", "", regex=False).str.upper()
    linhas_cnpj = df[df["PALAVRA_CHAVE_LIMPA"] == cnpj_limpo]

    if linhas_cnpj.empty: return None
    if len(linhas_cnpj) == 1:
        linha = linhas_cnpj.iloc[0]
        return {"ID": linha["ID"], "DESCR": linha["DESCR"], "material_cod": str(linha["material_cod"])}

    if valor_float < 400.00: id_alvo = 2
    elif valor_float < 2000.00: id_alvo = 7
    elif valor_float >= 40000.00: id_alvo = 8
    else: raise ValueError(f"⚠️ Valor R$ {valor_str} não se encaixa nas faixas.")

    linha_alvo = linhas_cnpj[linhas_cnpj["ID"] == id_alvo]
    if linha_alvo.empty: raise ValueError(f"⚠️ ID {id_alvo} não encontrado.")

    linha = linha_alvo.iloc[0]
    return {"ID": linha["ID"], "DESCR": linha["DESCR"], "material_cod": str(linha["material_cod"])}

def click_continuar_proximo_ao_select(driver, wait):
    container = driver.find_element(By.XPATH, "//mat-form-field[.//mat-select[@formcontrolname='justificativa']]")
    locator_local = (By.XPATH, ".//following::a[.//span[normalize-space(.)='CONTINUAR'] and (not(@aria-disabled) or @aria-disabled='false') and not(contains(@class,'mat-button-disabled'))][1]")
    try: link = container.find_element(*locator_local)
    except Exception: raise RuntimeError("Não achei o CONTINUAR referente a este passo.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", link)
    time.sleep(0.05)
    try: wait.until(EC.element_to_be_clickable((By.XPATH, "//a[.//span[normalize-space(.)='CONTINUAR']]")))
    except TimeoutException: pass

    try: link.click()
    except Exception:
        try: ActionChains(driver).move_to_element(link).pause(0.05).click().perform()
        except Exception: driver.execute_script("arguments[0].click();", link)

def preencher_codigo_material_ultima_linha(driver, wait, material_cod, timeout=3):
    xpath_all = ("//input[(@formcontrolname='frmCodigoMaterial' or @name='codigoMaterial' or @placeholder='CÓDIGO DO MATERIAL') and not(@disabled)]")
    end = time.time() + timeout
    visiveis = []
    while time.time() < end:
        elems = driver.find_elements(By.XPATH, xpath_all)
        visiveis = [e for e in elems if e.is_displayed()]
        if visiveis: break
        time.sleep(0.2)

    if not visiveis: raise RuntimeError("Nenhum campo 'CÓDIGO DO MATERIAL' visível encontrado.")
    alvo = visiveis[-1]
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", alvo)
    time.sleep(0.05)
    try: wait.until(EC.element_to_be_clickable((By.XPATH, xpath_all)))
    except TimeoutException: pass

    try: alvo.click()
    except Exception: driver.execute_script("arguments[0].focus();", alvo)
    clear_and_type(alvo, material_cod)

def click_incluir_produtos(driver, wait):
    wait_overlays_to_hide(wait)
    locator_btn = (By.XPATH, "//mat-action-row//button[.//span[normalize-space(.)='INCLUIR PRODUTO(S)'] and not(@disabled) and not(contains(@class,'mat-button-disabled'))]")
    try: btn = wait.until(EC.presence_of_element_located(locator_btn))
    except TimeoutException: raise RuntimeError("Botão 'INCLUIR PRODUTO(S)' não encontrado.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
    time.sleep(0.05)
    try: wait.until(EC.element_to_be_clickable(locator_btn))
    except TimeoutException: pass

    try:
        if not is_center_clickable_js(driver, btn):
            driver.execute_script("window.scrollBy(0, -80);")
            time.sleep(0.05)
        btn.click()
        return
    except Exception: pass

    try:
        ActionChains(driver).move_to_element(btn).pause(0.05).click().perform()
        return
    except Exception: pass

    try:
        driver.execute_script("arguments[0].click();", btn)
        return
    except Exception as e3: raise RuntimeError(f"Falha ao clicar em 'INCLUIR PRODUTO(S)': {repr(e3)}")

def abrir_select_justificativa(driver, wait):
    wait_overlays_to_hide(wait)
    loc_select = (By.CSS_SELECTOR, "mat-select[formcontrolname='justificativa']")
    try: sel = wait.until(EC.presence_of_element_located(loc_select))
    except TimeoutException:
        try: sel = wait.until(EC.presence_of_element_located((By.XPATH, "//mat-select[@formcontrolname='justificativa' or @aria-label='Por quê o Pedido não foi criado antes da emissão da Nota Fiscal?' or @placeholder='Por quê o Pedido não foi criado antes da emissão da Nota Fiscal?']")))
        except TimeoutException: raise RuntimeError("Campo select 'Justificativa' não encontrado.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", sel)
    time.sleep(0.05)
    try: sel.click()
    except Exception:
        try:
            trigger = sel.find_element(By.CSS_SELECTOR, ".mat-select-trigger")
            trigger.click()
        except Exception:
            try: driver.execute_script("arguments[0].click();", sel)
            except Exception as e3: raise RuntimeError(f"Falha ao abrir o select 'Justificativa': {repr(e3)}")

    try: wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".cdk-overlay-pane .mat-select-panel")))
    except TimeoutException: raise RuntimeError("Painel de opções do 'Justificativa' não apareceu.")

def verificar_textos_na_tabela(driver, wait, textos, timeout=3):
    locator_container = (By.CSS_SELECTOR, "div.table-container")
    try: container = wait.until(EC.presence_of_element_located(locator_container))
    except TimeoutException: raise RuntimeError("Não encontrei o container da tabela.")

    conteudo = " ".join(container.text.split())
    encontrados = [t for t in textos if t and t in conteudo]
    faltando = [t for t in textos if t and t not in conteudo]
    return encontrados, faltando, conteudo

def click_ok_confirm(driver, wait, timeout=3, max_tentativas=3):
    cliques = 0
    for tentativa in range(max_tentativas):
        locator_dialog = (By.CSS_SELECTOR, "mat-dialog-container")
        try:
            WebDriverWait(driver, timeout if tentativa == 0 else 5, poll_frequency=0.3, ignored_exceptions=[StaleElementReferenceException]).until(EC.presence_of_element_located(locator_dialog))
        except TimeoutException:
            if cliques == 0: return 0
            else: return cliques

        locators_btn = [
            (By.CSS_SELECTOR, "mat-dialog-container button#btnTipoDoc"),
            (By.CSS_SELECTOR, "mat-dialog-container button.confirm"),
            (By.XPATH, "//mat-dialog-container//button[normalize-space()='OK']"),
        ]

        btn = None
        for by, sel in locators_btn:
            try:
                elementos = driver.find_elements(by, sel)
                visiveis = [e for e in elementos if _is_displayed_safe(e)]
                if visiveis:
                    btn = visiveis[-1]  
                    break
            except Exception: continue

        if btn is None:
            try:
                btn = WebDriverWait(driver, 5, poll_frequency=0.2, ignored_exceptions=[StaleElementReferenceException]).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "mat-dialog-container button#btnTipoDoc")))
            except TimeoutException:
                if cliques > 0: return cliques
                continue

        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            time.sleep(0.1)
        except StaleElementReferenceException: continue  

        metodo = _clicar_botao_ok(driver, btn)
        if metodo: cliques += 1
        else: continue

        try: WebDriverWait(driver, 5, poll_frequency=0.2).until(_dialog_desapareceu())
        except TimeoutException: continue
        time.sleep(0.5)
    return cliques

def _is_displayed_safe(element):
    try: return element.is_displayed()
    except (StaleElementReferenceException, Exception): return False

def _clicar_botao_ok(driver, btn):
    try:
        btn.click()
        return "clique direto"
    except (ElementClickInterceptedException, ElementNotInteractableException): pass
    except StaleElementReferenceException: return None

    try:
        ActionChains(driver).move_to_element(btn).pause(0.1).click().perform()
        return "ActionChains"
    except Exception: pass

    try:
        driver.execute_script("arguments[0].click();", btn)
        return "JavaScript"
    except Exception: pass
    return None

class _dialog_desapareceu:
    def __call__(self, driver):
        dialogs = driver.find_elements(By.CSS_SELECTOR, "mat-dialog-container")
        visiveis = [d for d in dialogs if _is_displayed_safe(d)]
        return len(visiveis) == 0
    
def wait_no_overlay(driver, wait):
    try: wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".cdk-overlay-backdrop.cdk-overlay-backdrop-showing, .mat-progress-spinner, .mat-progress-bar")))
    except TimeoutException: pass

def js_set_value_and_dispatch(driver, el, value: str):
    driver.execute_script("""
        const el = arguments[0], v = arguments[1];
        el.focus();
        const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value').set;
        setter.call(el, v);
        el.dispatchEvent(new Event('input', { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
        el.blur();
    """, el, value)

def get_input_by_formcontrol(driver, wait, formcontrol):
    sel = f"input[formcontrolname='{formcontrol}']"
    el = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, sel)))
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, sel)))
    scroll_center(driver, el)
    return el    

def esperar_transicao_apos_primeiro(wait, btn_clique=None, timeout=40):
    if btn_clique is not None:
        try:
            wait.until(EC.staleness_of(btn_clique))
            return
        except TimeoutException: pass
        try:
            wait.until(EC.invisibility_of_element(btn_clique))
            return
        except TimeoutException: pass
    time.sleep(0.3)
    wait_overlays_to_hide(wait)

def click_with_fallback(driver, el):
    try: el.click()
    except (ElementClickInterceptedException, ElementNotInteractableException): driver.execute_script("arguments[0].click();", el)

def type_safely(driver, wait, el, value: str):
    try:
        click_with_fallback(driver, el)
        el.send_keys(Keys.CONTROL, 'a', Keys.DELETE)
        if value is not None: el.send_keys(value)
        time.sleep(0.05)
        v = el.get_attribute("value") or ""
        if v.strip() != (value or "").strip(): js_set_value_and_dispatch(driver, el, value or "")
    except StaleElementReferenceException: raise

def is_center_clickable_js(driver, el):
    try:
        return driver.execute_script("""
            const el = arguments[0];
            const r = el.getBoundingClientRect();
            if (r.width === 0 || r.height === 0) return false;
            const cx = r.left + r.width/2, cy = r.top + r.height/2;
            const e = document.elementFromPoint(cx, cy);
            return e && (e === el || el.contains(e));
        """, el)
    except Exception: return False

def debug_dump(driver, prefix):
    try: driver.save_screenshot(f"{prefix}.png")
    except: pass
    try:
        with open(f"{prefix}.html","w",encoding="utf-8") as f: f.write(driver.page_source)
    except: pass

def ensure_enabled_and_editable(driver, el, allow_force=False):
    readonly = el.get_attribute("readonly")
    disabled = el.get_attribute("disabled")
    if readonly is not None or disabled is not None:
        if not allow_force: return False
        driver.execute_script("arguments[0].removeAttribute('readonly'); arguments[0].removeAttribute('disabled');", el)
    return True

def click_primeiro_continuar(driver, wait, campo_desc_css='input[formcontrolname="frmDescNota"]'):
    try:
        campo_desc = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, campo_desc_css)))
        campo_desc.send_keys(Keys.TAB)
        time.sleep(0.1)
    except TimeoutException: pass

    wait_overlays_to_hide(wait)
    candidatos = driver.find_elements(By.XPATH, "//button[.//span[normalize-space(.)='CONTINUAR']]")
    if not candidatos: raise RuntimeError("Nenhum <button> CONTINUAR encontrado.")

    alvo = None
    for btn in candidatos:
        try:
            vis = btn.is_displayed()
            hab = btn.is_enabled() and btn.get_attribute("disabled") in (None, "", "false")
            aria = btn.get_attribute("aria-disabled")
            if vis and hab and (aria in (None, "", "false")) and "mat-button-disabled" not in (btn.get_attribute("class") or ""):
                alvo = btn
                break
        except Exception: continue

    if not alvo:
        for btn in candidatos:
            try:
                if btn.is_displayed():
                    alvo = btn
                    break
            except Exception: continue

    if not alvo: raise RuntimeError("Nenhum <button> CONTINUAR visível.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", alvo)
    time.sleep(0.05)
    try: wait.until(EC.element_to_be_clickable((By.XPATH, "//button[.//span[normalize-space(.)='CONTINUAR']]")))
    except TimeoutException: pass

    try:
        if not is_center_clickable_js(driver, alvo):
            driver.execute_script("window.scrollBy(0, -80);")
            time.sleep(0.05)
        alvo.click()
        return alvo  
    except Exception: pass

    try:
        ActionChains(driver).move_to_element(alvo).pause(0.05).click().perform()
        return alvo
    except Exception: pass

    try:
        driver.execute_script("arguments[0].click();", alvo)
        return alvo
    except Exception as e3: raise RuntimeError(f"Falha ao clicar no primeiro CONTINUAR: {repr(e3)}")

def click_segundo_continuar(driver, wait):
    wait_overlays_to_hide(wait)
    locator_link = (By.XPATH, "//a[.//span[normalize-space(.)='CONTINUAR'] and (not(@aria-disabled) or @aria-disabled='false') and not(contains(@class,'mat-button-disabled'))]")
    try: wait.until(EC.presence_of_element_located(locator_link))
    except TimeoutException: raise RuntimeError("Nenhum <a> CONTINUAR presente após o primeiro clique.")

    links = driver.find_elements(*locator_link)
    if not links: raise RuntimeError("Nenhum <a> CONTINUAR encontrado (filtro).")

    alvo = None
    for a in links:
        try:
            vis = a.is_displayed()
            hab = a.is_enabled()
            aria = a.get_attribute("aria-disabled")
            cls = a.get_attribute("class") or ""
            if vis and hab and (aria in (None, "", "false")) and "mat-button-disabled" not in cls:
                alvo = a
                break
        except Exception: continue

    if not alvo:
        for a in links:
            try:
                if a.is_displayed():
                    alvo = a
                    break
            except Exception: continue

    if not alvo: raise RuntimeError("Nenhum <a> CONTINUAR visível.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", alvo)
    time.sleep(0.05)
    try: wait.until(EC.element_to_be_clickable(locator_link))
    except TimeoutException: pass

    try:
        if not is_center_clickable_js(driver, alvo):
            driver.execute_script("window.scrollBy(0, -80);")
            time.sleep(0.05)
        alvo.click()
        return
    except Exception: pass

    try:
        ActionChains(driver).move_to_element(alvo).pause(0.05).click().perform()
        return
    except Exception: pass

    try:
        driver.execute_script("arguments[0].click();", alvo)
        return
    except Exception as e3: raise RuntimeError(f"Falha ao clicar no segundo CONTINUAR (<a>): {repr(e3)}")

def click_pesquisar(driver, wait):
    wait_overlays_to_hide(wait)
    locator_btn = (By.XPATH, "//button[.//span[normalize-space(.)='Pesquisar'] and not(@disabled) and not(contains(@class,'mat-button-disabled'))]")
    try: btn = wait.until(EC.presence_of_element_located(locator_btn))
    except TimeoutException: raise RuntimeError("Botão 'Pesquisar' não encontrado na tela.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
    time.sleep(0.05)
    try: wait.until(EC.element_to_be_clickable(locator_btn))
    except TimeoutException: pass

    try:
        if not is_center_clickable_js(driver, btn):
            driver.execute_script("window.scrollBy(0, -80);")
            time.sleep(0.05)
        btn.click()
        return
    except Exception: pass

    try:
        ActionChains(driver).move_to_element(btn).pause(0.05).click().perform()
        return
    except Exception: pass

    try:
        driver.execute_script("arguments[0].click();", btn)
        return
    except Exception as e3: raise RuntimeError(f"Falha ao clicar no botão 'Pesquisar': {repr(e3)}")

def _aguardar_checkbox_interagivel(driver, timeout=40):
    locator_checkbox = (By.XPATH, "(//td[contains(@class,'mat-column-select')]//mat-checkbox//input[@type='checkbox'])[1]")
    wait = WebDriverWait(driver, timeout, poll_frequency=0.3, ignored_exceptions=[StaleElementReferenceException])
    try: cb_input = wait.until(EC.presence_of_element_located(locator_checkbox))
    except TimeoutException: raise RuntimeError(f"Checkbox não apareceu no DOM após {timeout}s.")

    try: cb_input = wait.until(EC.visibility_of_element_located(locator_checkbox))
    except TimeoutException:
        locator_mat_checkbox = (By.XPATH, "(//td[contains(@class,'mat-column-select')]//mat-checkbox)[1]")
        try: wait.until(EC.visibility_of_element_located(locator_mat_checkbox))
        except TimeoutException: raise RuntimeError(f"mat-checkbox não ficou visível após {timeout}s.")

    _aguardar_posicao_estavel(driver, cb_input, tentativas=5, intervalo=0.3)
    return cb_input

def _aguardar_posicao_estavel(driver, elemento, tentativas=5, intervalo=0.3):
    pos_anterior = None
    for _ in range(tentativas):
        try:
            rect = driver.execute_script(
                "var r = arguments[0].getBoundingClientRect();"
                "return {top: r.top, left: r.left, width: r.width, height: r.height};",
                elemento,
            )
        except StaleElementReferenceException:
            time.sleep(intervalo)
            continue

        if pos_anterior and rect == pos_anterior: return  
        pos_anterior = rect
        time.sleep(intervalo)

def _obter_label_do_checkbox(driver, cb_input):
    try:
        input_id = cb_input.get_attribute("id")
        if input_id: return driver.find_element(By.XPATH, f"//label[@for='{input_id}']")
    except Exception: pass

    try: return driver.find_element(By.XPATH, "(//td[contains(@class,'mat-column-select')]//mat-checkbox//label)[1]")
    except Exception: return None

def _clicar_com_fallback(driver, alvo, descricao="elemento"):
    try:
        alvo.click()
        return "Clique direto"
    except (ElementClickInterceptedException, Exception): pass

    try:
        ActionChains(driver).move_to_element(alvo).pause(0.1).click().perform()
        return "ActionChains"
    except Exception: pass

    try:
        driver.execute_script("arguments[0].click();", alvo)
        return "Clique via JS"
    except Exception as e: raise RuntimeError(f"Todas as estratégias de clique falharam para {descricao}: {repr(e)}")

def selecionar_primeira_linha_checkbox(driver, wait, timeout=40, textos_para_verificar=None, exigir_todos=False, clicar_mesmo_se_faltar=False):
    wait_overlays_to_hide(wait)
    resultado_textos = {"verificacao_feita": False, "encontrados": [], "faltando": []}

    if textos_para_verificar:
        encontrados, faltando, _conteudo = verificar_textos_na_tabela(driver, wait, textos_para_verificar, timeout=timeout)
        resultado_textos.update({"verificacao_feita": True, "encontrados": encontrados, "faltando": faltando})
        if faltando and exigir_todos and not clicar_mesmo_se_faltar: raise RuntimeError(f"Textos faltando na tabela: {faltando} | Encontrados: {encontrados}")
        if faltando and not clicar_mesmo_se_faltar: return {**resultado_textos, "clicou": False, "motivo": f"Faltando textos: {faltando}"}

    try: cb_input = _aguardar_checkbox_interagivel(driver, timeout=timeout)
    except RuntimeError: raise

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cb_input)
    time.sleep(0.15)
    label = _obter_label_do_checkbox(driver, cb_input)
    alvo = label if label else cb_input

    try:
        if not is_center_clickable_js(driver, alvo):
            driver.execute_script("window.scrollBy(0, -100);")
            time.sleep(0.15)
    except Exception: pass  

    metodo = _clicar_com_fallback(driver, alvo, descricao="checkbox primeira linha")
    return {**resultado_textos, "clicou": True, "motivo": metodo}

def preencher_quantidade_e_valor(driver, wait, quantidade="1", valor_boleto="123,45"):
    wait_overlays_to_hide(wait)
    locator_qtd = (By.XPATH, "//input[@id='quantidade' or @name='quantidade' or @formcontrolname='frmQuantidade']")
    try: qtd = wait.until(EC.presence_of_element_located(locator_qtd))
    except TimeoutException: raise RuntimeError("Campo 'Quantidade' não encontrado.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", qtd)
    time.sleep(0.05)
    try: wait.until(EC.element_to_be_clickable(locator_qtd))
    except TimeoutException: pass

    try: clear_and_type(qtd, quantidade)
    except Exception: driver.execute_script("arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input',{bubbles:true}));", qtd, quantidade)

    locator_valor = (By.XPATH, "//input[@id='valorUnitario' or @name='valorUnitario' or @formcontrolname='frmValor']")
    try: valor = wait.until(EC.presence_of_element_located(locator_valor))
    except TimeoutException: raise RuntimeError("Campo 'Valor Unitário' não encontrado.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", valor)
    time.sleep(0.05)
    try: wait.until(EC.element_to_be_clickable(locator_valor))
    except TimeoutException: pass

    try: clear_and_type(valor, valor_boleto)
    except Exception:
        somente_digitos = re.sub(r"\D", "", str(valor_boleto))
        try: clear_and_type(valor, somente_digitos)
        except Exception: driver.execute_script("arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input',{bubbles:true}));", valor, somente_digitos)

def selecionar_opcao_justificativa_com_hover(driver, wait, texto_alvo="2 - Orientações do gestor/coordendor da área"):
    locator_option_exata = (By.XPATH, f"//mat-option//span[contains(normalize-space(.), '{texto_alvo.split(' - ')[-1].split()[0]}') and contains(normalize-space(.), 'Orientações do gestor')]")
    try: opt_span = wait.until(EC.presence_of_element_located(locator_option_exata))
    except TimeoutException:
        try: opt_span = wait.until(EC.presence_of_element_located((By.XPATH, "//mat-option//span[contains(normalize-space(.), 'Orientações do gestor')]")))
        except TimeoutException: raise RuntimeError("Não encontrei a opção de justificativa no painel.")

    try: mat_option = opt_span.find_element(By.XPATH, "./ancestor::mat-option[1]")
    except Exception: mat_option = opt_span

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", mat_option)
    time.sleep(0.05)

    try:
        ActionChains(driver).move_to_element(mat_option).pause(0.15).click().perform()
        return
    except Exception: pass

    try:
        mat_option.click()
        return
    except Exception:
        try:
            driver.execute_script("arguments[0].click();", mat_option)
            return
        except Exception: pass

    try:
        from selenium.webdriver.common.keys import Keys
        body = driver.find_element(By.TAG_NAME, "body")
        for _ in range(10):  
            body.send_keys(Keys.ARROW_DOWN)
            time.sleep(0.05)
        body.send_keys(Keys.ENTER)
    except Exception as e3: raise RuntimeError(f"Falha ao selecionar a justificativa (mesmo com hover): {repr(e3)}")

def _strip_accents(s: str) -> str:
    if pd.isna(s) or s is None: return ''
    s = unicodedata.normalize('NFKD', str(s))
    return ''.join(ch for ch in s if not unicodedata.combining(ch))

def _norm_colname(s: str) -> str:
    return _strip_accents(str(s)).lower().strip()

def _clean_str(x) -> str:
    if pd.isna(x): return ''
    if isinstance(x, float):
        if x.is_integer(): return str(int(x))
        return str(x)
    s = str(x).strip()
    if s.endswith('.0'): s = s[:-2]
    return s

def _norm_coletor(x: str) -> str:
    up = _strip_accents(_clean_str(x)).upper()
    return re.sub(r'[^A-Z0-9]', '', up)

def _is_valid_coletor(coletor: str) -> bool:
    if pd.isna(coletor) or str(coletor).strip() == '' or str(coletor).strip().lower() == 'nan': return False
    c = _strip_accents(_clean_str(coletor)).upper()
    c_clean = re.sub(r'[^A-Z0-9]', '', c)
    if re.fullmatch(r'\d{6,}', c_clean): return True
    if 6 <= len(c_clean) <= 12 and any(x.isalpha() for x in c_clean) and any(x.isdigit() for x in c_clean): return True
    return False

def _tipo_de_coletor(coletor: str) -> str:
    c = _norm_coletor(coletor)
    if c == 'SEMCENTRODECUSTO': return '-'
    if re.fullmatch(r'\d+', c): return 'N'
    if any(x.isalpha() for x in c) and any(x.isdigit() for x in c): return 'K'
    return ''

def _clean_valor_series(s: pd.Series) -> pd.Series:
    def limpa_valor(val):
        if pd.isna(val) or str(val).strip() == '': return None
        if isinstance(val, (int, float)): return float(val)
        v = str(val).upper().replace('R$', '').replace('\xa0', '').replace(' ', '').strip()
        if '.' in v and ',' in v: v = v.replace('.', '').replace(',', '.')
        elif ',' in v: v = v.replace(',', '.')
        try: return float(v)
        except ValueError: return None
    return s.apply(limpa_valor)

def ler_rr_bruto(caminho_rr: Union[str, Path]) -> pd.DataFrame:
    df_raw = pd.read_excel(caminho_rr, header=None, engine='openpyxl')
    
    col_coletor = -1
    col_valor = -1
    linha_cabecalho = -1
    
    # Procura onde estão os cabeçalhos "COLETOR" e "VALOR"
    for i, row in df_raw.head(20).iterrows():
        row_str = [str(x).strip().upper() for x in row.values]
        
        achou_coletor = False
        achou_valor = False
        
        for j, val in enumerate(row_str):
            # Agora aceita "CC/ OBRA", "OBRA", "CENTRO DE CUSTO", "COLETOR", etc.
            if any(palavra in val for palavra in ['COLETOR', 'CENTRO DE CUSTO', 'CC/', 'CC /', 'OBRA']) or val == 'CC':
                col_coletor = j
                achou_coletor = True
            elif 'VALOR' in val:
                col_valor = j
                achou_valor = True
                
        if achou_coletor and achou_valor:
            linha_cabecalho = i
            break
            
    if linha_cabecalho == -1 or col_coletor == -1 or col_valor == -1:
        print("⚠️ Aviso: Não encontrei as colunas COLETOR e VALOR no Rateio Recebido.")
        return pd.DataFrame(columns=['TIPOCOLETOR', 'COLETOR', 'VALOR'])
        
    # Extrai os dados abaixo do cabeçalho
    df_clean = df_raw.iloc[linha_cabecalho + 1:].copy()
    df_clean = df_clean[[col_coletor, col_valor]]
    df_clean.columns = ['COLETOR_ORIG', 'VALOR']
    
    # FILTRO INTELIGENTE: Remove as linhas de "Total" ou rodapés inválidos
    df_clean = df_clean[df_clean['COLETOR_ORIG'].apply(_is_valid_coletor)].copy()
    
    # Limpa os valores usando a sua função existente
    df_clean['VALOR'] = _clean_valor_series(df_clean['VALOR'])
    df_clean = df_clean.dropna(subset=['VALOR'])
    
    # Padroniza os coletores
    df_clean['COLETOR'] = df_clean['COLETOR_ORIG'].apply(_norm_coletor)
    
    # Define o tipo (K ou N) automaticamente
    df_clean['TIPOCOLETOR'] = df_clean['COLETOR'].apply(_tipo_de_coletor)
    
    return df_clean[['TIPOCOLETOR', 'COLETOR', 'VALOR']]

def _extrair_coletor_de_titular(texto: str) -> str:
    if pd.isna(texto) or str(texto).strip() == '': return "SEM CENTRO DE CUSTO"
    t = _strip_accents(str(texto)).upper()
    
    m = re.search(r'(?<!\d)(\d{6,})(?!\d)', t)
    if m: return _norm_coletor(m.group(1))
    
    palavras = t.split()
    for p in palavras:
        p_clean = re.sub(r'[^A-Z0-9]', '', p)
        if 6 <= len(p_clean) <= 12 and any(c.isalpha() for c in p_clean) and any(c.isdigit() for c in p_clean):
            return p_clean
            
    for i in range(len(palavras) - 1):
        p1 = re.sub(r'[^A-Z0-9]', '', palavras[i])
        p2 = re.sub(r'[^A-Z0-9]', '', palavras[i+1])
        comb = p1 + p2
        if 8 <= len(comb) <= 12 and any(c.isalpha() for c in comb) and any(c.isdigit() for c in comb):
            return comb
            
    return "SEM CENTRO DE CUSTO"

def ler_correios_bruto(caminho_correios: Union[str, Path]) -> Tuple[pd.DataFrame, float]:
    df_raw = pd.read_excel(caminho_correios, header=None, engine='openpyxl')
    idx_header = -1
    col_titular, col_valor = -1, -1
    
    for i, row in df_raw.head(20).iterrows():
        row_norm = [_norm_colname(str(x)) for x in row.values]
        if any('titular do cartao' in c for c in row_norm) and any('valor do servico' in c for c in row_norm):
            idx_header = i
            for j, c in enumerate(row_norm):
                if 'titular do cartao' in c: col_titular = j
                if 'valor do servico' in c: col_valor = j
            break
            
    valor_liquido = 0.0
    idx_fim_tabela = len(df_raw)
    
    for i in range(len(df_raw) - 1, -1, -1):
        row_norm = [_norm_colname(str(x)) for x in df_raw.iloc[i].values]
        if any('valor liquido' in c for c in row_norm):
            idx_fim_tabela = i
            for j, c in enumerate(row_norm):
                if 'valor liquido' in c:
                    if i + 1 < len(df_raw):
                        val_raw = df_raw.iloc[i + 1, j]
                        valor_liquido = _clean_valor_series(pd.Series([val_raw])).iloc[0]
                    break
            break

    if idx_header == -1: 
        return pd.DataFrame(columns=['TIPOCOLETOR', 'COLETOR', 'VALOR']), valor_liquido
    
    df = df_raw.iloc[idx_header + 1 : idx_fim_tabela, [col_titular, col_valor]].copy()
    df.columns = ['TITULAR', 'VALOR']
    
    mask_ignorar = df['TITULAR'].astype(str).str.upper().str.contains('ENCARGO|DESCONTO|CREDITO')
    df = df[~mask_ignorar].copy()
    
    df['VALOR'] = _clean_valor_series(df['VALOR'])
    df = df.dropna(subset=['VALOR'])
    df['COLETOR'] = df['TITULAR'].apply(_extrair_coletor_de_titular)
    df['TIPOCOLETOR'] = df['COLETOR'].apply(_tipo_de_coletor)
    
    return df[['TIPOCOLETOR', 'COLETOR', 'VALOR']], valor_liquido

def _formatar_planilha_final(arquivo_xlsx: Union[str, Path], sheet='Planilha1'):
    wb = load_workbook(arquivo_xlsx)
    if sheet not in wb.sheetnames:
        wb.close(); return
    ws = wb[sheet]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        idx_valor = header.index('VALOR') + 1
        idx_op = header.index('OPERACAO') + 1
    except ValueError:
        wb.close(); return
        
    for row in ws.iter_rows(min_row=2):
        cell_valor = row[idx_valor - 1]
        if isinstance(cell_valor.value, (int, float)):
            cell_valor.number_format = numbers.FORMAT_NUMBER_00
            
    wb.save(arquivo_xlsx)
    wb.close()

def gerar_rateio_pag(
    caminho_correios: Union[str, Path],
    caminho_rr: Union[str, Path],
    saida: Union[str, Path] = 'RATEIO PAG.xlsx',
    operacao_para_diagrama: int = 10,
    tolerancia_igual: float = 0.05, 
    debug: bool = True
) -> pd.DataFrame:

    # 1. Lê os arquivos
    df_rr_raw = ler_rr_bruto(caminho_rr)
    df_corr_raw, valor_liquido_correios = ler_correios_bruto(caminho_correios)

    df_rr_raw['VALOR'] = pd.to_numeric(df_rr_raw['VALOR'], errors='coerce').fillna(0.0)
    df_corr_raw['VALOR'] = pd.to_numeric(df_corr_raw['VALOR'], errors='coerce').fillna(0.0)

    total_rr = float(df_rr_raw['VALOR'].sum()) if not df_rr_raw.empty else 0.0
    total_corr_soma = float(df_corr_raw['VALOR'].sum()) if not df_corr_raw.empty else 0.0
    total_corr = valor_liquido_correios if valor_liquido_correios > 0 else total_corr_soma

    if debug:
        print(f"[DEBUG] TOTAL RR               = R$ {total_rr:.2f}")
        print(f"[DEBUG] TOTAL CORREIOS (SOMA)  = R$ {total_corr_soma:.2f}")
        print(f"[DEBUG] TOTAL CORREIOS (LÍQ)   = R$ {valor_liquido_correios:.2f}")

    # 2. Agrupa os valores por Centro de Custo para evitar duplicidades
    df_corr_grouped = df_corr_raw.groupby(['TIPOCOLETOR', 'COLETOR'], as_index=False)['VALOR'].sum()
    
    if not df_rr_raw.empty:
        df_rr_grouped = df_rr_raw.groupby(['TIPOCOLETOR', 'COLETOR'], as_index=False)['VALOR'].sum()
    else:
        df_rr_grouped = pd.DataFrame(columns=['TIPOCOLETOR', 'COLETOR', 'VALOR'])

    linhas_finais = []
    
    # Cria uma lista com os Centros de Custo que vieram no e-mail
    ccs_no_email = set(df_rr_grouped['COLETOR'].tolist())

    # PASSO A: Adiciona tudo que veio no e-mail (Prioridade Máxima)
    for _, row in df_rr_grouped.iterrows():
        linhas_finais.append({
            'TIPOCOLETOR': row['TIPOCOLETOR'],
            'COLETOR': row['COLETOR'],
            'VALOR': row['VALOR']
        })

    # PASSO B: Adiciona os CCs dos Correios que NÃO foram mencionados no e-mail
    # (Ex: PRIMDF3023, PRIMGO3023 vão entrar aqui com seus valores originais intactos)
    for _, row in df_corr_grouped.iterrows():
        if row['COLETOR'] not in ccs_no_email:
            linhas_finais.append({
                'TIPOCOLETOR': row['TIPOCOLETOR'],
                'COLETOR': row['COLETOR'],
                'VALOR': row['VALOR']
            })

    # 3. Transforma na base final
    final_base = pd.DataFrame(linhas_finais)

    # 4. Rateio Proporcional de Diferenças (Centavos, Encargos, Descontos)
    soma_atual = final_base['VALOR'].sum() if not final_base.empty else 0.0
    diferenca_rateio = round(total_corr - soma_atual, 2)
    
    if abs(diferenca_rateio) > 0.02 and not final_base.empty:
        if debug: print(f"[DEBUG] Rateando R$ {diferenca_rateio:.2f} (Encargos/Descontos) proporcionalmente...")
        
        soma_validos = final_base['VALOR'].sum()
        if soma_validos > 0:
            final_base['VALOR_ADD'] = (final_base['VALOR'] / soma_validos) * diferenca_rateio
            final_base['VALOR_ADD'] = final_base['VALOR_ADD'].round(2)
            
            # Ajuste de centavos no maior valor para bater exatamente com o boleto
            diff_centavos = round(diferenca_rateio - final_base['VALOR_ADD'].sum(), 2)
            if diff_centavos != 0:
                idx_max = final_base['VALOR'].idxmax() 
                final_base.loc[idx_max, 'VALOR_ADD'] += diff_centavos 
                
            final_base['VALOR'] += final_base['VALOR_ADD']
            final_base = final_base.drop(columns=['VALOR_ADD'])

    # 5. Formatar para o MRV Pag
    final = pd.DataFrame()
    if not final_base.empty:
        final['ITEM'] = [1] * len(final_base)
        final['TIPOCOLETOR'] = final_base['TIPOCOLETOR']
        final['COLETOR'] = final_base['COLETOR']
        final['OPERACAO'] = final_base['TIPOCOLETOR'].apply(lambda t: operacao_para_diagrama if t == 'N' else '')
        final['SUBNUMERO'] = ''
        final['VALOR'] = final_base['VALOR']
        final['DESCRICAO'] = ''

        final['__ord'] = final['TIPOCOLETOR'].map({'K': 0, 'N': 1}).fillna(2)
        final = final.sort_values(['__ord', 'COLETOR']).drop(columns='__ord').reset_index(drop=True)

    saida = Path(saida)
    with pd.ExcelWriter(saida, engine='openpyxl') as writer:
        final.to_excel(writer, sheet_name='Planilha1', index=False)

    _formatar_planilha_final(saida, 'Planilha1')

    if debug: print(f"[DEBUG] Arquivo gerado com sucesso: {saida.resolve()}")

    return final

