import time
import getpass
import os
import glob
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime, timedelta
import win32com.client
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import unicodedata
import re
import sys
from pathlib import Path
import config

# Configuração de caminhos dinâmicos
from config import EMAIL_MRV, SENHA_MRV 

def normalizar_texto(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s)
    return s

def executar_juridico(pular_download=False):
    print("[PROGRESSO: 5]")
    print("Iniciando automação do Jurídico Montreal...")
    
    if not pular_download:
        print("Opção: Baixar do Podio ativada. Abrindo navegador...")
        try:
            try:
                driver = webdriver.Chrome() 
                driver.get("https://podio.com/login")
                driver.maximize_window()
            except Exception as e:
                print(f"Erro ao iniciar o Chrome. Verifique seu chromedriver. {e}")
                sys.exit(1)
                
            # --- Etapa 1: Aceitar Cookies E Clicar no Login da Microsoft ---
            try:
                print("Aguardando página de login...")
                try:
                    print("Procurando o banner de cookies (OneTrust)...")
                    accept_cookies_id = "onetrust-accept-btn-handler"
                    cookie_button = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.ID, accept_cookies_id))
                    )
                    print("Banner de cookies encontrado. Clicando em 'Aceitar'...")
                    cookie_button.click()
                    time.sleep(1) 
                except TimeoutException:
                    print("Banner de cookies não encontrado ou já aceito. Continuando...")

                microsoft_login_xpath = "//a[@data-provider='live']"
                print("Procurando o botão de login da Microsoft...")
                microsoft_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, microsoft_login_xpath))
                )
                print("Clicando no botão da Microsoft...")
                microsoft_button.click()

            except Exception as e:
                print(f"Erro na Etapa 1: {e}")
                driver.save_screenshot("erro_etapa_1_cookies.png")
                driver.quit()
                sys.exit(1)
                
            # --- Etapa 1.5: Lidar com o login da Microsoft ---
            try:
                print("[PROGRESSO: 15]")
                print("Aguardando a nova janela/aba de login da Microsoft...")
                WebDriverWait(driver, 12).until(EC.number_of_windows_to_be(2))

                popup_window = None
                original_window = driver.current_window_handle 
                for window_handle in driver.window_handles:
                    if window_handle != original_window:
                        popup_window = window_handle
                        break
                        
                driver.switch_to.window(popup_window)
                print("Foco mudado para a janela de login da Microsoft.")
                        
                print("Aguardando a tela de login da Microsoft...")
                
                email_field_microsoft = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.ID, "i0116")) 
                )
                print("Preenchendo e-mail da Microsoft...")
                email_field_microsoft.send_keys(EMAIL_MRV)
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "idSIButton9"))).click()
                
                password_field_microsoft = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.ID, "i0118")) 
                )

                print("Preenchendo senha da Microsoft...")
                password_field_microsoft.send_keys(SENHA_MRV)
                
                print("Procurando o botão 'Entrar'...")
                tentativas = 0
                clicado_entrar = False
                while not clicado_entrar and tentativas < 5:
                    try:
                        entrar_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "idSIButton9")))
                        entrar_button.click()
                        clicado_entrar = True
                        print("Botão 'Entrar' clicado.")
                    except StaleElementReferenceException:
                        tentativas += 1; time.sleep(0.5)
                if not clicado_entrar: raise Exception("Falha ao clicar em Entrar")

                print("!!! AÇÃO MANUAL NECESSÁRIA !!!")
                print("Aguardando aprovação do MFA no seu celular (até 180s)...")
                
                tentativas = 0
                clicado_manter = False
                while not clicado_manter and tentativas < 5:
                    try:
                        keep_logged_in_button = WebDriverWait(driver, 180).until( 
                            EC.element_to_be_clickable((By.ID, "idSIButton9"))
                        )
                        keep_logged_in_button.click() 
                        clicado_manter = True
                        print("MFA Aprovado! Botão 'Manter conectado' clicado.")
                    except StaleElementReferenceException:
                        tentativas += 1; time.sleep(0.5)
                    except TimeoutException:
                        print("Erro: Timeout após 180s. Você não aprovou o MFA a tempo?")
                        clicado_manter = False; break
                if not clicado_manter: raise Exception("Falha ao clicar em Manter Conectado")

                print("Login da Microsoft concluído na janela pop-up.")
                
                print("Aguardando janela pop-up fechar...")
                WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(1))
                
                nova_janela_principal = driver.window_handles[0]
                driver.switch_to.window(nova_janela_principal)
                print("Foco retornado para a janela principal do Podio.")

                print("Forçando navegação para https://podio.com/home")
                driver.get("https://podio.com/home")
                print("Página principal carregada com sucesso.")
                
            except Exception as e:
                print(f"Erro durante o login na Microsoft (Etapa 1.5): {e}")
                driver.save_screenshot("erro_etapa_1-5.png") 
                driver.quit()
                sys.exit(1)

            # --- Início da Navegação no Podio (Etapas 2-12) ---
            try:
                print("[PROGRESSO: 30]")
                print("Etapa 2: Procurando 'Vá para uma área de trabalho'...")
                parent_element_xpath = "//div[contains(@class, 'space-switcher-wrapper')]"
                parent_element = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.XPATH, parent_element_xpath))
                )
                
                print("Simulando passagem do mouse (hover)...")
                actions = ActionChains(driver)
                actions.move_to_element(parent_element).perform()
                
                print("Etapa 3: Aguardando a lista de áreas de trabalho...")
                adm_link_xpath = "//a[contains(text(), 'ADM - Núcleo Contratos')]"
                adm_link = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, adm_link_xpath))
                )
                print("Clicando em 'ADM - Núcleo Contratos'...")
                adm_link.click()

                print("Etapa 4: Procurando o app 'Mensageria'...")
                mensageria_app_xpath = "//li[@data-app-id='22830484']"
                mensageria_app = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, mensageria_app_xpath))
                )
                print("Clicando no app 'Mensageria'...")
                mensageria_app.click()
                
                print("Etapa 5: Aguardando a página 'Mensageria' carregar...") 
                x = 0
                clicado_filtro = False
                while x <=2: 
                    try:
                        parent_filter_xpath = "//ul[@class='app-filter-tools']"
                        parent_filter_element = WebDriverWait(driver, 2).until(
                            EC.presence_of_element_located((By.XPATH, parent_filter_xpath))
                        )
                        print(f"Tentativa {x+1}: Container <ul> (pai) encontrado.")

                        child_items_xpath = ".//li"
                        child_items = parent_filter_element.find_elements(By.XPATH, child_items_xpath)
                        
                        if not child_items:
                            raise Exception("Container <ul> encontrado, mas nenhum <li> filho foi encontrado.")
                        
                        print(f"Encontrados {len(child_items)} itens. Passando o mouse sobre eles...")

                        actions_filter = ActionChains(driver)
                        for item in child_items:
                            actions_filter.move_to_element(item)
                        actions_filter.perform() 
                        
                        target_filter_xpath = ".//li[@data-original-title='Filtros']" 
                        target_filter = WebDriverWait(parent_filter_element, 2).until(
                            EC.element_to_be_clickable((By.XPATH, target_filter_xpath))
                        )
                        
                        print("Ícone 'Filtros' acordado e clicável. Clicando...")
                        target_filter.click()
                        
                        clicado_filtro = True 
                        print("Ícone 'Filtros' clicado com sucesso.")
                        break
                    except (StaleElementReferenceException, TimeoutException) as e:
                        print(f"Tentativa {x} falhou (Stale ou Timeout). Página recarregando... Tentando de novo.")
                        time.sleep(1) 
                        x += 1
                        
                if not clicado_filtro:
                    raise Exception("Falha ao encontrar o ícone 'Filtros' após 2 tentativas.")

                print("Ação concluída com sucesso!")
                time.sleep(1)

                print("[PROGRESSO: 45]")
                print("Etapa 6: Procurando o 'Criado em'...")
                Criado_em_xpath = "//li[@data-id='created_on']"
                Criado_em = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((By.XPATH, Criado_em_xpath))
                )
                print("Clicando no 'Criado em'...")
                Criado_em.click()

                print("Etapa 7: Procurando o 'Hoje'...")
                Criado_em_hoje = "//li[@data-id='-0dr:-0dr']"
                Criado_hoje = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, Criado_em_hoje))
                )
                print("Clicando no 'Hoje'...")
                Criado_hoje.click()

                try:
                    seletor_css = ".app-header__app-menu"
                    print(f"Procurando todos os elementos com a classe: {seletor_css}")
                    WebDriverWait(driver, 10).until(
                        lambda d: len(d.find_elements(By.CSS_SELECTOR, seletor_css)) >= 2
                    )

                    elementos = driver.find_elements(By.CSS_SELECTOR, seletor_css)
                    print(f"Encontrados {len(elementos)} elementos.")

                    if len(elementos) > 0:
                        print("Clicando no primeiro elemento (índice 0)...")
                        elementos[0].click()
                
                    print("Aguardando 2 segundos para a página/menu reagir...")
                    time.sleep(2) 

                    print("Re-encontrando os elementos (para segurança)...")
                    elementos = driver.find_elements(By.CSS_SELECTOR, seletor_css)

                    if len(elementos) > 1:
                        print("Clicando no segundo elemento (índice 1)...")
                        elementos[1].click()
                    else:
                        print("Erro: Não foi possível encontrar o segundo elemento após o primeiro clique.")
                    
                    print("Ações nos dois elementos concluídas!")
                    time.sleep(3)

                except Exception as e:
                    print(f"Ocorreu um erro: {e}")

                print("Etapa 9: Aguardando o menu dropdown abrir...")
                exportar_excel_selector = "a.app-box-supermenu-v2__link.app-export-excel"
                exportar_link = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, exportar_excel_selector))
                )
                print("Link 'Exportar Excel' encontrado. Clicando...")
                exportar_link.click()
                time.sleep(3)

                try:
                    print("Procurando o ícone de 'Notificação' (Inbox)...")
                    notificacao_selector = "li.navigation-link.inbox"
                    notificacao_icon = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, notificacao_selector))
                    )
                    print("Ícone de 'Notificação' encontrado. Clicando...")
                    notificacao_icon.click()
                    time.sleep(1) 
                except Exception as e:
                    print(f"Erro ao tentar clicar no ícone de Notificação: {e}")
                    driver.save_screenshot("erro_notificacao.png")

                css_corrigido = "a.PodioUI__Notifications__NotificationGroup"
                item_notificacao = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, css_corrigido))
                )
                item_notificacao.click()

                print("[PROGRESSO: 60]")
                print("Etapa 12: Aguardando a página de exportação carregar e o status ser 'Completado'...")
                status_completado_xpath = "//div[contains(@class, 'field-type-text')]"
                WebDriverWait(driver, 180).until(
                    EC.presence_of_element_located((By.XPATH, status_completado_xpath))
                )
                print("Exportação 'Completado'!")
                
                print("Etapa 13: Procurando o link de download do arquivo...")
                nome_do_arquivo = "Mensageria - Última vista usada.xlsx"
                link_download = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((By.LINK_TEXT, nome_do_arquivo))
                )
                print("Link encontrado! Clicando para baixar...")
                link_download.click()
                
                print("Ação final concluída com sucesso! O download deve começar.")
                time.sleep(10) 
                driver.quit()

            except Exception as e:
                print(f"Erro durante a navegação no Podio (Etapas 2-5): {e}")
                print("Verifique os seletores XPath. Um deles pode ter mudado.")
                driver.save_screenshot("erro_de_navegacao.png")
                driver.quit()
                sys.exit(1)
        
        except Exception as e:
            print(f"Erro inesperado durante o processo: {e}")
            driver.quit()
            sys.exit(1)         
    
    else:
        print("[PROGRESSO: 60]")
        print("Opção: Pular download ativada. Buscando arquivo já existente na pasta Downloads...")

    # ==========================================================================
    # PROCESSAMENTO DO EXCEL (Roda independente de ter baixado agora ou não)
    # ==========================================================================
    try:
        print("[PROGRESSO: 75]")
        print("\n--- INICIANDO PROCESSAMENTO DO EXCEL ---")
        
        download_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
        list_of_files = glob.glob(os.path.join(download_dir, 'Mensageria - Última vista usada*.xlsx'))
        
        if not list_of_files:
            raise Exception("Arquivo Excel não encontrado na pasta Downloads.\nCertifique-se de que você baixou a planilha do Podio.")
            
        latest_file = max(list_of_files, key=os.path.getctime)
        print(f"Arquivo encontrado: {latest_file}")
        
        workbook = openpyxl.load_workbook(latest_file)
        sheet = workbook.active
        print(f"Planilha '{sheet.title}' aberta.")

        print("1. Excluindo colunas A, I, J, K, L, M...")
        sheet.delete_cols(13) 
        sheet.delete_cols(12) 
        sheet.delete_cols(11) 
        sheet.delete_cols(10) 
        sheet.delete_cols(9)  
        sheet.delete_cols(1)  

        print("2. Ajustando tamanho das colunas para 20...")
        for i in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(i)
            sheet.column_dimensions[col_letter].width = 20

        print("3. Aplicando filtro...")
        sheet.auto_filter.ref = sheet.dimensions

        print("4. Filtrando por 'JURIDICO MONTREAL' na coluna F (ignorando acentos)...")
        linhas_para_excluir = []
        alvo = normalizar_texto("juridico montreal")

        for row_num in range(2, sheet.max_row + 1):
            cell_value = sheet[f'F{row_num}'].value
            valor_norm = normalizar_texto(cell_value)
            if alvo not in valor_norm:
                linhas_para_excluir.append(row_num)

        for row_num in reversed(linhas_para_excluir):
            sheet.delete_rows(row_num)

        print(f"{len(linhas_para_excluir)} linhas excluídas.")

        if sheet.max_row >= 2:
            tem_dados_juridico = True
            print("A planilha POSSUI dados do Jurídico.")
        else:
            tem_dados_juridico = False
            print("A planilha ESTÁ VAZIA (sem itens Jurídico Montreal).")

        print("5. Renomeando coluna G para 'Data de recebimento'...")
        sheet['G1'] = "Data de recebimento"

        print("6. Aplicando bordas...")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border

        print("7. Formatando cabeçalho...")
        header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in sheet[1]: 
            cell.fill = header_fill
            cell.font = header_font
            
        today_date = time.strftime("%d-%m-%Y") 
        nome_arquivo_final = f"juridico montreal - {today_date}.xlsx"
        
        excel_file_path = os.path.join(download_dir, nome_arquivo_final)
        workbook.save(excel_file_path) 
        
        print("\n--- PROCESSAMENTO DO EXCEL CONCLUÍDO! ---")
        print(f"Novo arquivo salvo como: {excel_file_path}")

    except Exception as e:
        print(f"ERRO DURANTE O PROCESSAMENTO DO EXCEL: {e}")
        excel_file_path = None 
        sys.exit(1)

    # ==========================================================================
    # CRIAÇÃO DO E-MAIL NO OUTLOOK
    # ==========================================================================
    if excel_file_path and os.path.exists(excel_file_path):
        try:
            print("[PROGRESSO: 90]")
            print("\n--- INICIANDO CRIAÇÃO DE E-MAIL NO OUTLOOK ---")

            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)

            mail.To = (
                "kelly.paixao@parceiro.mrv.com.br;"
                "reinaldo.reis@parceiro.mrv.com.br;"
                "maria.rferreira@parceiro.mrv.com.br"
            )
            mail.CC = "correiosbh@mrv.com.br"

            hoje = datetime.now()
            data_formatada = hoje.strftime("%d/%m/%Y")
            mail.Subject = f"Jurídico Montreal - {data_formatada}"

            if tem_dados_juridico:
                mensagem_principal = f"Segue em anexo os documentos que chegaram para a Montreal no dia {data_formatada}."
            else:
                mensagem_principal = f"No dia {data_formatada}, não chegou nenhum documento para a Montreal."

            mail.BodyFormat = 2  
            mail.Display(False)
            assinatura_outlook = mail.HTMLBody

            corpo_email = f"""
            <p style="font-family: Calibri, sans-serif; font-size: 11pt;">
                Bom dia, Prezado(s)!
            </p>
            <p style="font-family: Calibri, sans-serif; font-size: 11pt;">
                {mensagem_principal}
            </p>
            <p style="font-family: Calibri, sans-serif; font-size: 11pt;">
                Atenciosamente;
            </p>
            """

            mail.HTMLBody = f"""
            <html>
            <body>
                {corpo_email}
                {assinatura_outlook}
            </body>
            </html>
            """

            if tem_dados_juridico:
                mail.Attachments.Add(excel_file_path)
                print("Arquivo Excel anexado.")
            else:
                print("Planilha vazia: Arquivo NÃO anexado.")

            print("[PROGRESSO: 100]")
            print("E-mail criado no Outlook com sucesso.")

        except Exception as e:
            print(f"ERRO AO CRIAR E-MAIL NO OUTLOOK: {e}")
            sys.exit(1)
    else:
        print("Não foi possível criar o e-mail: arquivo Excel não encontrado.")
        sys.exit(1)

    print("Programa concluído.")

if __name__ == "__main__":
    executar_juridico()
