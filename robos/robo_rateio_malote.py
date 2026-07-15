#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
RATEIO MALOTE v14 - Correção de Valor Órfão e Filtro de Mesma Cidade
"""
import sys
import os
import re
import glob
import warnings
from collections import OrderedDict, defaultdict
from itertools import combinations
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ==============================================================================
# CONFIGURAÇÃO DE PASTAS DINÂMICAS
# ==============================================================================
sys.path.insert(0, str(Path(__file__).parent)) 
sys.path.append(str(Path(__file__).parent.parent)) 

PASTA_MALOTE = Path(__file__).parent.parent / "arquivos" / "rateio_malote"

try:
    from malote_web_scraper import ResolvedorCC, SELENIUM_AVAILABLE
    RESOLVER_AVAILABLE = True
except ImportError as e:
    RESOLVER_AVAILABLE = False
    SELENIUM_AVAILABLE = False
    print(f"⚠️ Módulo malote_web_scraper não encontrado: {e}")

# ============================================================
# ESTILOS
# ============================================================
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
HEADER_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def find_agilis_file():
    if not PASTA_MALOTE.exists(): return None
    for f in PASTA_MALOTE.glob("*.xlsx"):
        if "agilis" in f.name.lower(): return str(f)
    for f in PASTA_MALOTE.glob("*.xls"):
        if "agilis" in f.name.lower(): return str(f)
    return None

def find_correios_file():
    if not PASTA_MALOTE.exists(): return None
    for f in PASTA_MALOTE.glob("*.xlsx"):
        if re.match(r'^\d+$', f.stem): return str(f)
    for f in PASTA_MALOTE.glob("*.xls"):
        if re.match(r'^\d+$', f.stem): return str(f)
    return None

def find_previous_rateios():
    rateios = []
    pasta_arquivos = PASTA_MALOTE.parent 
    if not pasta_arquivos.exists(): return rateios
    
    for f in pasta_arquivos.rglob("*.xlsx"):
        if "rateio" in f.name.lower() and f.name != "Rateio Malote.xlsx":
            rateios.append(str(f))
    return rateios

def get_sheet_with_most_data(wb, preferred_names=None):
    if preferred_names:
        for name in preferred_names:
            for sn in wb.sheetnames:
                if name.lower() in sn.lower():
                    return wb[sn]
    best = None
    mx = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        if (ws.max_row or 0) > mx:
            mx = ws.max_row
            best = ws
    return best

def extract_text_after(text, marker):
    if not text or not isinstance(text, str):
        return ""
    idx = text.find(marker)
    return text[idx + len(marker):] if idx != -1 else ""

def extract_text_before(text, marker):
    if not text or not isinstance(text, str):
        return ""
    idx = text.find(marker)
    return text[:idx] if idx != -1 else text

def extract_origin_from_e(text):
    after = extract_text_after(str(text), "Documento:")
    return extract_text_before(after, "*").strip()

def extract_destination_from_e(text):
    after = extract_text_after(str(text), "Destino:")
    return extract_text_before(after, "*").strip()

def extract_centro_custo(text):
    if not text or not isinstance(text, str):
        return ""
    start = None
    for marker, offset in [("Código:", 8), ("Codigo:", 8),
                            ("Coletor de Custo ADM", 22),
                            ("Centro de custo:", 17),
                            (". Código:", 9), (". Codigo:", 9)]:
        idx = text.find(marker)
        if idx != -1:
            start = idx + offset
            break
    if start is None:
        return ""
    raw = text[start:start + 999]
    if not raw:
        return ""
    chunk = ' '.join(raw.replace('\xa0', ' ').replace('\n', ' ').replace('\r', ' ').split())
    if not chunk:
        return ""
    if chunk[0].isdigit():
        si = chunk.find(' ')
        return chunk[:si] if si != -1 else chunk
    return chunk[:10].strip()

def normalize_city(name):
    if not name:
        return ""
    result = str(name).upper().strip()
    result = re.sub(r'\s*-\s*SEDE', '', result, flags=re.IGNORECASE)
    result = re.sub(r'\s*/\s*\w{2}\s*', ' ', result)
    result = re.sub(r'\s*-\s*\d+', '', result)
    for k, v in {'Á': 'A', 'À': 'A', 'Ã': 'A', 'Â': 'A', 'É': 'E', 'Ê': 'E',
                  'Í': 'I', 'Ó': 'O', 'Ô': 'O', 'Õ': 'O', 'Ú': 'U', 'Ç': 'C'}.items():
        result = result.replace(k, v)
    return result.strip()

def extract_percurso_from_text(text):
    if not text:
        return ""
    match = re.search(r'-\s*(\d{7,})', str(text))
    return match.group(1) if match else ""

def extract_city_from_agilis_field(text):
    if not text:
        return ""
    t = str(text).strip()
    t = re.sub(r'\s*-\s*SEDE', '', t, flags=re.IGNORECASE)
    match = re.match(r'^(.+?)\s*/\s*\w{2}', t)
    if match:
        return match.group(1).strip().upper()
    match = re.match(r'^(.+?)\s*-\s*\d+', t)
    if match:
        return match.group(1).strip().upper()
    return t.upper()

def get_percurso_from_row(origem, destino):
    perc_orig = extract_percurso_from_text(origem)
    perc_dest = extract_percurso_from_text(destino)

    if perc_orig and perc_dest:
        if perc_orig == perc_dest:
            return perc_orig
        
        cartoes_sede = ['10320249', '10320238', '10320240', '10320255', '10320246', '10320251']
        
        if perc_dest in cartoes_sede and perc_orig not in cartoes_sede:
            return perc_orig
            
        if perc_orig in cartoes_sede and perc_dest not in cartoes_sede:
            return perc_dest
            
        return max(perc_orig, perc_dest)

    if perc_orig:
        return perc_orig
    if perc_dest:
        return perc_dest
    return ""

def similarity(a, b):
    if not a or not b:
        return 0
    a_norm = normalize_city(a)
    b_norm = normalize_city(b)
    if a_norm == b_norm:
        return 1.0
    if a_norm in b_norm or b_norm in a_norm:
        return 0.85
    return SequenceMatcher(None, a_norm, b_norm).ratio()

# ============================================================
# CLASSE PRINCIPAL
# ============================================================

class RateioMalote:
    def __init__(self, agilis_path=None, correios_path=None, output_path="Rateio Malote.xlsx"):
        self.agilis_path = agilis_path or find_agilis_file()
        self.correios_path = correios_path or find_correios_file()
        self.output_path = output_path

        # ======================================================================
        # TRAVA DE SEGURANÇA: VALIDAÇÃO GLOBAL DE ARQUIVOS
        # ======================================================================
        erros_arquivos = []
        
        if not self.agilis_path:
            erros_arquivos.append("👉 Relatório Agilis (deve conter 'agilis' no nome)")
            
        if not self.correios_path:
            erros_arquivos.append("👉 Extrato dos Correios (o nome deve conter APENAS números, ex: '2554871.xlsx')")

        base_cc_encontrada = False
        vsc_encontrado = False
        pasta_arquivos = PASTA_MALOTE.parent
        
        if pasta_arquivos.exists():
            for f in pasta_arquivos.rglob("*.xlsx"):
                fl = f.name.lower()
                if ("centro" in fl and "custo" in fl) or "diagrama" in fl or ("base" in fl and "cc" in fl):
                    base_cc_encontrada = True
                if "acompanhamento" in fl and "vsc" in fl:
                    vsc_encontrado = True

        if not base_cc_encontrada:
            erros_arquivos.append("👉 Base Centro de Custo (deve conter 'base' e 'centro de custo' no nome)")
            
        if not vsc_encontrado:
            erros_arquivos.append("👉 Acompanhamento VSC (deve conter 'acompanhamento' e 'vsc' no nome)")

        if erros_arquivos:
            mensagem_erro = "O robô não pode iniciar porque faltam arquivos obrigatórios:\n\n" + "\n".join(erros_arquivos) + "\n\nPor favor, verifique a pasta 'arquivos/rateio_malote' e tente novamente."
            raise RuntimeError(mensagem_erro)
        # ======================================================================

        print(f"📄 Agilis: {Path(self.agilis_path).name}")
        print(f"📄 Correios: {Path(self.correios_path).name}")
        print(f"📄 Saída: {Path(self.output_path).name}")
        print("-" * 60)

        self.agilis_data = []
        self.agilis_headers = []
        self.correios_malote_data = []
        self.correios_malote_headers = []
        self.vsc_data = []
        self.vsc_processed = []
        self.pivot_correios = OrderedDict()
        self.pivot_correios_totals = OrderedDict()
        self.pivot_agilis = OrderedDict()
        self.rateio_data = []
        self.previous_rateio_data = []
        self.resolvedor_cc = None

        self.status_cartoes = {}
        self.vsc_resolvido_obra = set()

    def run(self):
        print("🔄 Etapa 1-2: Lendo dados...")
        self._read_data()
        print("🔄 Etapa 2.5: Lendo rateios anteriores...")
        self._read_previous_rateios()
        print("🔄 Etapa 3: Colunas Origem/Destino/CC...")
        self._process_agilis_columns()
        print("🔄 Etapa 4: Processando VSC...")
        self._process_vsc()
        print("🔄 Etapa 5: Tabelas dinâmicas...")
        self._create_pivot_tables()
        print("🔄 Etapa 6: Criando Rateio...")
        self._create_rateio()
        print("💾 Salvando...")
        self._save()
        print(f"✅ Concluído! → {self.output_path}")

    def _read_data(self):
        wb_ag = openpyxl.load_workbook(self.agilis_path, data_only=True, read_only=True)
        ws_ag = get_sheet_with_most_data(wb_ag, ["relatório", "relatorio"])
        print(f"  Agilis aba: '{ws_ag.title}' ({ws_ag.max_row} linhas)")

        header_row = 1
        cat_col = None
        subcat_col = None
        for row in ws_ag.iter_rows(min_row=1, max_row=10, values_only=False):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    vl = cell.value.strip().lower()
                    if vl == "subcategoria":
                        subcat_col = cell.column - 1
                        header_row = cell.row
                    elif vl == "categoria":
                        cat_col = cell.column - 1
                        header_row = cell.row

        hdr_raw = list(ws_ag.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        
        last_valid_col = 0
        for i, h in enumerate(hdr_raw):
            if h and str(h).strip():
                last_valid_col = i
                
        hdr = hdr_raw[:last_valid_col + 1]
        self.agilis_headers = [str(h).strip() if h else f"Col{i}" for i, h in enumerate(hdr)]

        target = "SOLICITAÇÃO DE ENVIO DE MALOTE"
        for row in ws_ag.iter_rows(min_row=header_row + 1, values_only=True):
            rl = list(row)[:last_valid_col + 1]
            found = False
            if cat_col is not None and cat_col < len(rl):
                if rl[cat_col] and target.upper() in str(rl[cat_col]).upper():
                    found = True
            if not found and subcat_col is not None and subcat_col < len(rl):
                if rl[subcat_col] and target.upper() in str(rl[subcat_col]).upper():
                    found = True
            if not found and cat_col is None and subcat_col is None:
                for cv in rl:
                    if cv and target.upper() in str(cv).upper():
                        found = True
                        break
            if found:
                self.agilis_data.append(rl)
        print(f"  ✅ Agilis: {len(self.agilis_data)} linhas filtradas")
        wb_ag.close()

        wb_cor = openpyxl.load_workbook(self.correios_path, data_only=True, read_only=True)
        
        ws_mal = None
        ws_rel = None
        
        for name in wb_cor.sheetnames:
            nl = name.lower()
            if "malote" in nl:
                ws_mal = wb_cor[name]
            elif "relat" in nl:
                ws_rel = wb_cor[name]

        if ws_mal:
            print(f"  Correios MALOTE: '{ws_mal.title}' (Aba dedicada encontrada)")
            first = True
            for row in ws_mal.iter_rows(values_only=True):
                if not any(row): continue
                if first:
                    self.correios_malote_headers = [str(h).strip() if h else f"Col{i}" for i, h in enumerate(row)]
                    first = False
                else:
                    self.correios_malote_data.append(list(row))

        if ws_rel:
            print(f"  Correios RELATÓRIO: '{ws_rel.title}'")
            header_found = False
            
            for row in ws_rel.iter_rows(values_only=True):
                if not any(row): continue
                rl = list(row)
                
                if rl[0] and isinstance(rl[0], str) and "VSC" in str(rl[0]).upper():
                    self.vsc_data.append({
                        'texto': rl[0],
                        'valor': rl[1] if len(rl) > 1 else None,
                        'has_credito': "CREDITO" in str(rl[0]).upper() or "CRÉDITO" in str(rl[0]).upper(),
                        'has_debito': "DEBITO" in str(rl[0]).upper() or "DÉBITO" in str(rl[0]).upper()
                    })
                    continue 
                
                if not ws_mal:
                    if not header_found:
                        row_str = " ".join([str(c).lower() for c in rl if c])
                        if "cartao" in row_str or "postagem" in row_str or (len(rl) > 4 and rl[4] and "servico" in str(rl[4]).lower()):
                            self.correios_malote_headers = [str(h).strip() if h else f"Col{i}" for i, h in enumerate(rl)]
                            header_found = True
                    else:
                        if len(rl) > 4 and rl[4] and isinstance(rl[4], str) and "MALOTE" in rl[4].upper():
                            self.correios_malote_data.append(rl)

        print(f"  📋 Cabeçalhos Malote: {self.correios_malote_headers}")
        print(f"  ✅ Linhas Malote: {len(self.correios_malote_data)}")
        print(f"  ✅ VSC: {len(self.vsc_data)} registros")

        wb_cor.close()

    def _read_previous_rateios(self):
        rateio_files = find_previous_rateios()
        print(f"  📂 Encontrados {len(rateio_files)} rateios anteriores")

        for filepath in rateio_files:
            print(f"    📄 Lendo: {os.path.basename(filepath)}")
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)

                ws_rateio = None
                for name in wb.sheetnames:
                    if "rateio" in name.lower():
                        ws_rateio = wb[name]
                        break

                if ws_rateio:
                    headers = []
                    for cell in ws_rateio[1]:
                        headers.append(str(cell.value).strip().upper() if cell.value else "")

                    col_map = {}
                    for i, h in enumerate(headers):
                        hl = h.lower().strip()
                        if 'origem' in hl and 'destino' not in hl:
                            col_map['origem'] = i
                        elif 'destino' in hl:
                            col_map['destino'] = i
                        elif 'centro' in hl or 'centrodecusto' in hl:
                            col_map['cc'] = i
                        elif 'vsc' in hl and 'local' not in hl and 'nao' not in hl and 'não' not in hl:
                            col_map['vsc'] = i
                        elif 'total' in hl:
                            col_map['total'] = i
                        elif 'utiliza' in hl:
                            col_map['utilizacao'] = i

                    if 'origem' in col_map and 'destino' in col_map:
                        for row in ws_rateio.iter_rows(min_row=2, values_only=True):
                            rl = list(row)
                            if len(rl) <= max(col_map.values()):
                                continue

                            origem = str(rl[col_map['origem']]).strip() if rl[col_map['origem']] else ""
                            destino = str(rl[col_map['destino']]).strip() if rl[col_map['destino']] else ""
                            cc = str(rl[col_map.get('cc', 0)]).strip() if col_map.get('cc') is not None and rl[col_map.get('cc', 0)] else ""

                            vsc_val = 0
                            if 'vsc' in col_map:
                                try:
                                    vsc_val = float(rl[col_map['vsc']]) if rl[col_map['vsc']] else 0
                                except:
                                    vsc_val = 0

                            total_val = 0
                            if 'total' in col_map:
                                try:
                                    total_val = float(rl[col_map['total']]) if rl[col_map['total']] else 0
                                except:
                                    total_val = 0

                            util_val = 0
                            if 'utilizacao' in col_map:
                                try:
                                    util_val = float(rl[col_map['utilizacao']]) if rl[col_map['utilizacao']] else 0
                                except:
                                    util_val = 0

                            if origem or destino:
                                perc = get_percurso_from_row(origem, destino)
                                self.previous_rateio_data.append({
                                    'origem': origem,
                                    'destino': destino,
                                    'cc': cc,
                                    'percurso': perc,
                                    'vsc': vsc_val,
                                    'total': total_val,
                                    'utilizacao': util_val,
                                    'arquivo': os.path.basename(filepath),
                                    'origem_city': normalize_city(extract_city_from_agilis_field(origem)),
                                    'destino_city': normalize_city(extract_city_from_agilis_field(destino)),
                                })

                ws_vsc_prev = None
                for name in wb.sheetnames:
                    if "vsc" in name.lower():
                        ws_vsc_prev = wb[name]
                        break

                if ws_vsc_prev:
                    vsc_headers = []
                    for cell in ws_vsc_prev[1]:
                        vsc_headers.append(str(cell.value).strip().upper() if cell.value else "")

                    vsc_col_map = {}
                    for i, h in enumerate(vsc_headers):
                        hl = h.lower()
                        if 'percurso' in hl or 'perc' in hl:
                            vsc_col_map['percurso'] = i
                        elif 'origem' in hl:
                            vsc_col_map['origem'] = i
                        elif 'destino' in hl:
                            vsc_col_map['destino'] = i
                        elif 'valor' in hl:
                            vsc_col_map['valor'] = i

                    if 'percurso' in vsc_col_map:
                        for row in ws_vsc_prev.iter_rows(min_row=2, values_only=True):
                            rl = list(row)
                            if len(rl) <= max(vsc_col_map.values()):
                                continue
                            perc = str(rl[vsc_col_map['percurso']]).strip() if rl[vsc_col_map['percurso']] else ""
                            orig = str(rl[vsc_col_map.get('origem', 0)]).strip() if vsc_col_map.get('origem') is not None and rl[vsc_col_map.get('origem', 0)] else ""
                            dest = str(rl[vsc_col_map.get('destino', 0)]).strip() if vsc_col_map.get('destino') is not None and rl[vsc_col_map.get('destino', 0)] else ""

                            if perc and (orig or dest):
                                self.previous_rateio_data.append({
                                    'origem': orig,
                                    'destino': dest,
                                    'cc': '',
                                    'percurso': perc,
                                    'vsc': 0,
                                    'total': 0,
                                    'utilizacao': 0,
                                    'arquivo': os.path.basename(filepath) + " [VSC]",
                                    'origem_city': normalize_city(extract_city_from_agilis_field(orig)),
                                    'destino_city': normalize_city(extract_city_from_agilis_field(dest)),
                                    'is_vsc_tab': True,
                                })

                wb.close()
            except Exception as e:
                print(f"    ⚠️ Erro ao ler {filepath}: {e}")

        print(f"  ✅ Total de registros anteriores: {len(self.previous_rateio_data)}")

    def _process_agilis_columns(self):
        e_idx = 4
        for row in self.agilis_data:
            e_val = str(row[e_idx]) if len(row) > e_idx and row[e_idx] else ""
            row.append(extract_origin_from_e(e_val))
            row.append(extract_destination_from_e(e_val))
            row.append(extract_centro_custo(e_val))
        self.agilis_headers.extend(["ORIGEM", "DESTINO", "CENTRO DE CUSTO AJUSTADO"])

    def _process_vsc(self):
        self.vsc_processed = []
        origem_idx = len(self.agilis_headers) - 3
        destino_idx = len(self.agilis_headers) - 2

        if self.resolvedor_cc is None and RESOLVER_AVAILABLE:
            print("  📊 Inicializando integração Web/Cache para buscar Origens e Destinos vazios...")
            self.resolvedor_cc = ResolvedorCC(use_selenium=SELENIUM_AVAILABLE)

        for vsc_item in self.vsc_data:
            texto = str(vsc_item['texto'])
            valor = vsc_item['valor']

            perc_matches = re.findall(r'PERC\s*[:\-]?\s*(\d+)', texto, re.IGNORECASE)
            if not perc_matches:
                perc_matches = re.findall(r'(\d{7,})', texto)

            percurso = ""
            if perc_matches:
                nums = sorted(set(int(n) for n in perc_matches), reverse=True)
                percurso = str(nums[1]) if len(nums) >= 2 else str(nums[0])

            origem_enc = ""
            destino_enc = ""
            
            if percurso:
                for row in self.agilis_data:
                    if len(row) > max(origem_idx, destino_idx):
                        orig_ag = str(row[origem_idx]) if row[origem_idx] else ""
                        dest_ag = str(row[destino_idx]) if row[destino_idx] else ""
                        perc_found = get_percurso_from_row(orig_ag, dest_ag)
                        if perc_found == percurso:
                            origem_enc = orig_ag
                            destino_enc = dest_ag
                            break

            if not origem_enc and not destino_enc and percurso:
                origem_enc, destino_enc = self._find_vsc_origin_dest_from_previous(percurso)

            if not origem_enc and not destino_enc and percurso and self.resolvedor_cc:
                cached = self.resolvedor_cc.cache.get(percurso, {})
                
                if cached.get('valor_origem') or cached.get('valor_destino'):
                    origem_enc = cached.get('valor_origem', '')
                    destino_enc = cached.get('valor_destino', '')
                    print(f"      💾 Origem/Destino recuperados do Cache: {origem_enc} → {destino_enc}")
                
                elif self.resolvedor_cc.use_selenium:
                    print(f"      🌐 Buscando Origem/Destino no site dos Correios para o percurso {percurso}...")
                    dados_web = self.resolvedor_cc._consultar_selenium(percurso)
                    
                    if dados_web:
                        origem_enc = dados_web.get('valor_origem', '')
                        destino_enc = dados_web.get('valor_destino', '')
                        
                        if origem_enc or destino_enc:
                            print(f"        ✅ Encontrado na Web: {origem_enc} → {destino_enc}")
                            self.resolvedor_cc._save_cache(
                                percurso, 
                                cached.get('cc', ''), 
                                cached.get('descricao', ''), 
                                cached.get('fonte', 'selenium_orig_dest'), 
                                extra=dados_web
                            )

            self.vsc_processed.append({
                'percurso': percurso,
                'origem': origem_enc,
                'destino': destino_enc,
                'valor': valor,
                'has_credito': vsc_item['has_credito'],
                'has_debito': vsc_item['has_debito'],
                'texto_original': texto
            })

        print(f"  ✅ VSC: {len(self.vsc_processed)} processados")
        for v in self.vsc_processed:
            status = "✅" if v['origem'] or v['destino'] else "⚠️ sem origem/destino"
            print(f"      {status} Perc:{v['percurso']} Orig:{v['origem']} Dest:{v['destino']} Val:{v['valor']}")

    def _find_vsc_origin_dest_from_previous(self, percurso):
        for prev in self.previous_rateio_data:
            if prev.get('is_vsc_tab') and prev['percurso'] == percurso:
                if prev['origem'] or prev['destino']:
                    print(f"      🔍 VSC anterior encontrado: {prev['origem']} → {prev['destino']} ({prev['arquivo']})")
                    return prev['origem'], prev['destino']

        for prev in self.previous_rateio_data:
            if prev['percurso'] == percurso and (prev['origem'] or prev['destino']):
                print(f"      🔍 Rateio anterior encontrado: {prev['origem']} → {prev['destino']} ({prev['arquivo']})")
                return prev['origem'], prev['destino']

        return "", ""

    def _create_pivot_tables(self):
        cartao_col = None
        municipio_col = None
        valor_col = None

        for i, h in enumerate(self.correios_malote_headers):
            hl = h.lower().strip()
            if ("numero" in hl or "nº" in hl) and "cart" in hl:
                cartao_col = i
            elif "cartao" in hl or "cartão" in hl:
                if cartao_col is None:
                    cartao_col = i
            if "munic" in hl and "origem" in hl:
                municipio_col = i
            if "valor" in hl and "serv" in hl:
                valor_col = i

        if cartao_col is None:
            for i, h in enumerate(self.correios_malote_headers):
                if "postagem" in h.lower():
                    cartao_col = i
                    break

        if municipio_col is None:
            for i, h in enumerate(self.correios_malote_headers):
                if "municipio" in h.lower() or "município" in h.lower():
                    municipio_col = i
                    break

        if valor_col is None:
            for i, h in enumerate(self.correios_malote_headers):
                if "valor" in h.lower():
                    valor_col = i
                    break

        if cartao_col is not None:
            samples = []
            for row in self.correios_malote_data[:20]:
                if len(row) > cartao_col and row[cartao_col]:
                    samples.append(str(row[cartao_col]).strip())
            if samples and not all(re.match(r'^\d{7,8}$', s) for s in samples if s):
                for ct in range(len(self.correios_malote_headers)):
                    tv = [str(row[ct]).strip() for row in self.correios_malote_data[:20]
                          if len(row) > ct and row[ct]]
                    if tv and sum(1 for v in tv if re.match(r'^\d{7,8}$', v)) > len(tv) * 0.5:
                        cartao_col = ct
                        break

        print(f"  Colunas → Cartão:{cartao_col} Município:{municipio_col} Valor:{valor_col}")

        if cartao_col is not None and municipio_col is not None and valor_col is not None:
            for row in self.correios_malote_data:
                if len(row) > max(cartao_col, municipio_col, valor_col):
                    cr = row[cartao_col]
                    cartao = str(int(cr)) if isinstance(cr, (int, float)) else str(cr).strip() if cr else ""
                    municipio = str(row[municipio_col]).upper().strip() if row[municipio_col] else ""

                    try:
                        valor = float(row[valor_col]) if row[valor_col] else 0
                    except:
                        try:
                            valor = float(str(row[valor_col]).replace(',', '.').replace('R$', '').strip())
                        except:
                            valor = 0

                    if cartao and municipio:
                        if cartao not in self.pivot_correios:
                            self.pivot_correios[cartao] = OrderedDict()
                            self.pivot_correios_totals[cartao] = 0
                        if municipio not in self.pivot_correios[cartao]:
                            self.pivot_correios[cartao][municipio] = 0
                        self.pivot_correios[cartao][municipio] += valor
                        self.pivot_correios_totals[cartao] += valor

            print(f"  ✅ Pivot Correios: {len(self.pivot_correios)} cartões")

        origem_idx = len(self.agilis_headers) - 3
        destino_idx = len(self.agilis_headers) - 2
        cc_idx = len(self.agilis_headers) - 1

        for row in self.agilis_data:
            if len(row) > max(origem_idx, destino_idx, cc_idx):
                orig = str(row[origem_idx]).strip() if row[origem_idx] else ""
                dest = str(row[destino_idx]).strip() if row[destino_idx] else ""
                cc = str(row[cc_idx]).strip() if row[cc_idx] else ""

                if orig not in self.pivot_agilis:
                    self.pivot_agilis[orig] = OrderedDict()
                if dest not in self.pivot_agilis[orig]:
                    self.pivot_agilis[orig][dest] = OrderedDict()
                if cc not in self.pivot_agilis[orig][dest]:
                    self.pivot_agilis[orig][dest][cc] = 0
                self.pivot_agilis[orig][dest][cc] += 1

        print(f"  ✅ Pivot Agilis: {len(self.pivot_agilis)} origens")

    def _find_cc_from_previous_by_percurso(self, percurso, origem="", destino=""):
        if not percurso:
            return "", False

        candidates = []
        for prev in self.previous_rateio_data:
            if prev.get('is_vsc_tab'):
                continue
            if prev['percurso'] != percurso or not prev['cc']:
                continue
            candidates.append({
                'cc': prev['cc'],
                'total': abs(prev['total']) if prev['total'] else 0,
                'vsc': abs(prev['vsc']) if prev['vsc'] else 0,
                'arquivo': prev['arquivo'],
            })

        if not candidates:
            return "", False

        cc_totals = defaultdict(lambda: {'total': 0, 'vsc': 0, 'count': 0, 'arquivo': ''})
        for c in candidates:
            cc = c['cc']
            cc_totals[cc]['total'] += c['total']
            cc_totals[cc]['vsc'] += c['vsc']
            cc_totals[cc]['count'] += 1
            if not cc_totals[cc]['arquivo']:
                cc_totals[cc]['arquivo'] = c['arquivo']

        best = sorted(cc_totals.items(),
                       key=lambda x: (x[1]['total'] + x[1]['vsc'], x[1]['count']),
                       reverse=True)

        if best:
            cc_found = best[0][0]
            info = best[0][1]
            print(f"      🔍 CC por percurso exato: {cc_found} "
                  f"(despesa:{info['total']:.2f}, vsc:{info['vsc']:.2f}, {info['arquivo']})")
            return cc_found, True

        return "", False

    def _find_cc_from_previous_by_cities(self, origem, destino):
        if not origem or not destino:
            return "", False
            
        orig_norm = normalize_city(extract_city_from_agilis_field(origem))
        dest_norm = normalize_city(extract_city_from_agilis_field(destino))
        
        if not orig_norm or not dest_norm:
            return "", False

        candidates = []
        for prev in self.previous_rateio_data:
            if prev.get('is_vsc_tab'):
                continue
            if not prev['cc']:
                continue
            
            p_orig_norm = prev.get('origem_city', '')
            p_dest_norm = prev.get('destino_city', '')
            
            if p_orig_norm == orig_norm and p_dest_norm == dest_norm:
                candidates.append({
                    'cc': prev['cc'],
                    'total': abs(prev['total']) if prev['total'] else 0,
                    'vsc': abs(prev['vsc']) if prev['vsc'] else 0,
                    'arquivo': prev['arquivo'],
                })

        if not candidates:
            return "", False

        cc_totals = defaultdict(lambda: {'total': 0, 'vsc': 0, 'count': 0, 'arquivo': ''})
        for c in candidates:
            cc = c['cc']
            cc_totals[cc]['total'] += c['total']
            cc_totals[cc]['vsc'] += c['vsc']
            cc_totals[cc]['count'] += 1
            if not cc_totals[cc]['arquivo']:
                cc_totals[cc]['arquivo'] = c['arquivo']

        best = sorted(cc_totals.items(),
                       key=lambda x: (x[1]['count'], x[1]['total'] + x[1]['vsc']),
                       reverse=True)

        if best:
            cc_found = best[0][0]
            info = best[0][1]
            print(f"      🔍 CC por Rota (Origem->Destino): {cc_found} "
                  f"(usado {info['count']} vezes, {info['arquivo']})")
            return cc_found, True

        return "", False

    def _create_rateio(self):
        self.rateio_data = []

        for cartao in self.pivot_correios.keys():
            self.status_cartoes[cartao] = "ok"

        pivot_percurso = self._build_pivot_percurso()
        vsc_por_percurso = self._build_vsc_summary()

        for cartao, municipios_correios in self.pivot_correios.items():
            print(f"\n  🔄 Cartão {cartao}: {dict(municipios_correios)}")
            start_idx = len(self.rateio_data)

            vsc_valor, vsc_credito, has_credito, vsc_sp_injetado, is_sp_exc = self._get_vsc_state(cartao, vsc_por_percurso)

            self._distribute_municipality_pairs(cartao, municipios_correios, pivot_percurso, vsc_valor)
            self._apply_sp_fallback(cartao, municipios_correios, pivot_percurso, start_idx)
            self._sweep_orphan_values(cartao, municipios_correios, pivot_percurso, start_idx)

            linhas_deste_cartao = self.rateio_data[start_idx:]
            has_valid_cc = any(r.get('centro_custo') and r.get('centro_custo') != "CC NÃO ENCONTRADO" for r in linhas_deste_cartao)
            
            if not has_valid_cc:
                for r in linhas_deste_cartao:
                    orig_str = str(r.get('origem', '')).upper()
                    dest_str = str(r.get('destino', '')).upper()
                    if 'PAULO' in orig_str or 'PAULO' in dest_str or 'MOGI' in orig_str or 'MOGI' in dest_str:
                        self.status_cartoes[cartao] = "N/L valor rateado entre o percurso SP - SP"
                        break

        self._absorb_global_sp_orphans(pivot_percurso)
        self._compact_rateio_data()

        print(f"\n  🔄 Pós-processamento: VSC sem CC...")
        self._process_vsc_without_cc()

        print(f"\n  ✅ Rateio: {len(self.rateio_data)} linhas")

    def _build_pivot_percurso(self):
        pivot_percurso = defaultdict(list)
        for orig_raw, destinos in self.pivot_agilis.items():
            for dest_raw, ccs in destinos.items():
                perc = get_percurso_from_row(orig_raw, dest_raw)
                if not perc: continue
                orig_city = normalize_city(extract_city_from_agilis_field(orig_raw))
                dest_city = normalize_city(extract_city_from_agilis_field(dest_raw))
                for cc, count in ccs.items():
                    pivot_percurso[perc].append({
                        'origem': orig_raw, 'destino': dest_raw, 'cc': cc,
                        'count': count, 'orig_city': orig_city, 'dest_city': dest_city,
                    })
        print(f"\n  📊 Pivot por percurso: {len(pivot_percurso)} percursos")
        return pivot_percurso

    def _build_vsc_summary(self):
        vsc_por_percurso = {}
        for vi in self.vsc_processed:
            perc = vi['percurso']
            if not perc: continue
            if perc not in vsc_por_percurso:
                vsc_por_percurso[perc] = {'debito': 0, 'credito': 0}
            val = 0
            try: val = float(vi['valor']) if vi['valor'] else 0
            except:
                try: val = float(str(vi['valor']).replace(',', '.').replace('R$', '').strip())
                except: pass

            if vi['has_credito']: vsc_por_percurso[perc]['credito'] += abs(val)
            else: vsc_por_percurso[perc]['debito'] += val
        return vsc_por_percurso

    def _get_vsc_state(self, cartao, vsc_por_percurso):
        vsc_info = vsc_por_percurso.get(cartao, {'debito': 0, 'credito': 0})
        vsc_credito_valor = vsc_info['credito']
        vsc_has_credito = vsc_credito_valor > 0
        
        is_sp_exception = (cartao == '10320249')
        vsc_valor = vsc_info['debito']
        vsc_sp_injetado = 0

        if is_sp_exception:
            if vsc_valor >= 1205.84:
                vsc_sp_injetado = 1205.84
                vsc_valor -= 1205.84
            elif vsc_valor > 0:
                vsc_sp_injetado = vsc_valor
                vsc_valor = 0
                
        return vsc_valor, vsc_credito_valor, vsc_has_credito, vsc_sp_injetado, is_sp_exception

    def _distribute_municipality_pairs(self, cartao, municipios_correios, pivot_percurso, vsc_valor):
        mun_list = list(municipios_correios.keys())
        mun_norms = {m: normalize_city(m) for m in mun_list}
        linhas_percurso = pivot_percurso.get(cartao, [])
        
        linhas_diferentes_global = [l for l in linhas_percurso if l['orig_city'] != l['dest_city']]
        if linhas_diferentes_global:
            linhas_percurso = linhas_diferentes_global
            
        soma_total_cartao = sum(l['count'] for l in linhas_percurso)

        if len(mun_list) >= 2: pares = list(combinations(range(len(mun_list)), 2))
        elif len(mun_list) == 1: pares = [(0, 0)]
        else: return

        municipios_processados_ok = set()
        valores_pendentes = {m: float(v) for m, v in municipios_correios.items()}

        for idx_a, idx_b in pares:
            mun_a = mun_list[idx_a]
            mun_b = mun_list[idx_b]
            mun_a_norm = mun_norms[mun_a]
            mun_b_norm = mun_norms[mun_b]
            
            valor_a = valores_pendentes[mun_a]
            valor_b = valores_pendentes[mun_b]

            linhas_a_b = []
            linhas_b_a = []

            if idx_a == idx_b:
                linhas_a_b = list(linhas_percurso)
            else:
                for linha in linhas_percurso:
                    oc = linha['orig_city']
                    dc = linha['dest_city']
                    if oc == mun_a_norm and dc == mun_b_norm: linhas_a_b.append(linha)
                    elif oc == mun_b_norm and dc == mun_a_norm: linhas_b_a.append(linha)
                    elif oc == mun_a_norm: linhas_a_b.append(linha)
                    elif oc == mun_b_norm: linhas_b_a.append(linha)
                    elif dc == mun_a_norm: linhas_b_a.append(linha)
                    elif dc == mun_b_norm: linhas_a_b.append(linha)

            soma_a_b = sum(l['count'] for l in linhas_a_b)
            soma_b_a = sum(l['count'] for l in linhas_b_a)
            valor_a_usar = valor_a
            valor_b_usar = valor_b

            if not linhas_a_b and linhas_b_a:
                valor_b_usar = valor_b + valor_a
                valor_a_usar = 0
            elif not linhas_b_a and linhas_a_b:
                if idx_a == idx_b: valor_a_usar = valor_a
                else: valor_a_usar = valor_a + valor_b
                valor_b_usar = 0

            if valor_a_usar > 0 and linhas_a_b:
                valores_pendentes[mun_a] -= (valor_a_usar if valor_a_usar <= valor_a else valor_a)
                if valor_a_usar > valor_a: valores_pendentes[mun_b] -= (valor_a_usar - valor_a)

            if valor_b_usar > 0 and linhas_b_a:
                valores_pendentes[mun_b] -= (valor_b_usar if valor_b_usar <= valor_b else valor_b)
                if valor_b_usar > valor_b: valores_pendentes[mun_a] -= (valor_b_usar - valor_b)

            if linhas_a_b and valor_a_usar > 0:
                for linha in linhas_a_b:
                    count = linha['count']
                    utilizacao = (valor_a_usar / soma_a_b) * count if soma_a_b > 0 else 0
                    vsc_calc = round((vsc_valor / soma_total_cartao) * count, 2) if vsc_valor > 0 and soma_total_cartao > 0 else ''
                    self._append_rateio_row(linha['origem'], linha['destino'], linha['cc'], count, round(utilizacao, 2), '', vsc_calc, '')
                municipios_processados_ok.add(mun_a)

            if linhas_b_a and valor_b_usar > 0:
                for linha in linhas_b_a:
                    count = linha['count']
                    utilizacao = (valor_b_usar / soma_b_a) * count if soma_b_a > 0 else 0
                    vsc_calc = round((vsc_valor / soma_total_cartao) * count, 2) if vsc_valor > 0 and soma_total_cartao > 0 else ''
                    self._append_rateio_row(linha['origem'], linha['destino'], linha['cc'], count, round(utilizacao, 2), '', vsc_calc, '')
                municipios_processados_ok.add(mun_b)

            if not linhas_a_b and not linhas_b_a:
                if mun_a not in municipios_processados_ok and valor_a > 0:
                    vsc_nl = f"VSC s/ CC - {cartao}" if vsc_valor > 0 else ''
                    self._append_rateio_row(mun_a, mun_b, '', '', '', round(valor_a, 2), '', vsc_nl)
                    municipios_processados_ok.add(mun_a)
                    valores_pendentes[mun_a] = 0

                if mun_b not in municipios_processados_ok and idx_a != idx_b and valor_b > 0:
                    vsc_nl = f"VSC s/ CC - {cartao}" if vsc_valor > 0 else ''
                    self._append_rateio_row(mun_b, mun_a, '', '', '', round(valor_b, 2), '', vsc_nl)
                    municipios_processados_ok.add(mun_b)
                    valores_pendentes[mun_b] = 0

    def _apply_sp_fallback(self, cartao, municipios_correios, pivot_percurso, start_idx):
        def _to_float(x):
            try: return float(x) if x not in ('', None) else 0.0
            except: return 0.0

        valor_sp_correios = sum(_to_float(v) for m, v in municipios_correios.items() if normalize_city(m) == "SAO PAULO")
        if valor_sp_correios > 0:
            linhas_deste_cartao = self.rateio_data[start_idx:]
            valor_ja_distribuido_cartao = sum(_to_float(r.get('utilizacao')) + _to_float(r.get('nao_localizado')) for r in linhas_deste_cartao)
            valor_total_correios = _to_float(self.pivot_correios_totals.get(cartao, 0))
            faltante = round(valor_total_correios - valor_ja_distribuido_cartao, 2)

            if faltante > 0.01:
                candidatas_sp_sp = [l for l in pivot_percurso.get(cartao, []) if l.get('orig_city') == "SAO PAULO" and l.get('dest_city') == "SAO PAULO"]
                candidatas = candidatas_sp_sp or list(pivot_percurso.get(cartao, []))

                if candidatas:
                    soma_counts = sum(int(l.get('count', 0) or 0) for l in candidatas)
                    valor_restante = faltante
                    for i, l in enumerate(candidatas):
                        count = int(l.get('count', 0) or 0)
                        if i == len(candidatas) - 1: parcela = round(valor_restante, 2)
                        else:
                            parcela = round((faltante / soma_counts) * count, 2) if soma_counts > 0 else round(faltante / len(candidatas), 2)
                        valor_restante -= parcela
                        self._append_rateio_row(l['origem'], l['destino'], l['cc'], l['count'], parcela, '', '', '')
                    print(f"      🧩 Fallback UTILIZAÇÃO aplicado (SP pivot) no cartão {cartao}: R$ {faltante} distribuído.")

    def _sweep_orphan_values(self, cartao, municipios_correios, pivot_percurso, start_idx):
        valor_total_correios = self.pivot_correios_totals.get(cartao, 0)
        linhas_deste_cartao = self.rateio_data[start_idx:]
        
        linhas_vazias = [r for r in linhas_deste_cartao if not r.get('centro_custo') and (r.get('qnt_chamados') == '' or r.get('qnt_chamados') == 0)]
        linhas_validas = [r for r in linhas_deste_cartao if r.get('centro_custo')]

        if linhas_validas and linhas_vazias:
            print(f"      🧹 Removendo {len(linhas_vazias)} linha(s) vazia(s) para redistribuir o valor.")
            for linha_vazia in linhas_vazias:
                self.rateio_data.remove(linha_vazia)
        
        linhas_deste_cartao_atualizadas = self.rateio_data[start_idx:]
        valor_ja_distribuido = sum(
            (r['utilizacao'] if isinstance(r['utilizacao'], (int, float)) else 0) +
            (r['nao_localizado'] if isinstance(r['nao_localizado'], (int, float)) else 0)
            for r in linhas_deste_cartao_atualizadas
        )
        
        valor_orfao = round(valor_total_correios - valor_ja_distribuido, 2)
        
        if valor_orfao > 0.01 or valor_orfao < -0.01: 
            print(f"      💰 Valor órfão detectado: R${valor_orfao}")
            
            if linhas_validas:
                print(f"      🎯 Distribuindo R${valor_orfao} proporcionalmente entre os CCs válidos.")
                soma_chamados = sum(r.get('qnt_chamados', 0) for r in linhas_validas if isinstance(r.get('qnt_chamados'), (int, float)))
                valor_restante = valor_orfao
                
                for i, r in enumerate(linhas_validas):
                    count = r.get('qnt_chamados', 0) if isinstance(r.get('qnt_chamados'), (int, float)) else 0
                    
                    if soma_chamados > 0:
                        parcela = round((valor_orfao / soma_chamados) * count, 2)
                    else:
                        parcela = round(valor_orfao / len(linhas_validas), 2)

                    if i == len(linhas_validas) - 1: 
                        parcela = round(valor_restante, 2)
                        
                    valor_restante -= parcela
                    
                    atual = r['nao_localizado'] if isinstance(r['nao_localizado'], (int, float)) else 0
                    r['nao_localizado'] = round(atual + parcela, 2)
            else:
                self.status_cartoes[cartao] = "N/L valor rateado entre o percurso SP - SP"
                                
                mun_principal = list(municipios_correios.keys())[0] if municipios_correios else f"PERCURSO {cartao}"
                origem_str = f"{mun_principal}/XX - {cartao}"
                destino_str = f"{mun_principal}/XX - {cartao}"
                
                print(f"      ⚠️ Criando linha de aviso para absorver R${valor_orfao}.")
                self._append_rateio_row(origem_str, destino_str, "CC NÃO ENCONTRADO", 0, '', valor_orfao, '', '', highlight=True)

    def _apply_vsc_credits_and_exceptions(self, cartao, start_idx, vsc_credito_valor, vsc_has_credito, vsc_sp_injetado, is_sp_exception):
        if vsc_has_credito and vsc_credito_valor > 0:
            linhas_perc = self.rateio_data[start_idx:]
            if linhas_perc:
                soma_ch = sum(r['qnt_chamados'] for r in linhas_perc if isinstance(r['qnt_chamados'], (int, float)))
                if soma_ch > 0:
                    for r in linhas_perc:
                        if isinstance(r['qnt_chamados'], (int, float)):
                            r['desconto'] = round((vsc_credito_valor / soma_ch) * r['qnt_chamados'], 2)

        if is_sp_exception and vsc_sp_injetado > 0:
            print(f"      ⭐ Aplicando exceção de VSC para SP: R${vsc_sp_injetado} -> MRVHSP3023")
            self._append_rateio_row(f'SAO PAULO/SP - {cartao}', 'BELO HORIZONTE/MG -SEDE', 'MRVHSP3023', 0, '', '', round(vsc_sp_injetado, 2), '')
            self.vsc_resolvido_obra.add(cartao)

    def _absorb_global_sp_orphans(self, pivot_percurso):
        print("\n  🔄 Redistribuindo valores órfãos de SP para a rota SP->SP (10320249)...")
        linhas_orfas_sp = []
        valor_orfao_total = 0.0
        
        for r in self.rateio_data:
            if not r.get('centro_custo'):
                orig_str = str(r.get('origem', '')).upper()
                dest_str = str(r.get('destino', '')).upper()
                if 'PAULO' in orig_str or 'PAULO' in dest_str or 'MOGI' in orig_str or 'MOGI' in dest_str:
                    linhas_orfas_sp.append(r)
                    val = r.get('nao_localizado', 0)
                    if isinstance(val, (int, float)): valor_orfao_total += val

        if valor_orfao_total > 0:
            for r in linhas_orfas_sp:
                if r in self.rateio_data: self.rateio_data.remove(r)
            
            linhas_alvo_sp = [r for r in self.rateio_data if r.get('centro_custo') and '10320249' in str(r.get('origem', '')) and '10320249' in str(r.get('destino', ''))]
            
            if not linhas_alvo_sp:
                print("      ⚠️ Linhas SP->SP não encontradas no rateio atual. Resgatando do Agilis...")
                for l in pivot_percurso.get('10320249', []):
                    if l['orig_city'] == 'SAO PAULO' and l['dest_city'] == 'SAO PAULO':
                        self._append_rateio_row(l['origem'], l['destino'], l['cc'], l['count'], '', 0, '', '')
                        linhas_alvo_sp.append(self.rateio_data[-1])
            
            if linhas_alvo_sp:
                soma_chamados_sp = sum(r.get('qnt_chamados', 0) for r in linhas_alvo_sp if isinstance(r.get('qnt_chamados'), (int, float)))
                if soma_chamados_sp > 0:
                    valor_restante = valor_orfao_total
                    for i, r in enumerate(linhas_alvo_sp):
                        qnt = r.get('qnt_chamados', 0)
                        if i == len(linhas_alvo_sp) - 1: acrescimo = round(valor_restante, 2)
                        else:
                            acrescimo = round((valor_orfao_total / soma_chamados_sp) * qnt, 2)
                            valor_restante -= acrescimo
                        
                        atual = r.get('nao_localizado', 0)
                        if not isinstance(atual, (int, float)): atual = 0
                        r['nao_localizado'] = round(atual + acrescimo, 2)
                    print(f"      ✅ R$ {valor_orfao_total} redistribuídos entre {len(linhas_alvo_sp)} CCs da rota SP->SP.")

    def _compact_rateio_data(self):
        print("\n  🔄 Compactando linhas duplicadas...")
        aggregated = OrderedDict()
        for r in self.rateio_data:
            key = (r.get('origem', ''), r.get('destino', ''), r.get('centro_custo', ''))
            if key not in aggregated:
                aggregated[key] = r.copy()
            else:
                for field in ['utilizacao', 'nao_localizado', 'vsc', 'desconto']:
                    val1 = aggregated[key][field]
                    val2 = r[field]
                    v1 = float(val1) if val1 != '' and val1 is not None else 0
                    v2 = float(val2) if val2 != '' and val2 is not None else 0
                    soma = v1 + v2
                    if val1 == '' and val2 == '': aggregated[key][field] = ''
                    else: aggregated[key][field] = round(soma, 2)
                
                c1 = aggregated[key]['qnt_chamados']
                c2 = r['qnt_chamados']
                v_c1 = int(c1) if c1 != '' and c1 is not None else 0
                v_c2 = int(c2) if c2 != '' and c2 is not None else 0
                if c1 == '' and c2 == '': aggregated[key]['qnt_chamados'] = ''
                else: aggregated[key]['qnt_chamados'] = max(v_c1, v_c2)

        self.rateio_data = list(aggregated.values())

    def _append_rateio_row(self, origem, destino, cc, chamados, utilizacao, nao_localizado, vsc, vsc_nao_localizado, highlight=False):
        self.rateio_data.append({
            'origem': origem, 'destino': destino, 'centro_custo': cc,
            'qnt_chamados': chamados, 'utilizacao': utilizacao,
            'nao_localizado': nao_localizado, 'vsc': vsc,
            'vsc_nao_localizado': vsc_nao_localizado, 'desconto': '',
            'total': 0, 'obra': '', 'cc_sub': '', '_highlight': highlight
        })

    def _process_vsc_without_cc(self):
        if self.resolvedor_cc is None and RESOLVER_AVAILABLE:
            print("\n  📊 Inicializando resolvedor de CC...")
            self.resolvedor_cc = ResolvedorCC(use_selenium=SELENIUM_AVAILABLE)

        vsc_sem_cc = self._get_unprocessed_vscs()
        if not vsc_sem_cc:
            print("    ✅ Todos os VSC já têm CC distribuído")
            return

        print(f"\n    📋 {len(vsc_sem_cc)} VSCs sem CC para resolver:")

        for item in vsc_sem_cc:
            percurso = item['percurso']
            origem   = item['origem']
            destino  = item['destino']
            vsc_val  = item['valor']

            print(f"\n    VSC: Perc={percurso} Orig={origem} Dest={destino} Val={vsc_val}")

            cc_found, desc_found, fonte = self._resolve_cc_for_vsc(percurso, origem, destino)
            cc_was_found = bool(cc_found)

            if cc_was_found:
                self.vsc_resolvido_obra.add(percurso)

            if not cc_was_found:
                if self._distribute_vsc_to_same_city_lines(percurso, vsc_val):
                    continue 

                print(f"      ⚠️ Nenhuma linha encontrada para percurso {percurso} → criando linha nova")

            vsc_nl = f"VSC s/ CC - {percurso}" if not cc_found else ''
            self._append_rateio_row(
                origem if origem else f"PERCURSO {percurso}",
                destino if destino else f"PERCURSO {percurso}",
                cc_found, 0, '', '', round(vsc_val, 2), vsc_nl, highlight=not cc_was_found
            )

            if cc_found: print(f"      ✅ CC: {cc_found} ({desc_found}) via {fonte}")
            else: print(f"      ⚠️ CC não encontrado - marcado para revisão")

        if self.resolvedor_cc:
            self.resolvedor_cc.close()

    def _get_unprocessed_vscs(self):
        percursos_no_rateio = set()
        for rd in self.rateio_data:
            perc = get_percurso_from_row(rd.get('origem', ''), rd.get('destino', ''))
            if perc and (rd.get('vsc') and rd['vsc'] != '' and rd['vsc'] != 0):
                percursos_no_rateio.add(perc)

        vsc_sem_cc = []
        for vsc_item in self.vsc_processed:
            if vsc_item['has_credito']: continue
            percurso = vsc_item['percurso']
            if not percurso or percurso in percursos_no_rateio: continue

            try: vsc_val = float(vsc_item['valor']) if vsc_item['valor'] else 0
            except:
                try: vsc_val = float(str(vsc_item['valor']).replace(',', '.').replace('R$', '').strip())
                except: vsc_val = 0

            if vsc_val > 0:
                vsc_sem_cc.append({'percurso': percurso, 'origem': vsc_item['origem'], 'destino': vsc_item['destino'], 'valor': vsc_val})
        return vsc_sem_cc

    def _resolve_cc_for_vsc(self, percurso, origem, destino):
        nome_obra = ""
        if destino:
            nome_obra = re.sub(r'\s*/\s*\w+', '', destino)
            nome_obra = re.sub(r'\s*-\s*\d+', '', nome_obra)
            nome_obra = re.sub(r'\s*-\s*SEDE', '', nome_obra, flags=re.IGNORECASE).strip()

        cc_found, desc_found, fonte = "", "", ""

        if self.resolvedor_cc:
            cc_found, desc_found, fonte = self.resolvedor_cc.resolver_cc(percurso, nome_obra)

        if not cc_found:
            cc_found, _ = self._find_cc_from_previous_by_percurso(percurso, origem, destino)
            if cc_found: fonte = "rateio_anterior_percurso"

        if not cc_found and origem and destino:
            cc_found, _ = self._find_cc_from_previous_by_cities(origem, destino)
            if cc_found: fonte = "rateio_anterior_rota"

        return cc_found, desc_found, fonte

    def _distribute_vsc_to_same_city_lines(self, percurso, vsc_val):
        def _city(txt): return normalize_city(extract_city_from_agilis_field(str(txt)))

        linhas_mesmo_percurso = [r for r in self.rateio_data if percurso in str(r.get('origem', '')) or percurso in str(r.get('destino', ''))]
        linhas_diferentes = [r for r in linhas_mesmo_percurso if _city(r.get('origem', '')) != _city(r.get('destino', ''))]
        linhas_iguais = [r for r in linhas_mesmo_percurso if _city(r.get('origem', '')) == _city(r.get('destino', ''))]

        if linhas_diferentes:
            print(f"      ℹ️ Percurso {percurso} tem linhas com cidades diferentes → VSC R${vsc_val} não redistribuído (já coberto pelo fluxo normal)")
            return True 

        if linhas_iguais:
            candidatas = linhas_iguais
            soma_ch = sum(r['qnt_chamados'] for r in candidatas if isinstance(r['qnt_chamados'], (int, float)) and r['qnt_chamados'] > 0)
            valor_restante = vsc_val
            for i, r in enumerate(candidatas):
                count = r['qnt_chamados'] if isinstance(r['qnt_chamados'], (int, float)) else 0
                if soma_ch > 0:
                    if i == len(candidatas) - 1: acrescimo = round(valor_restante, 2)
                    else: acrescimo = round((vsc_val / soma_ch) * count, 2); valor_restante -= acrescimo
                else:
                    if i == len(candidatas) - 1: acrescimo = round(valor_restante, 2)
                    else: acrescimo = round(vsc_val / len(candidatas), 2); valor_restante -= acrescimo

                vsc_atual = r['vsc'] if isinstance(r['vsc'], (int, float)) else 0
                r['vsc'] = round(vsc_atual + acrescimo, 2)
                if r.get('vsc_nao_localizado'): r['vsc_nao_localizado'] = ''

            print(f"      ♻️ VSC R${vsc_val} redistribuído em {len(candidatas)} linha(s) de mesma cidade do percurso {percurso}")
            return True
            
        return False    

    def _save(self):
        wb = openpyxl.Workbook()

        ws_cor = wb.active
        ws_cor.title = "Relatório Correios"

        for ci, h in enumerate(self.correios_malote_headers, 1):
            c = ws_cor.cell(row=1, column=ci, value=h)
            c.font = HEADER_FONT; c.fill = LIGHT_BLUE_FILL; c.border = THIN_BORDER

        for ri, rd in enumerate(self.correios_malote_data, 2):
            for ci, v in enumerate(rd, 1):
                ws_cor.cell(row=ri, column=ci, value=v)

        pc = len(self.correios_malote_headers) + 3
        c = ws_cor.cell(row=1, column=pc, value="Rótulos de Linha")
        c.font = HEADER_FONT; c.fill = LIGHT_BLUE_FILL; c.border = THIN_BORDER
        c = ws_cor.cell(row=1, column=pc + 1, value="Soma de Valor do Servico")
        c.font = HEADER_FONT; c.fill = LIGHT_BLUE_FILL; c.border = THIN_BORDER

        pr = 2
        gt = 0
        for cartao in sorted(self.pivot_correios.keys()):
            total = self.pivot_correios_totals[cartao]
            gt += total
            c = ws_cor.cell(row=pr, column=pc, value=f"⊟ {cartao}")
            c.font = Font(bold=True); c.border = THIN_BORDER
            c = ws_cor.cell(row=pr, column=pc + 1, value=round(total, 2))
            c.font = Font(bold=True); c.number_format = '#,##0.00'
            c.border = THIN_BORDER; c.alignment = Alignment(horizontal='right')

            status_msg = self.status_cartoes.get(cartao, "ok")
            ws_cor.cell(row=pr, column=pc + 2, value=status_msg)

            pr += 1
            for mun in sorted(self.pivot_correios[cartao].keys()):
                val = self.pivot_correios[cartao][mun]
                c = ws_cor.cell(row=pr, column=pc, value=f"    {mun}")
                c.border = THIN_BORDER
                c = ws_cor.cell(row=pr, column=pc + 1, value=round(val, 2))
                c.number_format = '#,##0.00'; c.border = THIN_BORDER
                c.alignment = Alignment(horizontal='right')
                pr += 1

        c = ws_cor.cell(row=pr, column=pc, value="Total Geral")
        c.font = Font(bold=True); c.border = THIN_BORDER
        c = ws_cor.cell(row=pr, column=pc + 1, value=round(gt, 2))
        c.font = Font(bold=True); c.number_format = '#,##0.00'
        c.border = THIN_BORDER; c.alignment = Alignment(horizontal='right')

        ws_ag = wb.create_sheet("Relatório Agilis")
        num_orig = len(self.agilis_headers) - 3

        for ci, h in enumerate(self.agilis_headers, 1):
            c = ws_ag.cell(row=1, column=ci, value=h)
            c.font = HEADER_FONT; c.border = THIN_BORDER
            c.fill = YELLOW_FILL if ci > num_orig else LIGHT_BLUE_FILL

        for ri, rd in enumerate(self.agilis_data, 2):
            for ci, v in enumerate(rd, 1):
                ws_ag.cell(row=ri, column=ci, value=v)

        correios_percs = set(self.pivot_correios.keys())
        pcs = len(self.agilis_headers) + 2

        for ci, h in enumerate(["Rótulos de Linha", "DESTINO",
                                 "CENTRO DE CUSTO AJUSTADO",
                                 "Contagem de Identificação da solicitação"]):
            c = ws_ag.cell(row=1, column=pcs + ci, value=h)
            c.font = HEADER_FONT; c.fill = LIGHT_BLUE_FILL; c.border = THIN_BORDER

        pr = 2
        tg = 0
        for origem in sorted(self.pivot_agilis.keys()):
            for destino in sorted(self.pivot_agilis[origem].keys()):
                perc = get_percurso_from_row(origem, destino)
                is_match = perc in correios_percs if perc else False

                for cc in sorted(self.pivot_agilis[origem][destino].keys()):
                    count = self.pivot_agilis[origem][destino][cc]
                    tg += count

                    cells = []
                    cells.append(ws_ag.cell(row=pr, column=pcs, value=origem))
                    cells.append(ws_ag.cell(row=pr, column=pcs + 1, value=destino))
                    cells.append(ws_ag.cell(row=pr, column=pcs + 2, value=cc))
                    cells.append(ws_ag.cell(row=pr, column=pcs + 3, value=count))

                    for cell in cells:
                        cell.border = THIN_BORDER
                        if is_match:
                            cell.fill = YELLOW_FILL

                    pr += 1

        c = ws_ag.cell(row=pr, column=pcs, value="Total Geral")
        c.font = Font(bold=True); c.border = THIN_BORDER
        for ci in range(1, 3):
            ws_ag.cell(row=pr, column=pcs + ci, value="").border = THIN_BORDER
        c = ws_ag.cell(row=pr, column=pcs + 3, value=tg)
        c.font = Font(bold=True); c.border = THIN_BORDER

        ws_vsc = wb.create_sheet("VSC")
        for ci, h in enumerate(["Percurso", "Origem", "Destino", "Valor"], 1):
            c = ws_vsc.cell(row=1, column=ci, value=h)
            c.font = HEADER_FONT; c.fill = LIGHT_BLUE_FILL; c.border = THIN_BORDER

        for ri, vi in enumerate(self.vsc_processed, 2):
            ws_vsc.cell(row=ri, column=1, value=vi['percurso']).border = THIN_BORDER
            c2 = ws_vsc.cell(row=ri, column=2, value=vi['origem'])
            c2.border = THIN_BORDER
            c3 = ws_vsc.cell(row=ri, column=3, value=vi['destino'])
            c3.border = THIN_BORDER
            c = ws_vsc.cell(row=ri, column=4, value=vi['valor'])
            c.border = THIN_BORDER; c.number_format = 'R$ #,##0.00'

            if not vi['origem'] and not vi['destino']:
                c2.fill = ORANGE_FILL
                c3.fill = ORANGE_FILL

            if vi['percurso'] in self.vsc_resolvido_obra:
                ws_vsc.cell(row=ri, column=5, value="Valor rateado para o centro de custo da obra/escritório")

        ws_rat = wb.create_sheet("Rateio")
        rateio_headers = [
            "ORIGEM", "DESTINO", "CENTRODECUSTO", "QNT DE CHAMADOS",
            "UTILIZAÇÃO", " NÃO LOCALIZADO ", " VSC ",
            "VSC não localizado", "  DESCONTO  ", " TOTAL ",
            "OBRA", " CC SUB"
        ]

        for ci, h in enumerate(rateio_headers, 1):
            c = ws_rat.cell(row=1, column=ci, value=h)
            c.font = HEADER_FONT; c.fill = LIGHT_BLUE_FILL; c.border = THIN_BORDER

        for ri, rd in enumerate(self.rateio_data, 2):
            should_highlight = rd.get('_highlight', False)

            c1 = ws_rat.cell(row=ri, column=1, value=rd['origem'])
            c1.border = THIN_BORDER
            c2 = ws_rat.cell(row=ri, column=2, value=rd['destino'])
            c2.border = THIN_BORDER
            c3 = ws_rat.cell(row=ri, column=3, value=rd['centro_custo'])
            c3.border = THIN_BORDER

            if should_highlight:
                c1.fill = ORANGE_FILL
                c2.fill = ORANGE_FILL
                c3.fill = ORANGE_FILL

            c = ws_rat.cell(row=ri, column=4, value=rd['qnt_chamados'])
            c.border = THIN_BORDER

            c = ws_rat.cell(row=ri, column=5,
                            value=rd['utilizacao'] if rd['utilizacao'] != '' else 0)
            c.number_format = 'R$ #,##0.00'; c.border = THIN_BORDER
            if rd['utilizacao']:
                c.fill = LIGHT_RED_FILL

            c = ws_rat.cell(row=ri, column=6,
                            value=rd['nao_localizado'] if rd['nao_localizado'] != '' else 0)
            c.number_format = 'R$ #,##0.00'; c.border = THIN_BORDER
            if rd['nao_localizado']:
                c.fill = LIGHT_RED_FILL

            c = ws_rat.cell(row=ri, column=7,
                            value=rd['vsc'] if rd['vsc'] != '' else 0)
            c.number_format = 'R$ #,##0.00'; c.border = THIN_BORDER

            c = ws_rat.cell(row=ri, column=8,
                            value=rd['vsc_nao_localizado'] if rd['vsc_nao_localizado'] else 0)
            c.border = THIN_BORDER

            c = ws_rat.cell(row=ri, column=9,
                            value=rd['desconto'] if rd['desconto'] != '' else 0)
            c.number_format = 'R$ #,##0.00'; c.border = THIN_BORDER

            c = ws_rat.cell(row=ri, column=10)
            c.value = f'=E{ri}+F{ri}+G{ri}+H{ri}-I{ri}'
            c.number_format = 'R$ #,##0.00'; c.border = THIN_BORDER

            ws_rat.cell(row=ri, column=11, value=rd['obra']).border = THIN_BORDER
            ws_rat.cell(row=ri, column=12, value=rd['cc_sub']).border = THIN_BORDER

        pivot_cc = defaultdict(float)

        for ri, rd in enumerate(self.rateio_data, 2):
            cc = rd.get('centro_custo', '') or '(vazio)'
            util_val = rd['utilizacao'] if isinstance(rd['utilizacao'], (int, float)) else 0
            nao_loc_val = rd['nao_localizado'] if isinstance(rd['nao_localizado'], (int, float)) else 0
            vsc_val = rd['vsc'] if isinstance(rd['vsc'], (int, float)) else 0
            desc_val = rd['desconto'] if isinstance(rd['desconto'], (int, float)) else 0
            total = util_val + nao_loc_val + vsc_val - desc_val
            pivot_cc[cc] += total

        pivot_start_col = 14
        pr = 1

        c = ws_rat.cell(row=pr, column=pivot_start_col, value="Centro de Custo")
        c.font = HEADER_FONT; c.fill = LIGHT_BLUE_FILL; c.border = THIN_BORDER
        c = ws_rat.cell(row=pr, column=pivot_start_col + 1, value="Soma de Total")
        c.font = HEADER_FONT; c.fill = LIGHT_BLUE_FILL; c.border = THIN_BORDER
        pr += 1

        grand_total = 0
        for cc in sorted(pivot_cc.keys()):
            val = pivot_cc[cc]
            grand_total += val
            c = ws_rat.cell(row=pr, column=pivot_start_col, value=cc)
            c.border = THIN_BORDER
            c = ws_rat.cell(row=pr, column=pivot_start_col + 1, value=round(val, 2))
            c.number_format = 'R$ #,##0.00'; c.border = THIN_BORDER
            c.alignment = Alignment(horizontal='right')
            pr += 1

        c = ws_rat.cell(row=pr, column=pivot_start_col, value="Total Geral")
        c.font = Font(bold=True); c.border = THIN_BORDER
        c = ws_rat.cell(row=pr, column=pivot_start_col + 1, value=round(grand_total, 2))
        c.font = Font(bold=True); c.number_format = 'R$ #,##0.00'
        c.border = THIN_BORDER; c.alignment = Alignment(horizontal='right')

        for ws in [ws_cor, ws_ag, ws_vsc, ws_rat]:
            for col_cells in ws.columns:
                mx = 0
                cl = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    try:
                        if cell.value:
                            mx = max(mx, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[cl].width = max(min(mx + 3, 55), 12)

        wb.save(self.output_path)

# ==============================================================================
# FUNÇÃO MESTRE (CHAMADA PELO HUB CENTRAL)
# ==============================================================================
def executar_rateio_malote():
    print("=" * 60)
    print("  RATEIO MALOTE v14")
    print("  Com resolução avançada de CC e Correção de Valor Órfão")
    print("=" * 60)

    if not PASTA_MALOTE.exists():
        PASTA_MALOTE.mkdir(parents=True, exist_ok=True)
        raise RuntimeError(f"A pasta '{PASTA_MALOTE}' não existia e foi criada.\nColoque os arquivos necessários lá dentro e tente novamente.")

    output_file = str(PASTA_MALOTE / "Rateio Malote.xlsx")

    # Removido o try/except que causava o erro duplicado (nested exception)
    rateio = RateioMalote(output_path=output_file)
    rateio.run()

if __name__ == "__main__":
    executar_rateio_malote()
