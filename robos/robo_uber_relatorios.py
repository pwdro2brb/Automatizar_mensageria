import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import unicodedata
import re
import os
import glob
import sys
from pathlib import Path
import config
# ==============================================================================
# CONFIGURAÇÃO DE PASTAS DINÂMICAS
# ==============================================================================
PASTA_UBER = Path(config.PASTA_ARQUIVOS) / "uber"

# Garante que a pasta exista
if not PASTA_UBER.exists():
    PASTA_UBER.mkdir(parents=True, exist_ok=True)

# =========================
# FUNÇÕES AUXILIARES
# =========================
def get_file_by_pattern(pattern):
    """Busca o arquivo mais recente na pasta do Uber que bata com o padrão."""
    files = list(PASTA_UBER.glob(pattern))
    if not files: return None
    return max(files, key=os.path.getctime)

def to_float(v):
    if v is None or v == "": return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace("R$", "").strip()
    if "," in s and "." in s: s = s.replace(".", "").replace(",", ".")
    elif "," in s: s = s.replace(",", ".")
    try: return float(s)
    except: return None

def sanitize_filename(name: str) -> str:
    name = str(name).strip()
    name = re.sub(r'[\\/:*?"<>|]+', '-', name)
    return re.sub(r'\s+', ' ', name).strip()[:150]

def sanitize_sheetname(name: str) -> str:
    name = str(name).strip()
    name = re.sub(r'[:\\/?*\\[\\]]+', '-', name)
    return re.sub(r'\s+', ' ', name).strip()[:31] if name else "Planilha1"

def find_col_by_prefix(df_columns, wanted: str):
    w = norm(wanted)
    for c in df_columns:
        if norm(c) == w: return c
    for c in df_columns:
        if norm(c).startswith(w): return c
    return None

def norm(s):
    if s is None or (isinstance(s, float) and pd.isna(s)): return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.upper().split())

def norm_email(s):
    if s is None or (isinstance(s, float) and pd.isna(s)): return ""
    s = str(s).strip().lower()
    s = re.sub(r"^mailto:", "", s)
    return s.replace("[", "").replace("]", "").strip()

# =====================================================================
# INTELIGÊNCIA DE COMPARAÇÃO DE NOMES E DESEMPATE DE CARGOS
# =====================================================================
def is_word_match(w1, w2):
    if w1 == w2: return True
    if len(w1) >= 3 and len(w2) >= 3 and (w1.startswith(w2) or w2.startswith(w1)): return True
    if len(w1) <= 2 and w2.startswith(w1): return True
    if len(w2) <= 2 and w1.startswith(w2): return True
    return False

def names_match(name1, name2):
    n1 = norm(name1).split()
    n2 = norm(name2).split()
    if not n1 or not n2: return False
    
    if not is_word_match(n1[0], n2[0]):
        return False
    
    short_parts, long_parts = (n1, n2) if len(n1) <= len(n2) else (n2, n1)
    
    matches = 0
    for sw in short_parts:
        for lw in long_parts:
            if is_word_match(sw, lw):
                matches += 1
                break
                
    return matches == len(short_parts)

def get_cargo_weight(cargo):
    c = norm(cargo)
    if any(x in c for x in ["DIRETOR", "GESTOR", "GERENTE", "COORDENADOR", "SUPERINTENDENTE", "BUSINESS PARTNER", "PRESIDENTE", "CONSELHEIRO"]):
        return 100
    if any(x in c for x in ["ESPECIALISTA", "ENGENHEIRO", "SUPERVISOR", "ANALISTA", "ARQUITETO", "COMPRADOR"]):
        return 50
    if any(x in c for x in ["ASSISTENTE", "AUXILIAR", "TECNICO"]):
        return 20
    return 0

# =====================================================================
# ETAPA 1: ATUALIZAR PLANILHA DE RESPONSÁVEIS COM O SAP
# =====================================================================
def etapa_1_atualizar_responsaveis():
    print("[PROGRESSO: 5]")
    print("="*60)
    print(" UBER ETAPA 1: ATUALIZAR RESPONSÁVEIS ".center(60))
    print("="*60)
    
    # Validação de arquivos
    ativos_path = get_file_by_pattern("Base de Ativos*.xlsx")
    resp_path = PASTA_UBER / "Responsaveis Por Centro de Custos.xlsx"
    
    erros = []
    if not ativos_path: erros.append("👉 Base de Ativos (ex: 'Base de Ativos 16.06.2026.xlsx')")
    if not resp_path.exists(): erros.append("👉 Responsaveis Por Centro de Custos.xlsx")
    
    # Busca o arquivo SAP diretamente na pasta do Uber (muito mais seguro!)
    sap_files = [os.path.join(PASTA_UBER, f) for f in os.listdir(PASTA_UBER) 
                 if 'export' in f.lower() and f.lower().endswith(('.xls', '.xlsx'))]
                
    if not sap_files:
        erros.append("👉 Arquivo SAP (com 'export' no nome) na pasta do Uber")

        
    if erros:
        msg = "Faltam arquivos obrigatórios para a Etapa 1:\n\n" + "\n".join(erros) + f"\n\nColoque os arquivos na pasta:\n{PASTA_UBER}"
        raise FileNotFoundError(msg)

    sap_file = max(sap_files, key=os.path.getctime)
    print(f"✅ Arquivo SAP encontrado: {os.path.basename(sap_file)}")

    print("[PROGRESSO: 20]")
    # =================================================================
    # LEITURA BLINDADA DO SAP
    # =================================================================
    sap_df_raw = pd.read_excel(sap_file, header=None)
    header_idx = 0
    
    for i, row in sap_df_raw.iterrows():
        linha_texto = " ".join([norm(str(x)) for x in row.values])
        if "RESPONSAVEL" in linha_texto and ("CENTRO CUSTO" in linha_texto or "CENTRO DE CUSTO" in linha_texto):
            header_idx = i
            break
    
    print(f"-> Cabeçalho do SAP encontrado na linha: {header_idx}")
    sap_df = pd.read_excel(sap_file, header=header_idx)
    
    # Força a limpeza de TODOS os nomes de colunas do SAP
    sap_df.columns = [norm(str(c)) for c in sap_df.columns]
    print(f"-> Colunas limpas lidas do SAP: {list(sap_df.columns)}")
    
    mask_inativo = pd.Series(False, index=sap_df.index)
    for col in sap_df.columns:
        if sap_df[col].dtype == object:
            mask_inativo = mask_inativo | sap_df[col].astype(str).str.upper().str.contains(r'\bINATIVO|\bINAT\b|\bINAT-', regex=True)
    sap_df = sap_df[~mask_inativo]

    # Agora procura nas colunas já limpas
    col_cc_sap = next((c for c in sap_df.columns if "CENTRO CUSTO" in c or "CENTRO DE CUSTO" in c), None)
    col_resp_sap = next((c for c in sap_df.columns if "RESPONSAVEL" in c), None)
    
    if not col_cc_sap or not col_resp_sap:
        raise RuntimeError(f"⚠️ Colunas não encontradas! O robô enxergou estas colunas: {list(sap_df.columns)}")
    # =================================================================

    sap_dict = dict(zip(sap_df[col_cc_sap].astype(str).str.strip(), sap_df[col_resp_sap].astype(str).str.strip()))

    print("[PROGRESSO: 40]")
    wb_ativos = load_workbook(ativos_path, read_only=True, data_only=True)
    ws_ativos = wb_ativos.worksheets[0]
    
    header_row_idx = 7
    header_vals = []
    for r in range(1, 60):
        row_vals = [ws_ativos.cell(r, c).value for c in range(1, 40)]
        row_norm = [norm(v) for v in row_vals]
        if "NOME FUNCIONARIO" in row_norm or "NOME FUNCIONÁRIO" in row_norm or "NOME" in row_norm:
            header_row_idx = r
            header_vals = row_norm
            break
            
    def find_col_ativos(possible):
        for p in possible:
            p = norm(p)
            if p in header_vals: return header_vals.index(p) + 1
        return None

    col_nome_func = find_col_ativos(["Nome Funcionário", "Nome Funcionario", "Nome"]) or 2
    col_email_func = find_col_ativos(["E-mail", "Email", "E Mail"]) or 5
    col_cargo_func = find_col_ativos(["Nome da Função", "Nome da Funcao", "Função", "Funcao", "Cargo"]) or 7

    ativos_list = []
    for row in ws_ativos.iter_rows(min_row=header_row_idx + 1, values_only=True):
        nome = row[col_nome_func - 1] if col_nome_func - 1 < len(row) else None
        if nome:
            ativos_list.append({
                'nome': str(nome).strip(),
                'email': str(row[col_email_func - 1]).strip() if col_email_func - 1 < len(row) else "",
                'cargo': str(row[col_cargo_func - 1]).strip() if col_cargo_func - 1 < len(row) else ""
            })
    wb_ativos.close()

    print("[PROGRESSO: 60]")
    wb_resp = load_workbook(resp_path)
    ws_resp = wb_resp.active
    
    cc_col_idx = 1
    for c in range(1, ws_resp.max_column + 1):
        val = norm(ws_resp.cell(1, c).value)
        if "CODIGO" in val or "CC" in val or "CENTRO" in val:
            cc_col_idx = c
            break
            
    status_col_idx = ws_resp.max_column + 2
    ws_resp.cell(1, status_col_idx).value = "Status Atualização SAP"

    print("[PROGRESSO: 80]")
    for r in range(2, ws_resp.max_row + 1):
        cc_val = str(ws_resp.cell(r, cc_col_idx).value).strip()
        if not cc_val or cc_val == "None": continue

        if cc_val not in sap_dict:
            ws_resp.cell(r, status_col_idx).value = "Marcado: Não consta no SAP (Diagrama/Seq)"
            continue

        sap_resp_name = sap_dict[cc_val]
        current_resp_name = str(ws_resp.cell(r, cc_col_idx + 1).value).strip()

        if names_match(sap_resp_name, current_resp_name):
            ws_resp.cell(r, status_col_idx).value = "OK"
            continue
            
        found_ativos = []
        for ativo in ativos_list:
            if names_match(sap_resp_name, ativo['nome']):
                found_ativos.append(ativo)
                
        if found_ativos:
            found_ativos.sort(key=lambda x: get_cargo_weight(x['cargo']), reverse=True)
            found_ativo = found_ativos[0] 
            
            ws_resp.cell(r, cc_col_idx + 1).value = found_ativo['nome']
            ws_resp.cell(r, cc_col_idx + 2).value = found_ativo['email']
            ws_resp.cell(r, cc_col_idx + 3).value = found_ativo['cargo']
            ws_resp.cell(r, status_col_idx).value = "Alterado"
        else:
            ws_resp.cell(r, status_col_idx).value = f"Marcado: Nome SAP ({sap_resp_name}) não achado na Base de Ativos"

    novo_resp_path = PASTA_UBER / "Responsaveis_Atualizado_SAP.xlsx"
    wb_resp.save(novo_resp_path)
    print("[PROGRESSO: 100]")
    print(f"✅ Planilha de responsáveis atualizada salva como: {novo_resp_path.name}")
    print("\n⚠️ PROCESSO PAUSADO ⚠️")
    print("Abra o arquivo 'Responsaveis_Atualizado_SAP.xlsx', verifique os casos marcados manualmente, corrija o que for necessário e salve.")
    print("Depois, clique no botão 'Uber 2: Gerar Relatórios e Pastas'.")


# =====================================================================
# ETAPA 2: GERAR CONSOLIDADO E PASTAS
# =====================================================================
def etapa_2_gerar_relatorios():
    print("[PROGRESSO: 5]")
    print("="*60)
    print(" UBER ETAPA 2: GERAR RELATÓRIOS E PASTAS ".center(60))
    print("="*60)
    
    # Validação de arquivos
    uber_path = get_file_by_pattern("Relatório*.xlsx")
    ativos_path = get_file_by_pattern("Base de Ativos*.xlsx")
    resp_path_atualizado = PASTA_UBER / "Responsaveis_Atualizado_SAP.xlsx"
    resp_path_original = PASTA_UBER / "Responsaveis Por Centro de Custos.xlsx"
    
    erros = []
    if not uber_path: erros.append("👉 Relatório do Uber (ex: 'Relatório Maio - 2026.xlsx')")
    if not ativos_path: erros.append("👉 Base de Ativos (ex: 'Base de Ativos 16.06.2026.xlsx')")
    
    if resp_path_atualizado.exists():
        resp_path = resp_path_atualizado
        print(f"✅ Usando planilha de responsáveis revisada: {resp_path.name}")
    elif resp_path_original.exists():
        resp_path = resp_path_original
        print(f"⚠️ Planilha atualizada não encontrada. Usando a original: {resp_path.name}")
    else:
        erros.append("👉 Responsaveis Por Centro de Custos.xlsx")
        
    if erros:
        msg = "Faltam arquivos obrigatórios para a Etapa 2:\n\n" + "\n".join(erros) + f"\n\nColoque os arquivos na pasta:\n{PASTA_UBER}"
        raise FileNotFoundError(msg)

    CARGO_NAO_ENCONTRADO = "cargo não encontrado"
    MSG_CC_MISSING = "CC NÃO ESTÁ MAIS NA BASE UBER"

    BANIDOS_NOMES = {
        "JUNIA GALVAO", "THIAGO CORREA ELY", "RAFAEL PIRES E ALBUQUERQUE",
        "RONALDO PEDREIRA AYRES DA MOTTA FILHO", "RICARDO PAIXAO PINTO RODRIGUES",
        "RAPHAEL ROCHA LAFETA", "EDUARDO FISCHER TEIXEIRA DE SOUZA",
        "RODRIGO MARTINS DE RESENDE", "RAFAEL NAZARETH MENIN TEIXEIRA DE SOUZA",
        "RUBENS MENIN TEIXEIRA DE SOUZA"
    }
    CARGOS_PROIBIDOS_CONTEM = ["DIRETOR EXECUTIVO", "PRESIDENTE", "CONSELHEIRO"]

    def is_excluded(name, cargo):
        n, c = norm(name), norm(cargo)
        if n in {norm(x) for x in BANIDOS_NOMES}: return True
        for sub in [norm(x) for x in CARGOS_PROIBIDOS_CONTEM]:
            if sub in c: return True
        return False

    def fmt_date_as_text(x):
        if pd.isna(x): return ""
        dt = pd.to_datetime(x, errors="coerce")
        if pd.isna(dt): return str(x).strip()
        return dt.strftime("%d/%m/%Y")

    def fmt_time_as_text(x):
        if pd.isna(x): return ""
        dt = pd.to_datetime(x, errors="coerce")
        if pd.isna(dt): return str(x).strip()
        return dt.strftime("%I:%M%p").lstrip("0")

    print("[PROGRESSO: 15]")
    # =========================
    # 1) RESPONSÁVEIS
    # =========================
    resp_df = pd.read_excel(resp_path, engine="openpyxl")
    resp_df.columns = [str(c).strip() for c in resp_df.columns]

    col_cc = next((c for c in resp_df.columns if norm(c) in {
        "CODIGODEDESPESAS", "CODIGO DE DESPESAS", "CÓDIGO DE DESPESAS",
        "CODIGO DA DESPESA", "CÓDIGO DA DESPESA"
    }), resp_df.columns[0])

    col_resp = next((c for c in resp_df.columns if norm(c) in {"RESPONSAVEL", "RESPONSÁVEL"}), None)
    if col_resp is None: col_resp = next((c for c in resp_df.columns if "RESP" in norm(c)), None)

    col_email = next((c for c in resp_df.columns if "MAIL" in norm(c)), None)
    col_cargo_resp = next((c for c in resp_df.columns if "CARGO" in norm(c)), None)
    if col_cargo_resp is None and len(resp_df.columns) >= 4: col_cargo_resp = resp_df.columns[3]

    resp_df = resp_df.dropna(subset=[col_cc]).copy()
    resp_df["_cc_norm"] = resp_df[col_cc].astype(str).map(norm)
    resp_df["_resp_norm"] = resp_df[col_resp].astype(str).map(norm)
    resp_df["_email_norm"] = resp_df[col_email].map(norm_email)

    cc_to_resp = dict(zip(resp_df["_cc_norm"], resp_df[col_resp].astype(str)))
    cc_to_email = dict(zip(resp_df["_cc_norm"], resp_df[col_email].astype(str)))
    resp_name_to_cargo = resp_df.groupby("_resp_norm")[col_cargo_resp].first().to_dict()
    resp_name_to_email_norm = resp_df.groupby("_resp_norm")["_email_norm"].first().to_dict()

    print("[PROGRESSO: 25]")
    # =========================
    # 2) BASE ATIVOS
    # =========================
    wb = load_workbook(ativos_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    header_row_idx = None
    for r in range(1, 60):
        row = [ws.cell(r, c).value for c in range(1, 40)]
        if "NOME FUNCIONARIO" in [norm(v) for v in row] or "NOME FUNCIONÁRIO" in [norm(v) for v in row]:
            header_row_idx = r
            header_vals = row
            break

    if header_row_idx is None:
        header_row_idx = 7
        header_vals = [ws.cell(header_row_idx, c).value for c in range(1, 40)]

    header_norm = [norm(v) for v in header_vals]

    def find_col(possible):
        for p in possible:
            p = norm(p)
            if p in header_norm: return header_norm.index(p) + 1
        return None

    col_nome_func = find_col(["Nome Funcionário", "Nome Funcionario", "Nome"]) or 2
    col_email_func = find_col(["E-mail", "Email", "E Mail"]) or 5
    col_cargo_func = find_col(["Nome da Função", "Nome da Funcao", "Função", "Funcao", "Cargo"]) or 7
    col_diretor_func = find_col(["Nome Diretor", "Diretor"]) 

    email_to_cargo = {}
    name_to_cargo = {}
    colaborador_to_diretor = {} 

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        nome = row[col_nome_func - 1] if col_nome_func - 1 < len(row) else None
        email = row[col_email_func - 1] if col_email_func - 1 < len(row) else None
        cargo = row[col_cargo_func - 1] if col_cargo_func - 1 < len(row) else None
        diretor = row[col_diretor_func - 1] if col_diretor_func and col_diretor_func - 1 < len(row) else None
        
        cargo_str = "" if cargo is None else str(cargo).strip()

        if email:
            e = norm_email(email)
            if e and (e not in email_to_cargo or (not email_to_cargo[e] and cargo_str)):
                email_to_cargo[e] = cargo_str

        if nome:
            n = norm(nome)
            if n and (n not in name_to_cargo or (not name_to_cargo[n] and cargo_str)):
                name_to_cargo[n] = cargo_str
            if n and diretor:
                colaborador_to_diretor[n] = str(diretor).strip()

    wb.close()

    print("[PROGRESSO: 35]")
    # =========================
    # 3) UBER (consolidado)
    # =========================
    uber_df = pd.read_excel(uber_path, engine="openpyxl")
    uber_df.columns = [str(c).strip() for c in uber_df.columns]

    unnamed_cols = [c for c in uber_df.columns if str(c).startswith("Unnamed")]
    if unnamed_cols: uber_df = uber_df.drop(columns=unnamed_cols)

    drop_col = next((c for c in uber_df.columns if norm(c) == norm("Registro de data e hora da transação (UTC)")), None)
    if drop_col: uber_df = uber_df.drop(columns=[drop_col])

    col_cc_uber = next((c for c in uber_df.columns if norm(c) in {
        "CODIGO DA DESPESA", "CÓDIGO DA DESPESA", "CODIGO DE DESPESA", "CÓDIGO DE DESPESA"
    }), None)
    col_email_uber = next((c for c in uber_df.columns if "MAIL" in norm(c)), None)
    col_nome = next((c for c in uber_df.columns if norm(c) == "NOME"), None)
    col_sobrenome = next((c for c in uber_df.columns if norm(c) == "SOBRENOME"), None)

    for c in uber_df.columns:
        if norm(c).startswith("DATA "): uber_df[c] = uber_df[c].map(fmt_date_as_text)
        if norm(c).startswith("HORA "): uber_df[c] = uber_df[c].map(fmt_time_as_text)

    uber_df["_cc_norm"] = uber_df[col_cc_uber].astype(str).map(norm)
    uber_df["RESPONSÁVEL CC"] = uber_df["_cc_norm"].map(lambda x: cc_to_resp.get(x, MSG_CC_MISSING))

    def cargo_responsavel(resp_name, cc_norm):
        rn = norm(resp_name)
        if resp_name == MSG_CC_MISSING or rn == norm(MSG_CC_MISSING): return ""
        c = resp_name_to_cargo.get(rn, "")
        if c: return c
        email_resp = cc_to_email.get(cc_norm, "")
        c2 = email_to_cargo.get(norm_email(email_resp), "")
        if c2: return c2
        return name_to_cargo.get(rn, "")

    uber_df["CARGO RESPONSÁVEL CC"] = [
        cargo_responsavel(r, cc) for r, cc in zip(uber_df["RESPONSÁVEL CC"], uber_df["_cc_norm"])
    ]

    uber_df["COLABORADOR AJUSTADO"] = (
        uber_df[col_nome].fillna("").astype(str).str.strip() + " " +
        uber_df[col_sobrenome].fillna("").astype(str).str.strip()
    ).str.replace(r"\s+", " ", regex=True).str.strip()

    for i in uber_df.index:
        resp_atual = uber_df.at[i, "RESPONSÁVEL CC"]
        if resp_atual == MSG_CC_MISSING or pd.isna(resp_atual) or str(resp_atual).strip() == "":
            colab = uber_df.at[i, "COLABORADOR AJUSTADO"]
            diretor = colaborador_to_diretor.get(norm(colab))
            
            if diretor:
                uber_df.at[i, "RESPONSÁVEL CC"] = diretor.title()
                cargo_dir = name_to_cargo.get(norm(diretor), CARGO_NAO_ENCONTRADO)
                uber_df.at[i, "CARGO RESPONSÁVEL CC"] = cargo_dir

    def cargo_colaborador(email, nome):
        c = email_to_cargo.get(norm_email(email), "")
        if c: return c
        c2 = name_to_cargo.get(norm(nome), "")
        if c2: return c2
        return CARGO_NAO_ENCONTRADO

    uber_df["CARGO COLABORADOR"] = [
        cargo_colaborador(e, n) for e, n in zip(uber_df[col_email_uber], uber_df["COLABORADOR AJUSTADO"])
    ]

    uber_df["OBSERVAÇÃO"] = ""

    anchor = next((c for c in uber_df.columns if norm(c) == norm("Valor da transação em BRL (com tributos)")), None)
    new_cols = ["RESPONSÁVEL CC", "CARGO RESPONSÁVEL CC", "COLABORADOR AJUSTADO", "CARGO COLABORADOR", "OBSERVAÇÃO"]
    cols = [c for c in uber_df.columns if c not in new_cols]

    if anchor and anchor in cols:
        idx = cols.index(anchor) + 1
        cols = cols[:idx] + new_cols + cols[idx:]
    else:
        cols += new_cols

    uber_out_df = uber_df[cols].drop(columns=["_cc_norm"], errors="ignore")

    print("[PROGRESSO: 45]")
    # =========================
    # 4) SALVA CONSOLIDADO + APLICA ESTILOS
    # =========================
    out_consolidado = PASTA_UBER / "consolidado_para_envio_ATUALIZADO.xlsx"
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "consolidado para envio"

    for row in dataframe_to_rows(uber_out_df, index=False, header=True): ws_out.append(row)

    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = ws_out.dimensions

    fill_azul   = PatternFill("solid", fgColor="B7DEE8")
    fill_laranja= PatternFill("solid", fgColor="FFC000")
    fill_rosa   = PatternFill("solid", fgColor="FCE4D6")
    bold_font = Font(bold=True, color="000000")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="A6A6A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header = [cell.value for cell in ws_out[1]]
    col_index = {str(h).strip(): i + 1 for i, h in enumerate(header) if h is not None}

    def find_col_idx(col_name: str):
        target = norm(col_name)
        for h, idx in col_index.items():
            if norm(h) == target or norm(h).startswith(target): return idx
        return None

    def style_header(col_name, fill):
        idx = find_col_idx(col_name)
        if idx is None: return
        c = ws_out.cell(row=1, column=idx)
        c.fill = fill
        c.font = bold_font
        c.alignment = center
        c.border = border

    AZUL_HEADERS = [
        "Data da solicitação (UTC)", "Hora da solicitação (UTC)", "Data da solicitação (local)",
        "Hora da solicitação (local)", "Data de chegada (UTC)", "Hora de chegada (UTC)",
        "Data de chegada (local)", "Hora de chegada (local)", "Compensação do fuso horário de solicitação a partir do UTC",
        "Nome", "Sobrenome", "E-mail", "ID do funcionário", "Serviço", "Cidade", "Distância (mi)",
        "Duração (min)", "Endereço de partida", "Endereço de destino", "Detalhamento da despesa",
        "Faturas", "Programa", "Grupo", "Forma de pagamento", "Valor na moeda local (sem tributos)",
        "Tributos na moeda local", "Valor extra em moeda local", "Valor da transação na moeda local (com tributos)",
        "Código da moeda local", "Valor em BRL (sem tributos)", "Tributos em BRL", "Valor extra em BRL",
    ]

    for col_name in AZUL_HEADERS: style_header(col_name, fill_azul)
    for col_name in ["Código da despesa", "Tipo de transação", "Valor da transação em BRL (com tributos)"]: style_header(col_name, fill_rosa)
    for col_name in ["RESPONSÁVEL CC", "CARGO RESPONSÁVEL CC", "COLABORADOR AJUSTADO", "CARGO COLABORADOR", "OBSERVAÇÃO"]: style_header(col_name, fill_laranja)

    ws_out.row_dimensions[1].height = 22
    wb_out.save(out_consolidado)
    print("OK! Arquivo gerado:", out_consolidado.name)

    print("[PROGRESSO: 55]")
    # =========================
    # 5) TESTE MACRO + ENVIAR EMAIL
    # =========================
    used_resp = uber_out_df["RESPONSÁVEL CC"].fillna("").astype(str).str.strip()
    used_resp = used_resp[(used_resp != "") & (used_resp != MSG_CC_MISSING)]
    used_norm = set(used_resp.map(norm))

    macro_rows = []
    for rn in sorted(used_norm):
        orig = resp_df.loc[resp_df["_resp_norm"] == rn, col_resp].dropna()
        orig_name = orig.iloc[0] if len(orig) > 0 else rn.title()

        email = resp_name_to_email_norm.get(rn, "")
        cargo = resp_name_to_cargo.get(rn, "")
        if not cargo and email: cargo = email_to_cargo.get(email, "")
        if not cargo: cargo = name_to_cargo.get(rn, "")

        if is_excluded(orig_name, cargo) or not email: continue
        macro_rows.append({"Responsavel": orig_name, "E-mail": email})

    macro_unique = pd.DataFrame(macro_rows)
    if not macro_unique.empty:
        macro_unique["_email_norm"] = macro_unique["E-mail"].map(norm_email)
        macro_unique = macro_unique.drop_duplicates("_email_norm", keep="first").drop(columns=["_email_norm"])

    now = datetime.now()
    prev_month = now.month - 1
    prev_year = now.year
    if prev_month == 0:
        prev_month = 12
        prev_year -= 1

    macro_filename = PASTA_UBER / f"TESTE MACRO {prev_year},{prev_month:02d} COM_ESTILO.xlsx"
    email_filename = PASTA_UBER / "Enviar_e-mail original COM_ESTILO.xlsx"

    orange = PatternFill("solid", fgColor="FFA500")
    header_font = Font(bold=True, color="FFFFFF")

    wb_m = Workbook()
    ws_m = wb_m.active
    ws_m.title = "TESTE MACRO"
    ws_m["A1"], ws_m["B1"] = "Responsavel", "E-mail"
    for cell in ("A1", "B1"):
        ws_m[cell].fill = orange
        ws_m[cell].font = header_font

    for i, r in enumerate(macro_unique.itertuples(index=False), start=2):
        ws_m.cell(i, 1).value = r[0]
        ws_m.cell(i, 2).value = r[1]
    wb_m.save(macro_filename)

    wb_e = Workbook()
    ws_e = wb_e.active
    ws_e.title = "Enviar_e-mail original"
    ws_e["A1"], ws_e["B1"], ws_e["C1"] = "E-mail", "Responsavel", "Corpo"
    for cell in ("A1", "B1", "C1"):
        ws_e[cell].fill = orange
        ws_e[cell].font = header_font

    body_text = (
        f"Segue em anexo as utilizações do Uber coorporativo referente ao mês de {now.strftime('%B')} de {prev_year}, "
        "solicitados nos centros de custos sob sua responsabilidade. Caso estiver de acordo, não é necessário responder esse e-mail."
    )

    for i, r in enumerate(macro_unique.itertuples(index=False), start=2):
        ws_e.cell(i, 1).value = r[1]
        ws_e.cell(i, 2).value = r[0]
        ws_e.cell(i, 3).value = body_text
    wb_e.save(email_filename)

    print("OK! Arquivos gerados:", macro_filename.name, "e", email_filename.name)

    # =========================
    # 6) CRIAR PLANILHAS POR RESPONSÁVEL
    # =========================
    def criar_planilhas_por_responsavel(consolidado_path, teste_macro_path):
        # Cria a pasta com o mês passado automaticamente
        nome_pasta_mes = f"{prev_year},{prev_month:02d}"
        pasta_mes_path = PASTA_UBER / nome_pasta_mes
        pasta_mes_path.mkdir(parents=True, exist_ok=True)

        macro_df = pd.read_excel(teste_macro_path, engine="openpyxl")
        macro_df.columns = [str(c).strip() for c in macro_df.columns]
        col_resp_macro = macro_df.columns[0]
        responsaveis = macro_df[col_resp_macro].dropna().astype(str).map(lambda x: x.strip()).loc[lambda s: s != ""].unique().tolist()

        cons_df = pd.read_excel(consolidado_path, engine="openpyxl")
        cons_df.columns = [str(c).strip() for c in cons_df.columns]

        col_end_dest = find_col_by_prefix(cons_df.columns, "Endereço de destino")
        col_nome_aj = find_col_by_prefix(cons_df.columns, "COLABORADOR AJUSTADO")
        col_resp_cc = find_col_by_prefix(cons_df.columns, "RESPONSÁVEL CC")

        colunas_saida = [
            "Data da solicitação (local)", "Hora da solicitação (local)", "Data de chegada (local)",
            "Hora de chegada (local)", "Nome", "Sobrenome", "Nome ajustado", "E-mail", "Cidade",
            "Distância (mi)", "Duração (min)", "Endereço de partida", "Endereço de destino",
            "Código da despesa", "Responsavel", "Detalhamento da despesa", "Valor da transação em BRL (com tributos)",
        ]

        roxo = PatternFill("solid", fgColor="7030A0")
        laranja = PatternFill("solid", fgColor="ED7D31")
        font_white_bold = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="A6A6A6")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        total_resp = len(responsaveis)
        for i, resp in enumerate(responsaveis):
            resp_norm = norm(resp)
            if col_resp_cc is None: continue

            f = cons_df[col_resp_cc].fillna("").astype(str).map(norm) == resp_norm
            df_r = cons_df.loc[f].copy()
            if df_r.empty: continue

            saida = pd.DataFrame()
            for c in colunas_saida:
                if c == "Endereço de destino": saida[c] = df_r[col_end_dest] if col_end_dest else ""
                elif c == "Nome ajustado": saida[c] = df_r[col_nome_aj] if col_nome_aj else ""
                elif c == "Responsavel": saida[c] = df_r[col_resp_cc] if col_resp_cc else ""
                elif c in df_r.columns: saida[c] = df_r[c]
                else: saida[c] = ""

            saida = saida[colunas_saida]
            wb = Workbook()
            ws = wb.active
            ws.title = sanitize_sheetname(resp)

            for row in dataframe_to_rows(saida, index=False, header=True): ws.append(row)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            header_map = {ws.cell(1, j).value: j for j in range(1, ws.max_column + 1)}
            for col_name, idx in header_map.items():
                cell = ws.cell(1, idx)
                cell.fill = roxo
                cell.font = font_white_bold
                cell.alignment = center
                cell.border = border

            for special in ["Nome ajustado", "Responsavel"]:
                if special in header_map:
                    cell = ws.cell(1, header_map[special])
                    cell.fill = laranja

            max_row, max_col = ws.max_row, ws.max_column
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(r, c)
                    if cell.value not in (None, ""): cell.border = border

            cols_centralizar = list(range(1, 6)) + list(range(9, 12)) + [14, 17]
            for r in range(2, max_row + 1):
                for c in cols_centralizar:
                    if c <= max_col:
                        cell = ws.cell(r, c)
                        if cell.value not in (None, ""): cell.alignment = center

            ws.row_dimensions[1].height = 50
            for r in range(2, ws.max_row + 1): ws.row_dimensions[r].height = 15

            BRL_FORMAT = '"R$" #,##0.00'
            col_valor_brl = header_map.get("Valor da transação em BRL (com tributos)")
            if col_valor_brl:
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(r, col_valor_brl)
                    val = to_float(cell.value)
                    if val is not None:
                        cell.value = val
                        cell.number_format = BRL_FORMAT

            file_name = sanitize_filename(resp) + ".xlsx"
            wb.save(pasta_mes_path / file_name)
            
            # Progresso dinâmico de 60% a 95%
            progresso_atual = 60 + int(((i + 1) / total_resp) * 35)
            print(f"[PROGRESSO: {progresso_atual}]")

        print(f"✅ Pasta criada: {pasta_mes_path.name} | Arquivos por responsável gerados.")

    criar_planilhas_por_responsavel(consolidado_path=out_consolidado, teste_macro_path=macro_filename)

    # =========================
    # 7) GERAR PENDENCIAS_CARGO.xlsx
    # =========================
    PEND_FILE = PASTA_UBER / "PENDENCIAS_CARGO.xlsx"
    cols_needed = ["COLABORADOR AJUSTADO", "E-mail", "RESPONSÁVEL CC", "Código da despesa", "Cidade", "Serviço", "CARGO COLABORADOR", "Detalhamento da despesa", "Valor da transação em BRL (com tributos)"]
    for c in cols_needed:
        if c not in uber_out_df.columns: uber_out_df[c] = ""

    pend_df = uber_out_df.loc[uber_out_df["CARGO COLABORADOR"].fillna("").astype(str).str.strip().str.lower() == "cargo não encontrado"].copy()
    pend_out_cols = ["COLABORADOR AJUSTADO", "E-mail", "Cidade", "Serviço", "Código da despesa", "RESPONSÁVEL CC", "CARGO COLABORADOR"]
    pend_df = pend_df[pend_out_cols]

    resumo = pend_df.groupby(["COLABORADOR AJUSTADO", "E-mail"], dropna=False).size().reset_index(name="Qtde ocorrências").sort_values("Qtde ocorrências", ascending=False)

    with pd.ExcelWriter(PEND_FILE, engine="openpyxl") as writer:
        pend_df.to_excel(writer, index=False, sheet_name="Pendencias (linhas)")
        resumo.to_excel(writer, index=False, sheet_name="Resumo por pessoa")

    print("[PROGRESSO: 100]")
    print(f"✅ Arquivo gerado: {PEND_FILE.name}  | Linhas: {len(pend_df)} | Pessoas: {len(resumo)}")
